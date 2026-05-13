from datetime import datetime, timezone
import io
import logging
import re
import uuid
from typing import Optional

from fastapi import APIRouter, Body, Depends, File, HTTPException, Query, Response, UploadFile

from models import (
    COAImportCommitRequest,
    COAManualInput,
    COASettingsUpdate,
    DisputeAttachmentInput,
    DisputeCloseInput,
    DisputeNoteInput,
    UmpireProposal,
    UmpireResultInput,
)
from services.coa_reconciliation import (
    apply_preserved_coa_fields,
    build_combined_coa_import_preview,
    calculate_kpis,
    coa_records_differ,
    get_gcv_trend_data,
    get_radar_chart_data,
    merge_coa_data,
    normalize_coa_shipment,
    parse_combined_coa_workbook,
    parse_coa_excel,
)
from services.data_quality import summarize_coa_preview_quality
from utils.auth import get_current_user, require_role
from utils.database import db

router = APIRouter(tags=["COA"])
logger = logging.getLogger(__name__)


def _workflow_event(action: str, status: str, user: dict, notes: Optional[str] = None) -> dict:
    return {
        "id": str(uuid.uuid4()),
        "action": action,
        "status": status,
        "notes": notes,
        "actor_id": user.get("id"),
        "actor_name": user.get("name"),
        "created_at": datetime.now(timezone.utc).isoformat(),
    }


def _workflow_summary(record: dict) -> dict:
    history = record.get("dispute_history") or []
    notes = record.get("dispute_notes") or []
    attachments = record.get("dispute_attachments") or []
    opened_at = record.get("umpire_proposed_at") or record.get("created_at")
    closed_at = record.get("dispute_closed_at") or record.get("umpire_completed_at")
    return {
        "status": record.get("umpire_status", "none"),
        "opened_at": opened_at,
        "closed_at": closed_at,
        "history": history,
        "notes": notes,
        "attachments": attachments,
        "resolution": record.get("dispute_resolution"),
        "closure_notes": record.get("dispute_closure_notes"),
        "history_count": len(history),
        "note_count": len(notes),
        "attachment_count": len(attachments),
    }


def _existing_by_shipment(records: list[dict]) -> dict[str, dict]:
    by_key = {}
    for record in records:
        key = normalize_coa_shipment(record.get("shipment"))
        if key and key not in by_key:
            by_key[key] = record
    return by_key


def _has_critical_import_issues(preview: dict) -> bool:
    return any(issue.get("severity") == "critical" for issue in preview.get("issues") or [])


def _safe_regex(value: str) -> dict:
    return {"$regex": re.escape(str(value).strip()), "$options": "i"}


def _coa_filter_query(
    *,
    status: Optional[str] = None,
    search: Optional[str] = None,
    supplier: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    include_disputes: bool = False,
    umpire_status: Optional[str] = None,
) -> dict:
    conditions = []
    if include_disputes:
        conditions.append({"umpire_status": {"$ne": "none"}})
    if status and status != "all":
        conditions.append({"status": status})
    if umpire_status and umpire_status != "all":
        conditions.append({"umpire_status": umpire_status})
    if supplier and supplier != "all":
        conditions.append({"suppliers": _safe_regex(supplier)})
    if search:
        safe_search = re.escape(search)
        conditions.append({
            "$or": [
                {"shipment": {"$regex": safe_search, "$options": "i"}},
                {"suppliers": {"$regex": safe_search, "$options": "i"}}
            ]
        })
    if date_from or date_to:
        date_filter = {}
        if date_from:
            date_filter["$gte"] = date_from
        if date_to:
            date_filter["$lte"] = date_to + "T23:59:59"
        conditions.append({"completed_unloading": date_filter})

    if not conditions:
        return {}
    return {"$and": conditions} if len(conditions) > 1 else conditions[0]


def _public_preview_response(preview: dict) -> dict:
    return {
        "preview_id": preview["id"],
        "dataset": preview["dataset"],
        "filename": preview.get("filename"),
        "row_count": len(preview.get("records") or []),
        "source_counts": preview.get("source_counts") or {},
        "coverage": preview.get("coverage") or {},
        "validation_summary": preview.get("validation_summary") or {},
        "issue_count": len(preview.get("issues") or []),
        "issues": (preview.get("issues") or [])[:100],
        "diff_summary": preview.get("diff_summary") or {},
        "preservation_summary": preview.get("preservation_summary") or {},
        "data_quality": preview.get("data_quality") or summarize_coa_preview_quality(preview),
        "preview_rows": (preview.get("records") or [])[:10],
        "allowed_modes": ["merge", "replace"],
        "status": preview.get("status"),
        "created_at": preview.get("created_at"),
        "created_by": preview.get("created_by"),
    }


async def _commit_combined_coa_preview(preview: dict, mode: str, confirm_replace_all: bool, user: dict) -> dict:
    if mode not in {"merge", "replace"}:
        raise HTTPException(status_code=400, detail="Mode import COA harus merge atau replace")
    if mode == "replace" and not confirm_replace_all:
        raise HTTPException(status_code=400, detail="Replace-all wajib dikonfirmasi eksplisit")
    if _has_critical_import_issues(preview):
        raise HTTPException(status_code=400, detail="Preview masih memiliki isu critical. Perbaiki workbook sebelum commit.")

    records = preview.get("records") or []
    existing_records = await db.coa_reconciliation.find({}, {"_id": 0}).to_list(50000)
    existing_by_key = _existing_by_shipment(existing_records)
    now = datetime.now(timezone.utc).isoformat()

    prepared_records = []
    for record in records:
        key = normalize_coa_shipment(record.get("shipment"))
        prepared = apply_preserved_coa_fields(record, existing_by_key.get(key))
        prepared.update({
            "uploaded_by": user["id"],
            "uploaded_at": now,
            "updated_at": now,
            "import_source": preview.get("filename"),
            "import_preview_id": preview["id"],
        })
        prepared_records.append(prepared)

    snapshot_id = str(uuid.uuid4())
    await db.coa_import_snapshots.insert_one({
        "id": snapshot_id,
        "preview_id": preview["id"],
        "dataset": "coa-reconciliation",
        "mode": mode,
        "filename": preview.get("filename"),
        "records": existing_records,
        "record_count": len(existing_records),
        "created_at": now,
        "created_by": user["id"],
    })

    inserted = updated = deleted = unchanged = 0
    if mode == "replace":
        deleted = (await db.coa_reconciliation.delete_many({})).deleted_count
        if prepared_records:
            await db.coa_reconciliation.insert_many(prepared_records)
            inserted = len(prepared_records)
    else:
        for record in prepared_records:
            key = normalize_coa_shipment(record.get("shipment"))
            existing = existing_by_key.get(key)
            if existing and existing.get("id"):
                if not coa_records_differ(record, existing):
                    unchanged += 1
                    continue
                result = await db.coa_reconciliation.replace_one({"id": existing["id"]}, record)
                updated += result.modified_count or 1
            elif record.get("shipment"):
                result = await db.coa_reconciliation.update_one(
                    {"shipment": record["shipment"]},
                    {"$set": record},
                    upsert=True,
                )
                if result.upserted_id:
                    inserted += 1
                else:
                    updated += result.modified_count or 1

    after_total = await db.coa_reconciliation.count_documents({})
    history = {
        "id": str(uuid.uuid4()),
        "preview_id": preview["id"],
        "dataset": "coa-reconciliation",
        "filename": preview.get("filename"),
        "mode": mode,
        "row_count": len(records),
        "source_counts": preview.get("source_counts") or {},
        "validation_summary": preview.get("validation_summary") or {},
        "diff_summary": preview.get("diff_summary") or {},
        "preservation_summary": preview.get("preservation_summary") or {},
        "before_total": len(existing_records),
        "after_total": after_total,
        "inserted": inserted,
        "updated": updated,
        "unchanged": unchanged,
        "deleted": deleted,
        "snapshot_id": snapshot_id,
        "created_at": now,
        "created_by": user["id"],
        "created_by_name": user.get("name"),
    }
    await db.import_history.insert_one(history)
    await db.import_previews.update_one(
        {"id": preview["id"]},
        {"$set": {
            "status": "committed",
            "committed_at": now,
            "committed_by": user["id"],
            "commit_mode": mode,
            "history_id": history["id"],
        }},
    )
    return {key: value for key, value in history.items() if key != "_id"}


@router.get("/settings/coa")
async def get_coa_settings(user: dict = Depends(get_current_user)):
    """Get COA settings including price per kCal"""
    settings = await db.app_settings.find_one({"type": "coa"}, {"_id": 0})
    if not settings:
        return {"price_per_kcal_per_ton": None}
    return settings


@router.put("/settings/coa")
async def update_coa_settings(data: COASettingsUpdate, user: dict = Depends(require_role(["admin"]))):
    """Update COA settings"""
    await db.app_settings.update_one(
        {"type": "coa"},
        {"$set": {
            "type": "coa",
            "price_per_kcal_per_ton": data.price_per_kcal_per_ton,
            "updated_at": datetime.now(timezone.utc).isoformat(),
            "updated_by": user["id"]
        }},
        upsert=True
    )
    return {"message": "Pengaturan COA berhasil disimpan", "price_per_kcal_per_ton": data.price_per_kcal_per_ton}


@router.get("/coa-reconciliation")
async def get_coa_reconciliation(
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=50000),
    status: Optional[str] = None,
    search: Optional[str] = None,
    supplier: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    user: dict = Depends(get_current_user)
):
    """Get COA reconciliation data with pagination and drilldown filters"""
    query = _coa_filter_query(status=status, search=search, supplier=supplier, date_from=date_from, date_to=date_to)

    skip = (page - 1) * page_size
    total = await db.coa_reconciliation.count_documents(query)
    items = await db.coa_reconciliation.find(query, {"_id": 0}).sort("completed_unloading", -1).skip(skip).limit(page_size).to_list(page_size)

    return {
        "items": items,
        "total": total,
        "page": page,
        "page_size": page_size,
        "total_pages": (total + page_size - 1) // page_size
    }


@router.get("/coa-reconciliation/kpis")
async def get_coa_kpis(
    status: Optional[str] = None,
    supplier: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    user: dict = Depends(get_current_user)
):
    """Get KPIs for COA Reconciliation dashboard"""
    query = _coa_filter_query(status=status, supplier=supplier, date_from=date_from, date_to=date_to)
    all_data = await db.coa_reconciliation.find(query, {"_id": 0}).to_list(10000)

    coa_settings = await db.app_settings.find_one({"type": "coa"}, {"_id": 0})
    price_per_kcal = coa_settings.get("price_per_kcal_per_ton") if coa_settings else None

    if not all_data:
        return {
            "total_records": 0,
            "high_deviation_count": 0,
            "potential_loss_rp": 0,
            "total_tonnage_problem": 0,
            "umpire_status": {"proposed": 0, "in_progress": 0, "completed": 0, "total": 0},
            "supplier_deviations": [],
            "worst_supplier": None,
            "avg_accuracy": 100,
            "critical_count": 0,
            "warning_count": 0,
            "normal_count": 0,
            "price_per_kcal_per_ton": price_per_kcal,
            "price_not_set": price_per_kcal is None
        }

    kpis = calculate_kpis(all_data, price_per_kcal)
    kpis["price_per_kcal_per_ton"] = price_per_kcal
    kpis["price_not_set"] = price_per_kcal is None
    return kpis


@router.get("/coa-reconciliation/trend")
async def get_coa_trend(
    months: int = Query(3, ge=1, le=12),
    status: Optional[str] = None,
    supplier: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    user: dict = Depends(get_current_user)
):
    """Get GCV trend data for line chart"""
    query = _coa_filter_query(status=status, supplier=supplier, date_from=date_from, date_to=date_to)
    all_data = await db.coa_reconciliation.find(query, {"_id": 0}).to_list(10000)
    return get_gcv_trend_data(all_data, months)


@router.get("/coa-reconciliation/supplier-consistency")
async def get_supplier_consistency(
    status: Optional[str] = None,
    supplier: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    user: dict = Depends(get_current_user)
):
    """Get supplier consistency data for bar chart"""
    query = _coa_filter_query(status=status, supplier=supplier, date_from=date_from, date_to=date_to)
    all_data = await db.coa_reconciliation.find(query, {"_id": 0}).to_list(10000)
    kpis = calculate_kpis(all_data)
    return kpis.get("supplier_deviations", [])


@router.get("/coa-reconciliation/dispute-monitor")
async def get_dispute_monitor(
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=50000),
    umpire_status: Optional[str] = None,
    status: Optional[str] = None,
    supplier: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    user: dict = Depends(get_current_user)
):
    """Get all records that have umpire activity (proposed, in_progress, or completed)"""
    selected_status = umpire_status or status
    query = _coa_filter_query(
        include_disputes=True,
        umpire_status=selected_status,
        supplier=supplier,
        date_from=date_from,
        date_to=date_to,
    )

    skip = (page - 1) * page_size
    total = await db.coa_reconciliation.count_documents(query)
    items = await db.coa_reconciliation.find(query, {"_id": 0}).sort("umpire_proposed_at", -1).skip(skip).limit(page_size).to_list(page_size)
    for item in items:
        item["dispute_workflow"] = _workflow_summary(item)

    summary_query = _coa_filter_query(
        include_disputes=True,
        supplier=supplier,
        date_from=date_from,
        date_to=date_to,
    )
    all_disputes = await db.coa_reconciliation.find(summary_query, {"_id": 0, "umpire_status": 1}).to_list(10000)
    summary = {
        "proposed": sum(1 for d in all_disputes if d.get("umpire_status") == "proposed"),
        "in_progress": sum(1 for d in all_disputes if d.get("umpire_status") == "in_progress"),
        "completed": sum(1 for d in all_disputes if d.get("umpire_status") == "completed"),
        "total": len(all_disputes)
    }

    return {
        "items": items,
        "total": total,
        "page": page,
        "page_size": page_size,
        "total_pages": (total + page_size - 1) // page_size,
        "summary": summary
    }


@router.post("/coa-reconciliation/preview-combined")
async def preview_combined_coa_file(
    file: UploadFile = File(..., description="Workbook gabungan Rekapitulasi CoA"),
    user: dict = Depends(require_role(["admin", "operator"]))
):
    """Preview one combined COA workbook without mutating coa_reconciliation."""
    try:
        contents = await file.read()
        merged_data, source_counts = parse_combined_coa_workbook(contents)
        existing_records = await db.coa_reconciliation.find({}, {"_id": 0}).to_list(50000)
        preview = build_combined_coa_import_preview(merged_data, source_counts, existing_records)
        data_quality = summarize_coa_preview_quality(preview)
        preview_id = str(uuid.uuid4())
        now = datetime.now(timezone.utc).isoformat()
        preview_doc = {
            "id": preview_id,
            "dataset": "coa-reconciliation",
            "filename": file.filename or "coa-combined.xlsx",
            "records": merged_data,
            "source_counts": source_counts,
            "coverage": preview["coverage"],
            "issues": preview["issues"],
            "validation_summary": preview["validation_summary"],
            "diff_summary": preview["diff_summary"],
            "preservation_summary": preview["preservation_summary"],
            "data_quality": data_quality,
            "status": "previewed",
            "created_at": now,
            "created_by": user["id"],
            "created_by_name": user.get("name"),
        }
        await db.import_previews.insert_one(preview_doc)
        return _public_preview_response(preview_doc)
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error previewing combined COA workbook: {e}")
        raise HTTPException(status_code=400, detail=f"Gagal preview workbook gabungan: {str(e)}")


@router.get("/coa-reconciliation/import-preview/{preview_id}")
async def get_coa_import_preview(preview_id: str, user: dict = Depends(get_current_user)):
    """Get a stored COA import preview."""
    preview = await db.import_previews.find_one({"id": preview_id, "dataset": "coa-reconciliation"}, {"_id": 0})
    if not preview:
        raise HTTPException(status_code=404, detail="Preview import COA tidak ditemukan")
    return _public_preview_response(preview)


@router.post("/coa-reconciliation/import-preview/{preview_id}/commit")
async def commit_coa_import_preview(
    preview_id: str,
    request: COAImportCommitRequest = Body(default_factory=COAImportCommitRequest),
    user: dict = Depends(require_role(["admin", "operator"]))
):
    """Commit a previously previewed COA workbook using merge or confirmed replace-all mode."""
    preview = await db.import_previews.find_one({"id": preview_id, "dataset": "coa-reconciliation"}, {"_id": 0})
    if not preview:
        raise HTTPException(status_code=404, detail="Preview import COA tidak ditemukan")
    if preview.get("status") == "committed":
        raise HTTPException(status_code=400, detail="Preview import COA sudah pernah dicommit")

    result = await _commit_combined_coa_preview(preview, request.mode, request.confirm_replace_all, user)
    return {"message": "Import COA berhasil dicommit", **result}


@router.get("/coa-reconciliation/import-history")
async def get_coa_import_history(
    page: int = Query(1, ge=1),
    page_size: int = Query(10, ge=1, le=100),
    user: dict = Depends(get_current_user)
):
    """Get COA import commit history."""
    query = {"dataset": "coa-reconciliation"}
    skip = (page - 1) * page_size
    total = await db.import_history.count_documents(query)
    items = await db.import_history.find(query, {"_id": 0}).sort("created_at", -1).skip(skip).limit(page_size).to_list(page_size)
    return {"items": items, "total": total, "page": page, "page_size": page_size, "total_pages": (total + page_size - 1) // page_size}


@router.post("/coa-reconciliation/import-history/{history_id}/rollback")
async def rollback_coa_import(history_id: str, user: dict = Depends(require_role(["admin"]))):
    """Rollback a COA import using the snapshot captured immediately before commit."""
    history = await db.import_history.find_one({"id": history_id, "dataset": "coa-reconciliation"}, {"_id": 0})
    if not history:
        raise HTTPException(status_code=404, detail="Riwayat import COA tidak ditemukan")
    if history.get("rolled_back_at"):
        raise HTTPException(status_code=400, detail="Import ini sudah pernah di-rollback")

    snapshot_id = history.get("snapshot_id")
    snapshot = await db.coa_import_snapshots.find_one({"id": snapshot_id}, {"_id": 0})
    if not snapshot:
        raise HTTPException(status_code=404, detail="Snapshot rollback tidak ditemukan")

    before_total = await db.coa_reconciliation.count_documents({})
    await db.coa_reconciliation.delete_many({})
    snapshot_records = snapshot.get("records") or []
    if snapshot_records:
        await db.coa_reconciliation.insert_many(snapshot_records)
    after_total = await db.coa_reconciliation.count_documents({})
    now = datetime.now(timezone.utc).isoformat()
    await db.import_history.update_one(
        {"id": history_id},
        {"$set": {
            "rolled_back_at": now,
            "rolled_back_by": user["id"],
            "rollback_before_total": before_total,
            "rollback_after_total": after_total,
        }},
    )
    return {
        "message": "Rollback import COA berhasil",
        "history_id": history_id,
        "before_total": before_total,
        "after_total": after_total,
    }


@router.get("/coa-reconciliation/{record_id}")
async def get_coa_reconciliation_detail(record_id: str, user: dict = Depends(get_current_user)):
    """Get single COA reconciliation record with radar chart data"""
    record = await db.coa_reconciliation.find_one({"id": record_id}, {"_id": 0})
    if not record:
        raise HTTPException(status_code=404, detail="Data tidak ditemukan")

    radar_data = get_radar_chart_data(record)
    record["dispute_workflow"] = _workflow_summary(record)
    return {
        "record": record,
        "radar_chart": radar_data
    }


@router.get("/coa-reconciliation/shipment/{shipment}")
async def get_coa_by_shipment(shipment: str, user: dict = Depends(get_current_user)):
    """Get COA reconciliation record by shipment ID (string)"""
    record = await db.coa_reconciliation.find_one({"shipment": shipment}, {"_id": 0})
    if not record:
        raise HTTPException(status_code=404, detail="Data tidak ditemukan")

    radar_data = get_radar_chart_data(record)
    record["dispute_workflow"] = _workflow_summary(record)
    return {
        "record": record,
        "radar_chart": radar_data
    }


@router.post("/coa-reconciliation/propose-umpire")
async def propose_umpire(data: UmpireProposal, user: dict = Depends(require_role(["admin", "operator"]))):
    """Propose umpire testing for a reconciliation record"""
    event = _workflow_event("propose", "proposed", user, data.notes)
    result = await db.coa_reconciliation.update_one(
        {"id": data.reconciliation_id},
        {
            "$set": {
                "umpire_status": "proposed",
                "umpire_sample_number": data.sample_number,
                "umpire_notes": data.notes,
                "umpire_proposed_at": event["created_at"],
                "umpire_proposed_by": user["id"],
                "dispute_closed_at": None,
                "dispute_resolution": None,
                "dispute_closure_notes": None,
            },
            "$push": {"dispute_history": event},
        }
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Data tidak ditemukan")
    return {"message": "Umpire testing berhasil diajukan", "sample_number": data.sample_number}


@router.post("/coa-reconciliation/update-umpire-status/{record_id}")
async def update_umpire_status(
    record_id: str,
    status: str = Query(..., regex="^(none|proposed|in_progress|completed)$"),
    user: dict = Depends(require_role(["admin", "operator"]))
):
    """Update umpire testing status"""
    update_data = {"umpire_status": status}
    if status == "completed":
        update_data["umpire_completed_at"] = datetime.now(timezone.utc).isoformat()
    if status == "in_progress":
        update_data["umpire_started_at"] = datetime.now(timezone.utc).isoformat()
    event = _workflow_event("status_change", status, user)

    result = await db.coa_reconciliation.update_one(
        {"id": record_id},
        {"$set": update_data, "$push": {"dispute_history": event}},
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Data tidak ditemukan")
    return {"message": f"Status umpire berhasil diubah ke {status}"}


@router.post("/coa-reconciliation/submit-umpire-result")
async def submit_umpire_result(data: UmpireResultInput, user: dict = Depends(require_role(["admin", "operator"]))):
    """Submit umpire test results for a reconciliation record"""
    event = _workflow_event("submit_result", "completed", user, data.notes)
    result = await db.coa_reconciliation.update_one(
        {"id": data.reconciliation_id},
        {
            "$set": {
                "umpire_status": "completed",
                "umpire_gcv_arb": data.umpire_gcv_arb,
                "umpire_tm_arb": data.umpire_tm_arb,
                "umpire_ash_arb": data.umpire_ash_arb,
                "umpire_ts_arb": data.umpire_ts_arb,
                "umpire_lab_name": data.umpire_lab_name,
                "umpire_result_date": data.umpire_result_date,
                "umpire_result_notes": data.notes,
                "umpire_completed_at": event["created_at"],
                "umpire_completed_by": user["id"]
            },
            "$push": {"dispute_history": event},
        }
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Data tidak ditemukan")
    return {"message": "Hasil umpire berhasil disimpan", "reconciliation_id": data.reconciliation_id}


@router.post("/coa-reconciliation/{record_id}/dispute-notes")
async def add_dispute_note(
    record_id: str,
    data: DisputeNoteInput,
    user: dict = Depends(require_role(["admin", "operator"]))
):
    """Add a structured note to a dispute workflow."""
    note = {
        "id": str(uuid.uuid4()),
        "note": data.note,
        "visibility": data.visibility,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "created_by": user["id"],
        "created_by_name": user.get("name"),
    }
    event = _workflow_event("add_note", "note", user, data.note)
    result = await db.coa_reconciliation.update_one(
        {"id": record_id},
        {"$push": {"dispute_notes": note, "dispute_history": event}},
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Data tidak ditemukan")
    return {"message": "Catatan dispute berhasil ditambahkan", "note": note}


@router.post("/coa-reconciliation/{record_id}/dispute-attachments")
async def add_dispute_attachment(
    record_id: str,
    data: DisputeAttachmentInput,
    user: dict = Depends(require_role(["admin", "operator"]))
):
    """Attach document metadata to a dispute workflow."""
    attachment = {
        "id": str(uuid.uuid4()),
        "filename": data.filename,
        "url": data.url,
        "description": data.description,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "created_by": user["id"],
        "created_by_name": user.get("name"),
    }
    event = _workflow_event("add_attachment", "attachment", user, data.filename)
    result = await db.coa_reconciliation.update_one(
        {"id": record_id},
        {"$push": {"dispute_attachments": attachment, "dispute_history": event}},
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Data tidak ditemukan")
    return {"message": "Dokumen dispute berhasil dicatat", "attachment": attachment}


@router.post("/coa-reconciliation/{record_id}/close-dispute")
async def close_dispute(
    record_id: str,
    data: DisputeCloseInput,
    user: dict = Depends(require_role(["admin", "operator"]))
):
    """Close dispute workflow with explicit resolution."""
    event = _workflow_event("close", "closed", user, data.closure_notes)
    result = await db.coa_reconciliation.update_one(
        {"id": record_id},
        {
            "$set": {
                "umpire_status": "completed",
                "dispute_closed_at": event["created_at"],
                "dispute_closed_by": user["id"],
                "dispute_resolution": data.resolution,
                "dispute_closure_notes": data.closure_notes,
            },
            "$push": {"dispute_history": event},
        },
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Data tidak ditemukan")
    return {"message": "Dispute berhasil ditutup", "resolution": data.resolution}


@router.post("/coa-reconciliation/upload")
async def upload_coa_files(
    loading_file: UploadFile = File(..., description="Loading.xlsx"),
    unloading_file: UploadFile = File(..., description="Unloading.xlsx"),
    internal_file: UploadFile = File(..., description="Lab Internal.xlsx"),
    user: dict = Depends(require_role(["admin", "operator"]))
):
    """Upload and process COA files (Loading, Unloading, Lab Internal)"""
    try:
        loading_contents = await loading_file.read()
        unloading_contents = await unloading_file.read()
        internal_contents = await internal_file.read()

        loading_data = parse_coa_excel(loading_contents, "loading")
        unloading_data = parse_coa_excel(unloading_contents, "unloading")
        internal_data = parse_coa_excel(internal_contents, "internal")

        merged_data = merge_coa_data(loading_data, unloading_data, internal_data)

        if merged_data:
            await db.coa_reconciliation.delete_many({})
            for record in merged_data:
                record["uploaded_by"] = user["id"]
                record["uploaded_at"] = datetime.now(timezone.utc).isoformat()
            await db.coa_reconciliation.insert_many(merged_data)

        return {
            "message": f"Berhasil memproses dan menyimpan {len(merged_data)} data rekonsiliasi COA",
            "count": len(merged_data),
            "sources": {
                "loading": len(loading_data),
                "unloading": len(unloading_data),
                "internal": len(internal_data)
            }
        }
    except Exception as e:
        logger.error(f"Error processing COA files: {e}")
        raise HTTPException(status_code=400, detail=f"Gagal memproses file: {str(e)}")


@router.post("/coa-reconciliation/upload-combined")
async def upload_combined_coa_file(
    file: UploadFile = File(..., description="Workbook gabungan Rekapitulasi CoA"),
    user: dict = Depends(require_role(["admin", "operator"]))
):
    """Upload and process one combined COA workbook with Loading, Unloading, Internal, and Umpire data."""
    try:
        contents = await file.read()
        merged_data, source_counts = parse_combined_coa_workbook(contents)

        if merged_data:
            await db.coa_reconciliation.delete_many({})
            uploaded_at = datetime.now(timezone.utc).isoformat()
            for record in merged_data:
                record["uploaded_by"] = user["id"]
                record["uploaded_at"] = uploaded_at
                record["import_source"] = file.filename
            await db.coa_reconciliation.insert_many(merged_data)

        return {
            "message": f"Berhasil memproses workbook gabungan dan menyimpan {len(merged_data)} data rekonsiliasi COA",
            "count": len(merged_data),
            "sources": source_counts,
        }
    except Exception as e:
        logger.error(f"Error processing combined COA workbook: {e}")
        raise HTTPException(status_code=400, detail=f"Gagal memproses workbook gabungan: {str(e)}")


@router.delete("/coa-reconciliation")
async def delete_all_coa_reconciliation(user: dict = Depends(require_role(["admin"]))):
    """Delete all COA reconciliation data"""
    result = await db.coa_reconciliation.delete_many({})
    return {"message": f"Berhasil menghapus {result.deleted_count} data rekonsiliasi COA", "count": result.deleted_count}


@router.post("/coa-reconciliation/manual")
async def add_coa_manual(data: COAManualInput, user: dict = Depends(require_role(["admin", "operator"]))):
    """Add COA reconciliation data manually"""
    existing = await db.coa_reconciliation.find_one({"shipment": data.shipment})
    if existing:
        raise HTTPException(status_code=400, detail=f"Shipment {data.shipment} sudah ada. Gunakan fitur edit atau hapus terlebih dahulu.")

    delta_loading_internal = None
    delta_unloading_internal = None
    delta_loading_unloading = None

    if data.loading_gcv_arb and data.internal_gcv_arb:
        delta_loading_internal = data.loading_gcv_arb - data.internal_gcv_arb
    if data.unloading_gcv_arb and data.internal_gcv_arb:
        delta_unloading_internal = data.unloading_gcv_arb - data.internal_gcv_arb
    if data.loading_gcv_arb and data.unloading_gcv_arb:
        delta_loading_unloading = data.loading_gcv_arb - data.unloading_gcv_arb

    status = "normal"
    if delta_loading_internal is not None:
        if delta_loading_internal > 150:
            status = "critical"
        elif delta_loading_internal > 100:
            status = "warning"

    record = {
        "id": str(uuid.uuid4()),
        "shipment": data.shipment,
        "suppliers": data.suppliers,
        "periode": data.periode or datetime.now(timezone.utc).strftime("%Y-%m-%d"),
        "tb": data.tb or "",
        "bg": data.bg or "",
        "ds_mt": data.ds_mt,
        "completed_unloading": data.completed_unloading,
        "loading_gcv_arb": data.loading_gcv_arb,
        "loading_tm_arb": data.loading_tm_arb,
        "loading_ash_arb": data.loading_ash_arb,
        "loading_ts_arb": data.loading_ts_arb,
        "loading_surveyor": None,
        "loading_no_coa": None,
        "unloading_gcv_arb": data.unloading_gcv_arb,
        "unloading_tm_arb": data.unloading_tm_arb,
        "unloading_ash_arb": data.unloading_ash_arb,
        "unloading_ts_arb": data.unloading_ts_arb,
        "unloading_surveyor": None,
        "unloading_slagging": None,
        "unloading_fouling": None,
        "internal_gcv_arb": data.internal_gcv_arb,
        "internal_tm_arb": data.internal_tm_arb,
        "internal_ash_arb": data.internal_ash_arb,
        "internal_ts_arb": data.internal_ts_arb,
        "delta_loading_internal": delta_loading_internal,
        "delta_unloading_internal": delta_unloading_internal,
        "delta_loading_unloading": delta_loading_unloading,
        "status": status,
        "umpire_status": "none",
        "umpire_sample_number": None,
        "umpire_proposed_at": None,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "created_by": user["id"],
        "input_method": "manual"
    }

    await db.coa_reconciliation.insert_one(record)
    return {"message": f"Berhasil menambahkan data COA Shipment {data.shipment}", "id": record["id"], "status": status}


@router.get("/coa-reconciliation/export/excel")
async def export_coa_to_excel(
    status_filter: str = Query("all"),
    user: dict = Depends(get_current_user)
):
    """Export COA reconciliation data to Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    query = {}
    if status_filter == "kritis":
        query["status"] = "Kritis"
    elif status_filter == "umpire":
        query["umpire_status"] = {"$nin": [None, "none", ""]}

    data = await db.coa_reconciliation.find(query, {"_id": 0}).sort("completed_unloading", -1).to_list(10000)

    wb = Workbook()
    ws = wb.active
    ws.title = "COA Reconciliation"

    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    kritis_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    headers = [
        "No", "Shipment", "Supplier", "Tanggal Unloading", "DS (MT)",
        "Loading GCV", "Loading TM", "Loading Ash", "Loading S",
        "Unloading GCV", "Unloading TM", "Unloading Ash", "Unloading S",
        "Internal GCV", "Internal TM", "Internal Ash", "Internal S",
        "Delta GCV", "Status", "Umpire Status", "Umpire GCV"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    for row_idx, record in enumerate(data, 2):
        row_data = [
            row_idx - 1,
            record.get("shipment", ""),
            record.get("suppliers", ""),
            record.get("completed_unloading", "")[:10] if record.get("completed_unloading") else "",
            record.get("ds_mt", 0),
            record.get("loading_gcv_arb"),
            record.get("loading_tm_arb"),
            record.get("loading_ash_arb"),
            record.get("loading_ts_arb"),
            record.get("unloading_gcv_arb"),
            record.get("unloading_tm_arb"),
            record.get("unloading_ash_arb"),
            record.get("unloading_ts_arb"),
            record.get("internal_gcv_arb"),
            record.get("internal_tm_arb"),
            record.get("internal_ash_arb"),
            record.get("internal_ts_arb"),
            record.get("delta_loading_internal"),
            record.get("status", "normal").upper(),
            record.get("umpire_status", "none"),
            record.get("umpire_result", {}).get("gcv") if record.get("umpire_result") else None
        ]

        is_kritis = record.get("status") == "Kritis"

        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = border
            cell.alignment = Alignment(horizontal='center' if col <= 4 else 'right')
            if is_kritis:
                cell.fill = kritis_fill

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 30

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"COA_Reconciliation_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/coa-reconciliation/export/pdf")
async def export_coa_to_pdf(
    status_filter: str = Query("all"),
    user: dict = Depends(get_current_user)
):
    """Export COA reconciliation data to PDF"""
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import inch
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    query = {}
    if status_filter == "kritis":
        query["status"] = "Kritis"
    elif status_filter == "umpire":
        query["umpire_status"] = {"$nin": [None, "none", ""]}

    data = await db.coa_reconciliation.find(query, {"_id": 0}).sort("completed_unloading", -1).to_list(10000)

    all_coa = await db.coa_reconciliation.find({}, {"_id": 0}).to_list(10000)
    kritis_count = sum(1 for c in all_coa if c.get("status") == "Kritis")
    umpire_count = sum(1 for c in all_coa if c.get("umpire_status") not in [None, "none", ""])

    settings = await db.app_settings.find_one({"type": "coa"})
    price_per_kcal = settings.get("price_per_kcal_per_ton", 50) if settings else 50
    potential_loss = 0
    for c in all_coa:
        delta = c.get("delta_loading_internal", 0) or 0
        tonase = c.get("ds_mt", 0) or 0
        if delta > 0:
            potential_loss += delta * tonase * price_per_kcal

    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=landscape(A4), leftMargin=0.5*inch, rightMargin=0.5*inch)
    elements = []
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=16, alignment=TA_CENTER, spaceAfter=20)
    elements.append(Paragraph("LAPORAN COA RECONCILIATION - PLTU TENAYAN", title_style))
    elements.append(Paragraph(f"Tanggal: {datetime.now().strftime('%d %B %Y')}", ParagraphStyle('SubTitle', parent=styles['Normal'], alignment=TA_CENTER, spaceAfter=20)))

    kpi_data = [
        ["Total Records", "Status Kritis", "Proses Umpire", "Potential Loss"],
        [str(len(all_coa)), str(kritis_count), str(umpire_count), f"Rp {potential_loss:,.0f}"]
    ]
    kpi_table = Table(kpi_data, colWidths=[2.5*inch, 2.5*inch, 2.5*inch, 2.5*inch])
    kpi_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E3A5F')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F0F4F8')),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 11),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#CCCCCC')),
    ]))
    elements.append(kpi_table)
    elements.append(Spacer(1, 20))

    headers = ["No", "Shipment", "Supplier", "Tanggal", "Loading GCV", "Unloading GCV", "Internal GCV", "Delta", "Status"]
    table_data = [headers]

    for idx, record in enumerate(data[:100], 1):
        status = record.get("status", "normal")
        row = [
            str(idx),
            str(record.get("shipment", ""))[:15],
            str(record.get("suppliers", ""))[:20],
            str(record.get("completed_unloading", ""))[:10] if record.get("completed_unloading") else "-",
            str(record.get("loading_gcv_arb") or "-"),
            str(record.get("unloading_gcv_arb") or "-"),
            str(record.get("internal_gcv_arb") or "-"),
            str(int(record.get("delta_loading_internal") or 0)),
            status.upper()
        ]
        table_data.append(row)

    main_table = Table(table_data, colWidths=[0.4*inch, 1.2*inch, 1.8*inch, 0.9*inch, 1*inch, 1*inch, 1*inch, 0.8*inch, 0.8*inch])

    table_style = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E3A5F')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F5F5')]),
    ]

    for idx, record in enumerate(data[:100], 1):
        if record.get("status") == "Kritis":
            table_style.append(('BACKGROUND', (0, idx), (-1, idx), colors.HexColor('#FFCCCC')))

    main_table.setStyle(TableStyle(table_style))
    elements.append(main_table)

    elements.append(Spacer(1, 20))
    footer_style = ParagraphStyle('Footer', parent=styles['Normal'], fontSize=8, alignment=TA_LEFT)
    elements.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Filter: {status_filter.upper()}", footer_style))

    doc.build(elements)
    output.seek(0)

    filename = f"COA_Reconciliation_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"

    return Response(
        content=output.getvalue(),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
