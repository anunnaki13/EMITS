from fastapi import FastAPI, APIRouter, HTTPException, Depends, UploadFile, File, Query
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, EmailStr
from typing import List, Optional, Any, Dict
import uuid
from datetime import datetime, timezone, timedelta
import jwt
import bcrypt
import pandas as pd
import io
from emergentintegrations.llm.chat import LlmChat, UserMessage
import json
import asyncio

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# JWT Settings
JWT_SECRET = os.environ.get('JWT_SECRET', 'tenayan-fuel-management-secret-key-2024')
JWT_ALGORITHM = "HS256"
JWT_EXPIRATION_HOURS = 24

# Create the main app
app = FastAPI(title="PLTU Tenayan Fuel Management System")

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")
security = HTTPBearer()

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# ==================== MODELS ====================

class UserCreate(BaseModel):
    email: EmailStr
    password: str
    name: str
    role: str = "operator"  # admin, operator, viewer

class UserLogin(BaseModel):
    email: EmailStr
    password: str

class UserResponse(BaseModel):
    id: str
    email: str
    name: str
    role: str
    created_at: str

class TokenResponse(BaseModel):
    access_token: str
    token_type: str = "bearer"
    user: UserResponse

# Vessel TNY Model - Complete fields based on PLTU Tenayan Excel template
class VesselTNYCreate(BaseModel):
    # Informasi Shipment
    periode_ta: str
    periode_realisasi: str
    shipment_code: str
    voyage_code: str
    suppliers: str
    voyage: str
    name_of_vessel: str
    coal_from: str
    # Waktu Operasional
    time_arrival: Optional[str] = None
    berthed_time: Optional[str] = None
    commenced_unloading: Optional[str] = None
    completed_unloading: Optional[str] = None
    durasi_pembongkaran_hari: Optional[float] = None
    durasi_pembongkaran_jam: Optional[float] = None
    waktu_tunggu_jam: Optional[float] = None
    # Muatan
    bl_mt: Optional[float] = None
    ds_mt: Optional[float] = None
    # COW (Certificate of Weighing)
    no_cow: Optional[str] = None
    tgl_terbit_cow: Optional[str] = None
    # Kualitas - GCV
    gcv_arb: Optional[float] = None
    gcv_adb: Optional[float] = None
    gcv_db: Optional[float] = None
    # Kualitas - Moisture
    tm_arb: Optional[float] = None
    im_adb: Optional[float] = None
    # Kualitas - Ash Content
    ash_arb: Optional[float] = None
    ash_adb: Optional[float] = None
    ash_db: Optional[float] = None
    # Kualitas - VM & FC
    vm_arb: Optional[float] = None
    vm_adb: Optional[float] = None
    fc_arb: Optional[float] = None
    fc_adb: Optional[float] = None
    # Kualitas - Sulfur
    ts_arb: Optional[float] = None
    ts_adb: Optional[float] = None
    ts_db: Optional[float] = None
    ts_dafb: Optional[float] = None
    # Ultimate Analysis
    c_arb: Optional[float] = None
    c_adb: Optional[float] = None
    h_arb: Optional[float] = None
    h_adb: Optional[float] = None
    n_arb: Optional[float] = None
    n_adb: Optional[float] = None
    n_dafb: Optional[float] = None
    o_arb: Optional[float] = None
    o_adb: Optional[float] = None
    # HGI & Index
    hgi: Optional[float] = None
    slagging_index: Optional[str] = None
    fouling_index: Optional[str] = None
    idt_reducing: Optional[float] = None
    # Ash Composition
    sio2_db: Optional[float] = None
    al2o3_db: Optional[float] = None
    tio2_db: Optional[float] = None
    fe2o3_db: Optional[float] = None
    cao_db: Optional[float] = None
    mgo_db: Optional[float] = None
    k2o_db: Optional[float] = None
    na2o_db: Optional[float] = None
    so3_db: Optional[float] = None
    p2o5_db: Optional[float] = None
    mno2_db: Optional[float] = None
    mn3o4_db: Optional[float] = None
    # Size Analysis
    size_70mm: Optional[float] = None
    size_50mm: Optional[float] = None
    size_32mm: Optional[float] = None
    size_2_38mm: Optional[float] = None
    # COA (Certificate of Analysis)
    no_coa: Optional[str] = None
    tgl_terbit_coa: Optional[str] = None
    durasi_terbit_coa: Optional[str] = None

class VesselTNYResponse(BaseModel):
    id: str
    # Informasi Shipment
    periode_ta: str
    periode_realisasi: str
    shipment_code: str
    voyage_code: str
    suppliers: str
    voyage: str
    name_of_vessel: str
    coal_from: str
    # Waktu Operasional
    time_arrival: Optional[str] = None
    berthed_time: Optional[str] = None
    commenced_unloading: Optional[str] = None
    completed_unloading: Optional[str] = None
    durasi_pembongkaran_hari: Optional[float] = None
    durasi_pembongkaran_jam: Optional[float] = None
    waktu_tunggu_jam: Optional[float] = None
    # Muatan
    bl_mt: Optional[float] = None
    ds_mt: Optional[float] = None
    # COW
    no_cow: Optional[str] = None
    tgl_terbit_cow: Optional[str] = None
    # Kualitas - GCV
    gcv_arb: Optional[float] = None
    gcv_adb: Optional[float] = None
    gcv_db: Optional[float] = None
    # Kualitas - Moisture
    tm_arb: Optional[float] = None
    im_adb: Optional[float] = None
    # Kualitas - Ash Content
    ash_arb: Optional[float] = None
    ash_adb: Optional[float] = None
    ash_db: Optional[float] = None
    # Kualitas - VM & FC
    vm_arb: Optional[float] = None
    vm_adb: Optional[float] = None
    fc_arb: Optional[float] = None
    fc_adb: Optional[float] = None
    # Kualitas - Sulfur
    ts_arb: Optional[float] = None
    ts_adb: Optional[float] = None
    ts_db: Optional[float] = None
    ts_dafb: Optional[float] = None
    # Ultimate Analysis
    c_arb: Optional[float] = None
    c_adb: Optional[float] = None
    h_arb: Optional[float] = None
    h_adb: Optional[float] = None
    n_arb: Optional[float] = None
    n_adb: Optional[float] = None
    n_dafb: Optional[float] = None
    o_arb: Optional[float] = None
    o_adb: Optional[float] = None
    # HGI & Index
    hgi: Optional[float] = None
    slagging_index: Optional[str] = None
    fouling_index: Optional[str] = None
    idt_reducing: Optional[float] = None
    # Ash Composition
    sio2_db: Optional[float] = None
    al2o3_db: Optional[float] = None
    tio2_db: Optional[float] = None
    fe2o3_db: Optional[float] = None
    cao_db: Optional[float] = None
    mgo_db: Optional[float] = None
    k2o_db: Optional[float] = None
    na2o_db: Optional[float] = None
    so3_db: Optional[float] = None
    p2o5_db: Optional[float] = None
    mno2_db: Optional[float] = None
    mn3o4_db: Optional[float] = None
    # Size Analysis
    size_70mm: Optional[float] = None
    size_50mm: Optional[float] = None
    size_32mm: Optional[float] = None
    size_2_38mm: Optional[float] = None
    # COA
    no_coa: Optional[str] = None
    tgl_terbit_coa: Optional[str] = None
    durasi_terbit_coa: Optional[str] = None
    # Metadata
    created_at: str
    created_by: Optional[str] = None

# Barge TNY Model - Complete fields based on PLTU Tenayan Excel template
class BargeTNYCreate(BaseModel):
    # Informasi Shipment
    periode: str
    shipment_code: str
    voyage_code: str
    shipment: Optional[str] = None
    suppliers: str
    voyage: str
    tb: Optional[str] = None  # Tug Boat
    bg: Optional[str] = None  # Barge Name
    coal_from: str
    # Waktu Operasional
    ta: Optional[str] = None  # Time Arrival
    berthed_time: Optional[str] = None
    commenced_unloading: Optional[str] = None
    completed_unloading: Optional[str] = None
    durasi_pembongkaran_hari: Optional[float] = None
    durasi_pembongkaran_jam: Optional[float] = None
    waktu_tunggu_jam: Optional[float] = None
    # Muatan
    bl_mt: Optional[float] = None
    ds_mt: Optional[float] = None
    # COW
    no_cow: Optional[str] = None
    tgl_terbit_cow: Optional[str] = None
    # Kualitas - GCV
    gcv_arb: Optional[float] = None
    gcv_adb: Optional[float] = None
    gcv_db: Optional[float] = None
    # Kualitas - Moisture
    tm_arb: Optional[float] = None
    im_adb: Optional[float] = None
    # Kualitas - Ash Content
    ash_arb: Optional[float] = None
    ash_adb: Optional[float] = None
    ash_db: Optional[float] = None
    # Kualitas - VM & FC
    vm_arb: Optional[float] = None
    vm_adb: Optional[float] = None
    fc_arb: Optional[float] = None
    fc_adb: Optional[float] = None
    # Kualitas - Sulfur
    ts_arb: Optional[float] = None
    ts_adb: Optional[float] = None
    ts_db: Optional[float] = None
    ts_dafb: Optional[float] = None
    # Ultimate Analysis
    c_arb: Optional[float] = None
    c_adb: Optional[float] = None
    h_arb: Optional[float] = None
    h_adb: Optional[float] = None
    n_arb: Optional[float] = None
    n_adb: Optional[float] = None
    n_dafb: Optional[float] = None
    o_arb: Optional[float] = None
    o_adb: Optional[float] = None
    # HGI & Index
    hgi: Optional[float] = None
    slagging_index: Optional[str] = None
    fouling_index: Optional[str] = None
    idt_reducing: Optional[float] = None
    # Ash Composition
    sio2_db: Optional[float] = None
    al2o3_db: Optional[float] = None
    tio2_db: Optional[float] = None
    fe2o3_db: Optional[float] = None
    cao_db: Optional[float] = None
    mgo_db: Optional[float] = None
    k2o_db: Optional[float] = None
    na2o_db: Optional[float] = None
    so3_db: Optional[float] = None
    p2o5_db: Optional[float] = None
    mno2_db: Optional[float] = None
    mn3o4_db: Optional[float] = None
    # Size Analysis
    size_70mm: Optional[float] = None
    size_50mm: Optional[float] = None
    size_32mm: Optional[float] = None
    size_2_38mm: Optional[float] = None
    # COA
    no_coa: Optional[str] = None
    tgl_terbit_coa: Optional[str] = None
    durasi_terbit_coa: Optional[str] = None

class BargeTNYResponse(BargeTNYCreate):
    id: str
    created_at: str
    created_by: Optional[str] = None

# Trucking TNY Model - Complete fields based on PLTU Tenayan Excel template
class TruckingTNYCreate(BaseModel):
    # Informasi Shipment
    periode_ta: str
    periode_realisasi: str
    shipment_code: str
    voyage_code: str
    shipment: Optional[str] = None
    suppliers: str
    transportasi: Optional[str] = None
    coal_from: str
    # Waktu Operasional
    ta: Optional[str] = None
    berthed_time: Optional[str] = None
    commenced_unloading: Optional[str] = None
    completed_unloading: Optional[str] = None
    durasi_pembongkaran_hari: Optional[float] = None
    durasi_pembongkaran_jam: Optional[float] = None
    # Muatan
    bl_mt: Optional[float] = None
    ds_mt: Optional[float] = None
    rit: Optional[int] = None  # Jumlah RIT
    # COW
    no_cow: Optional[str] = None
    tgl_terbit_cow: Optional[str] = None
    # Kualitas - GCV
    gcv_arb: Optional[float] = None
    gcv_adb: Optional[float] = None
    gcv_db: Optional[float] = None
    # Kualitas - Moisture
    tm_arb: Optional[float] = None
    im_adb: Optional[float] = None
    # Kualitas - Ash Content
    ash_arb: Optional[float] = None
    ash_adb: Optional[float] = None
    ash_db: Optional[float] = None
    # Kualitas - VM & FC
    vm_arb: Optional[float] = None
    vm_adb: Optional[float] = None
    fc_arb: Optional[float] = None
    fc_adb: Optional[float] = None
    # Kualitas - Sulfur
    ts_arb: Optional[float] = None
    ts_adb: Optional[float] = None
    ts_db: Optional[float] = None
    ts_dafb: Optional[float] = None
    # Ultimate Analysis
    c_arb: Optional[float] = None
    c_adb: Optional[float] = None
    h_arb: Optional[float] = None
    h_adb: Optional[float] = None
    n_arb: Optional[float] = None
    n_adb: Optional[float] = None
    n_dafb: Optional[float] = None
    o_arb: Optional[float] = None
    o_adb: Optional[float] = None
    # HGI & Index
    hgi: Optional[float] = None
    slagging_index: Optional[str] = None
    fouling_index: Optional[str] = None
    idt_reducing: Optional[float] = None
    # Ash Composition
    sio2_db: Optional[float] = None
    al2o3_db: Optional[float] = None
    tio2_db: Optional[float] = None
    fe2o3_db: Optional[float] = None
    cao_db: Optional[float] = None
    mgo_db: Optional[float] = None
    k2o_db: Optional[float] = None
    na2o_db: Optional[float] = None
    so3_db: Optional[float] = None
    p2o5_db: Optional[float] = None
    mno2_db: Optional[float] = None
    mn3o4_db: Optional[float] = None
    # Size Analysis
    size_70mm: Optional[float] = None
    size_50mm: Optional[float] = None
    size_32mm: Optional[float] = None
    size_2_38mm: Optional[float] = None
    # COA
    no_coa: Optional[str] = None
    tgl_terbit_coa: Optional[str] = None

class TruckingTNYResponse(TruckingTNYCreate):
    id: str
    created_at: str
    created_by: Optional[str] = None

# Biomassa TNY Model - Complete fields based on PLTU Tenayan Excel template
class BiomassaTNYCreate(BaseModel):
    # Informasi Shipment
    periode: Optional[str] = None
    shipment_code: Optional[str] = None
    voyage_code: Optional[str] = None
    lot: Optional[str] = None
    suppliers: Optional[str] = None
    shipper: Optional[str] = None
    lot_1: Optional[str] = None
    tb: Optional[str] = None  # Tug Boat
    bg: Optional[str] = None  # Barge
    biomass_type: Optional[str] = None  # Jenis Biomassa
    # Waktu Operasional
    ta: Optional[str] = None  # Time Arrival
    berthed_time: Optional[str] = None
    commenced_unloading: Optional[str] = None
    completed_unloading: Optional[str] = None
    durasi_pembongkaran_hari: Optional[float] = None
    # Muatan
    bl_mt: Optional[float] = None
    jembatan_timbang_mt: Optional[float] = None
    surveyor_unloading: Optional[str] = None
    # COW/ROW
    no_cow_row: Optional[str] = None
    tgl_terbit_cow: Optional[str] = None
    lama_terbit_row: Optional[float] = None
    # Kualitas
    gcv_arb: Optional[float] = None
    gcv_adb: Optional[float] = None
    tm_arb: Optional[float] = None
    im_adb: Optional[float] = None
    # COA
    no_coa: Optional[str] = None
    tgl_terbit_coa: Optional[str] = None
    durasi_pembongkaran_hari_2: Optional[float] = None
    waktu_tunggu_jam: Optional[float] = None
    durasi_terbit_coa: Optional[float] = None

class BiomassaTNYResponse(BiomassaTNYCreate):
    id: str
    created_at: str
    created_by: Optional[str] = None

# PO Batubara Model
class POBatubaraCreate(BaseModel):
    district_code: Optional[str] = None
    district_name: Optional[str] = None
    periode: Optional[str] = None
    stock_code: Optional[float] = None
    warehouse: Optional[float] = None
    po_number: Optional[str] = None
    supplier_code: Optional[str] = None
    supplier_name: Optional[str] = None
    spec: Optional[str] = None
    vessel_tugboat: Optional[str] = None
    barge: Optional[str] = None
    no_jadwal: Optional[str] = None
    id_bbo_no_pengiriman: Optional[str] = None
    id_bbo_trans: Optional[str] = None
    no_shipment: Optional[str] = None
    time_arrival: Optional[str] = None
    completed: Optional[str] = None
    completed_year: Optional[int] = None
    completed_month: Optional[int] = None
    tonase_po: Optional[float] = None
    tonase_po_1000: Optional[float] = None
    inventory_price: Optional[float] = None
    freight_inventory_fob: Optional[float] = None
    total: Optional[float] = None

class POBatubaraResponse(POBatubaraCreate):
    id: str
    created_at: str
    created_by: Optional[str] = None

# Merit Order Model
class MeritOrderCreate(BaseModel):
    periode: Optional[str] = None
    periode_year: Optional[int] = None
    periode_month: Optional[int] = None
    pemasok: Optional[str] = None
    moda: Optional[str] = None  # Tongkang, Trucking, Vessel
    tipikal_kcal_kg: Optional[float] = None
    jenis_kontrak: Optional[str] = None  # CIF, CFR, FOB
    harga_batubara: Optional[float] = None  # RP/Ton
    harga_freight: Optional[float] = None  # RP/Ton
    harga_cif: Optional[float] = None  # RP/Ton
    rp_kg: Optional[float] = None
    rp_kcal: Optional[float] = None

class MeritOrderResponse(MeritOrderCreate):
    id: str
    created_at: str
    created_by: Optional[str] = None

# Dashboard Stats Model
class DashboardStats(BaseModel):
    total_vessel: int
    total_barge: int
    total_trucking: int
    total_biomassa: int
    total_tonase_batubara: float
    total_tonase_biomassa: float
    avg_gcv: float
    recent_shipments: List[dict]
    monthly_trend: List[dict]
    supplier_stats: List[dict]

# Advanced Dashboard Models
class DashboardAdvanced(BaseModel):
    # Contract Monitoring (Gauge)
    total_ds_mt: float
    total_tonase_po: float
    contract_percentage: float
    # Fuel Composition (Donut)
    fuel_composition: List[dict]
    # GCV Trend (Line Chart)
    gcv_trend: List[dict]
    # Supplier Economy Analysis (Bar Chart)
    supplier_economy: List[dict]
    # Slagging Risk Matrix (Heatmap)
    slagging_matrix: List[dict]
    # 6 Months Summary
    six_months_summary: List[dict]
    # Filter options
    available_periods: List[dict]
    available_moda: List[str]

# ==================== AUTH HELPERS ====================

def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')

def verify_password(password: str, hashed: str) -> bool:
    return bcrypt.checkpw(password.encode('utf-8'), hashed.encode('utf-8'))

def create_token(user_id: str, email: str, role: str) -> str:
    payload = {
        "user_id": user_id,
        "email": email,
        "role": role,
        "exp": datetime.now(timezone.utc) + timedelta(hours=JWT_EXPIRATION_HOURS)
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    try:
        token = credentials.credentials
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        user = await db.users.find_one({"id": payload["user_id"]}, {"_id": 0})
        if not user:
            raise HTTPException(status_code=401, detail="User tidak ditemukan")
        return user
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token sudah kadaluarsa")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Token tidak valid")

def require_role(allowed_roles: List[str]):
    async def role_checker(user: dict = Depends(get_current_user)):
        if user["role"] not in allowed_roles:
            raise HTTPException(status_code=403, detail="Akses ditolak")
        return user
    return role_checker

# ==================== AUTH ROUTES ====================

@api_router.post("/auth/register", response_model=TokenResponse)
async def register(data: UserCreate):
    existing = await db.users.find_one({"email": data.email})
    if existing:
        raise HTTPException(status_code=400, detail="Email sudah terdaftar")
    
    user_id = str(uuid.uuid4())
    user_doc = {
        "id": user_id,
        "email": data.email,
        "password": hash_password(data.password),
        "name": data.name,
        "role": data.role,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.users.insert_one(user_doc)
    
    token = create_token(user_id, data.email, data.role)
    return TokenResponse(
        access_token=token,
        user=UserResponse(
            id=user_id,
            email=data.email,
            name=data.name,
            role=data.role,
            created_at=user_doc["created_at"]
        )
    )

@api_router.post("/auth/login", response_model=TokenResponse)
async def login(data: UserLogin):
    user = await db.users.find_one({"email": data.email}, {"_id": 0})
    if not user or not verify_password(data.password, user["password"]):
        raise HTTPException(status_code=401, detail="Email atau password salah")
    
    token = create_token(user["id"], user["email"], user["role"])
    return TokenResponse(
        access_token=token,
        user=UserResponse(
            id=user["id"],
            email=user["email"],
            name=user["name"],
            role=user["role"],
            created_at=user["created_at"]
        )
    )

@api_router.get("/auth/me", response_model=UserResponse)
async def get_me(user: dict = Depends(get_current_user)):
    return UserResponse(
        id=user["id"],
        email=user["email"],
        name=user["name"],
        role=user["role"],
        created_at=user["created_at"]
    )

@api_router.get("/users", response_model=List[UserResponse])
async def get_users(user: dict = Depends(require_role(["admin"]))):
    users = await db.users.find({}, {"_id": 0, "password": 0}).to_list(1000)
    return [UserResponse(**u) for u in users]

# ==================== VESSEL TNY ROUTES ====================

@api_router.post("/vessels", response_model=VesselTNYResponse)
async def create_vessel(data: VesselTNYCreate, user: dict = Depends(require_role(["admin", "operator"]))):
    vessel_id = str(uuid.uuid4())
    vessel_doc = {
        "id": vessel_id,
        **data.model_dump(),
        "created_at": datetime.now(timezone.utc).isoformat(),
        "created_by": user["id"]
    }
    await db.vessels.insert_one(vessel_doc)
    vessel_doc.pop("_id", None)
    return VesselTNYResponse(**vessel_doc)

@api_router.get("/vessels")
async def get_vessels(
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=500),
    search: Optional[str] = None,
    user: dict = Depends(get_current_user)
):
    query = {}
    if search:
        query["$or"] = [
            {"shipment_code": {"$regex": search, "$options": "i"}},
            {"name_of_vessel": {"$regex": search, "$options": "i"}},
            {"suppliers": {"$regex": search, "$options": "i"}}
        ]
    
    # Server-side pagination
    skip = (page - 1) * page_size
    total = await db.vessels.count_documents(query)
    vessels = await db.vessels.find(query, {"_id": 0}).sort("time_arrival", -1).skip(skip).limit(page_size).to_list(page_size)
    
    return {
        "items": vessels,
        "total": total,
        "page": page,
        "page_size": page_size,
        "total_pages": (total + page_size - 1) // page_size
    }

@api_router.get("/vessels/{vessel_id}", response_model=VesselTNYResponse)
async def get_vessel(vessel_id: str, user: dict = Depends(get_current_user)):
    vessel = await db.vessels.find_one({"id": vessel_id}, {"_id": 0})
    if not vessel:
        raise HTTPException(status_code=404, detail="Data vessel tidak ditemukan")
    return VesselTNYResponse(**vessel)

@api_router.put("/vessels/{vessel_id}", response_model=VesselTNYResponse)
async def update_vessel(vessel_id: str, data: VesselTNYCreate, user: dict = Depends(require_role(["admin", "operator"]))):
    result = await db.vessels.update_one({"id": vessel_id}, {"$set": data.model_dump()})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Data vessel tidak ditemukan")
    vessel = await db.vessels.find_one({"id": vessel_id}, {"_id": 0})
    return VesselTNYResponse(**vessel)

@api_router.delete("/vessels/{vessel_id}")
async def delete_vessel(vessel_id: str, user: dict = Depends(require_role(["admin"]))):
    result = await db.vessels.delete_one({"id": vessel_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Data vessel tidak ditemukan")
    return {"message": "Data vessel berhasil dihapus"}

@api_router.delete("/vessels")
async def delete_all_vessels(user: dict = Depends(require_role(["admin"]))):
    result = await db.vessels.delete_many({})
    return {"message": f"Berhasil menghapus {result.deleted_count} data vessel", "count": result.deleted_count}

# ==================== BARGE TNY ROUTES ====================

@api_router.post("/barges", response_model=BargeTNYResponse)
async def create_barge(data: BargeTNYCreate, user: dict = Depends(require_role(["admin", "operator"]))):
    barge_id = str(uuid.uuid4())
    barge_doc = {
        "id": barge_id,
        **data.model_dump(),
        "created_at": datetime.now(timezone.utc).isoformat(),
        "created_by": user["id"]
    }
    await db.barges.insert_one(barge_doc)
    barge_doc.pop("_id", None)
    return BargeTNYResponse(**barge_doc)

@api_router.get("/barges")
async def get_barges(
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=500),
    search: Optional[str] = None,
    user: dict = Depends(get_current_user)
):
    query = {}
    if search:
        query["$or"] = [
            {"shipment_code": {"$regex": search, "$options": "i"}},
            {"name_of_barge": {"$regex": search, "$options": "i"}},
            {"suppliers": {"$regex": search, "$options": "i"}}
        ]
    
    skip = (page - 1) * page_size
    total = await db.barges.count_documents(query)
    barges = await db.barges.find(query, {"_id": 0}).sort("ta", -1).skip(skip).limit(page_size).to_list(page_size)
    
    return {
        "items": barges,
        "total": total,
        "page": page,
        "page_size": page_size,
        "total_pages": (total + page_size - 1) // page_size
    }

@api_router.get("/barges/{barge_id}", response_model=BargeTNYResponse)
async def get_barge(barge_id: str, user: dict = Depends(get_current_user)):
    barge = await db.barges.find_one({"id": barge_id}, {"_id": 0})
    if not barge:
        raise HTTPException(status_code=404, detail="Data barge tidak ditemukan")
    return BargeTNYResponse(**barge)

@api_router.put("/barges/{barge_id}", response_model=BargeTNYResponse)
async def update_barge(barge_id: str, data: BargeTNYCreate, user: dict = Depends(require_role(["admin", "operator"]))):
    result = await db.barges.update_one({"id": barge_id}, {"$set": data.model_dump()})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Data barge tidak ditemukan")
    barge = await db.barges.find_one({"id": barge_id}, {"_id": 0})
    return BargeTNYResponse(**barge)

@api_router.delete("/barges/{barge_id}")
async def delete_barge(barge_id: str, user: dict = Depends(require_role(["admin"]))):
    result = await db.barges.delete_one({"id": barge_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Data barge tidak ditemukan")
    return {"message": "Data barge berhasil dihapus"}

@api_router.delete("/barges")
async def delete_all_barges(user: dict = Depends(require_role(["admin"]))):
    result = await db.barges.delete_many({})
    return {"message": f"Berhasil menghapus {result.deleted_count} data barge", "count": result.deleted_count}

# ==================== TRUCKING TNY ROUTES ====================

@api_router.post("/trucking", response_model=TruckingTNYResponse)
async def create_trucking(data: TruckingTNYCreate, user: dict = Depends(require_role(["admin", "operator"]))):
    trucking_id = str(uuid.uuid4())
    trucking_doc = {
        "id": trucking_id,
        **data.model_dump(),
        "created_at": datetime.now(timezone.utc).isoformat(),
        "created_by": user["id"]
    }
    await db.trucking.insert_one(trucking_doc)
    trucking_doc.pop("_id", None)
    return TruckingTNYResponse(**trucking_doc)

@api_router.get("/trucking")
async def get_trucking(
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=500),
    search: Optional[str] = None,
    user: dict = Depends(get_current_user)
):
    query = {}
    if search:
        query["$or"] = [
            {"shipment_code": {"$regex": search, "$options": "i"}},
            {"no_truck": {"$regex": search, "$options": "i"}},
            {"suppliers": {"$regex": search, "$options": "i"}}
        ]
    
    skip = (page - 1) * page_size
    total = await db.trucking.count_documents(query)
    trucking_list = await db.trucking.find(query, {"_id": 0}).sort("ta", -1).skip(skip).limit(page_size).to_list(page_size)
    
    return {
        "items": trucking_list,
        "total": total,
        "page": page,
        "page_size": page_size,
        "total_pages": (total + page_size - 1) // page_size
    }

@api_router.get("/trucking/{trucking_id}", response_model=TruckingTNYResponse)
async def get_trucking_item(trucking_id: str, user: dict = Depends(get_current_user)):
    trucking = await db.trucking.find_one({"id": trucking_id}, {"_id": 0})
    if not trucking:
        raise HTTPException(status_code=404, detail="Data trucking tidak ditemukan")
    return TruckingTNYResponse(**trucking)

@api_router.put("/trucking/{trucking_id}", response_model=TruckingTNYResponse)
async def update_trucking(trucking_id: str, data: TruckingTNYCreate, user: dict = Depends(require_role(["admin", "operator"]))):
    result = await db.trucking.update_one({"id": trucking_id}, {"$set": data.model_dump()})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Data trucking tidak ditemukan")
    trucking = await db.trucking.find_one({"id": trucking_id}, {"_id": 0})
    return TruckingTNYResponse(**trucking)

@api_router.delete("/trucking/{trucking_id}")
async def delete_trucking(trucking_id: str, user: dict = Depends(require_role(["admin"]))):
    result = await db.trucking.delete_one({"id": trucking_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Data trucking tidak ditemukan")
    return {"message": "Data trucking berhasil dihapus"}

@api_router.delete("/trucking")
async def delete_all_trucking(user: dict = Depends(require_role(["admin"]))):
    result = await db.trucking.delete_many({})
    return {"message": f"Berhasil menghapus {result.deleted_count} data trucking", "count": result.deleted_count}

# ==================== BIOMASSA TNY ROUTES ====================

@api_router.post("/biomassa", response_model=BiomassaTNYResponse)
async def create_biomassa(data: BiomassaTNYCreate, user: dict = Depends(require_role(["admin", "operator"]))):
    biomassa_id = str(uuid.uuid4())
    biomassa_doc = {
        "id": biomassa_id,
        **data.model_dump(),
        "created_at": datetime.now(timezone.utc).isoformat(),
        "created_by": user["id"]
    }
    await db.biomassa.insert_one(biomassa_doc)
    biomassa_doc.pop("_id", None)
    return BiomassaTNYResponse(**biomassa_doc)

@api_router.get("/biomassa")
async def get_biomassa(
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=500),
    search: Optional[str] = None,
    user: dict = Depends(get_current_user)
):
    query = {}
    if search:
        query["$or"] = [
            {"shipment_code": {"$regex": search, "$options": "i"}},
            {"suppliers": {"$regex": search, "$options": "i"}},
            {"biomass_type": {"$regex": search, "$options": "i"}}
        ]
    
    skip = (page - 1) * page_size
    total = await db.biomassa.count_documents(query)
    biomassa_list = await db.biomassa.find(query, {"_id": 0}).sort("periode", -1).skip(skip).limit(page_size).to_list(page_size)
    
    return {
        "items": biomassa_list,
        "total": total,
        "page": page,
        "page_size": page_size,
        "total_pages": (total + page_size - 1) // page_size
    }

@api_router.get("/biomassa/{biomassa_id}", response_model=BiomassaTNYResponse)
async def get_biomassa_item(biomassa_id: str, user: dict = Depends(get_current_user)):
    biomassa = await db.biomassa.find_one({"id": biomassa_id}, {"_id": 0})
    if not biomassa:
        raise HTTPException(status_code=404, detail="Data biomassa tidak ditemukan")
    return BiomassaTNYResponse(**biomassa)

@api_router.put("/biomassa/{biomassa_id}", response_model=BiomassaTNYResponse)
async def update_biomassa(biomassa_id: str, data: BiomassaTNYCreate, user: dict = Depends(require_role(["admin", "operator"]))):
    result = await db.biomassa.update_one({"id": biomassa_id}, {"$set": data.model_dump()})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Data biomassa tidak ditemukan")
    biomassa = await db.biomassa.find_one({"id": biomassa_id}, {"_id": 0})
    return BiomassaTNYResponse(**biomassa)

@api_router.delete("/biomassa/{biomassa_id}")
async def delete_biomassa(biomassa_id: str, user: dict = Depends(require_role(["admin"]))):
    result = await db.biomassa.delete_one({"id": biomassa_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Data biomassa tidak ditemukan")
    return {"message": "Data biomassa berhasil dihapus"}

@api_router.delete("/biomassa")
async def delete_all_biomassa(user: dict = Depends(require_role(["admin"]))):
    result = await db.biomassa.delete_many({})
    return {"message": f"Berhasil menghapus {result.deleted_count} data biomassa", "count": result.deleted_count}

# ==================== PO BATUBARA ROUTES ====================

@api_router.get("/po-batubara")
async def get_po_batubara(
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=500),
    year: Optional[int] = None,
    month: Optional[int] = None,
    search: Optional[str] = None,
    user: dict = Depends(get_current_user)
):
    query = {}
    if year:
        query["completed_year"] = year
    if month:
        query["completed_month"] = month
    if search:
        query["$or"] = [
            {"po_number": {"$regex": search, "$options": "i"}},
            {"supplier_name": {"$regex": search, "$options": "i"}},
            {"no_shipment": {"$regex": search, "$options": "i"}}
        ]
    
    skip = (page - 1) * page_size
    total = await db.po_batubara.count_documents(query)
    po_list = await db.po_batubara.find(query, {"_id": 0}).sort("periode", -1).skip(skip).limit(page_size).to_list(page_size)
    
    return {
        "items": po_list,
        "total": total,
        "page": page,
        "page_size": page_size,
        "total_pages": (total + page_size - 1) // page_size
    }

@api_router.get("/po-batubara/years")
async def get_po_years(user: dict = Depends(get_current_user)):
    """Get list of available years with monthly summaries"""
    pipeline = [
        {"$match": {"completed_year": {"$ne": None}}},
        {"$group": {
            "_id": {"year": "$completed_year", "month": "$completed_month"},
            "count": {"$sum": 1},
            "total_tonase": {"$sum": "$tonase_po"},
            "total_value": {"$sum": "$total"}
        }},
        {"$sort": {"_id.year": -1, "_id.month": 1}}
    ]
    results = await db.po_batubara.aggregate(pipeline).to_list(1000)
    
    # Organize by year
    years_data = {}
    for r in results:
        year = r["_id"]["year"]
        month = r["_id"]["month"]
        if year not in years_data:
            years_data[year] = {"year": year, "months": {}, "total_count": 0, "total_tonase": 0, "total_value": 0}
        years_data[year]["months"][month] = {
            "month": month,
            "count": r["count"],
            "total_tonase": r["total_tonase"] or 0,
            "total_value": r["total_value"] or 0
        }
        years_data[year]["total_count"] += r["count"]
        years_data[year]["total_tonase"] += r["total_tonase"] or 0
        years_data[year]["total_value"] += r["total_value"] or 0
    
    return list(years_data.values())

@api_router.post("/po-batubara", response_model=POBatubaraResponse)
async def create_po_batubara(po: POBatubaraCreate, user: dict = Depends(require_role(["admin", "operator"]))):
    po_id = str(uuid.uuid4())
    po_doc = {
        "id": po_id,
        **po.model_dump(),
        "created_at": datetime.now(timezone.utc).isoformat(),
        "created_by": user["id"]
    }
    await db.po_batubara.insert_one(po_doc)
    return POBatubaraResponse(**{k: v for k, v in po_doc.items() if k != "_id"})

@api_router.get("/po-batubara/{po_id}", response_model=POBatubaraResponse)
async def get_po_batubara_by_id(po_id: str, user: dict = Depends(get_current_user)):
    po = await db.po_batubara.find_one({"id": po_id}, {"_id": 0})
    if not po:
        raise HTTPException(status_code=404, detail="Data PO tidak ditemukan")
    return POBatubaraResponse(**po)

@api_router.put("/po-batubara/{po_id}", response_model=POBatubaraResponse)
async def update_po_batubara(po_id: str, po: POBatubaraCreate, user: dict = Depends(require_role(["admin", "operator"]))):
    existing = await db.po_batubara.find_one({"id": po_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Data PO tidak ditemukan")
    await db.po_batubara.update_one({"id": po_id}, {"$set": po.model_dump()})
    updated = await db.po_batubara.find_one({"id": po_id}, {"_id": 0})
    return POBatubaraResponse(**updated)

@api_router.delete("/po-batubara/{po_id}")
async def delete_po_batubara(po_id: str, user: dict = Depends(require_role(["admin"]))):
    result = await db.po_batubara.delete_one({"id": po_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Data PO tidak ditemukan")
    return {"message": "Data PO berhasil dihapus"}

@api_router.delete("/po-batubara")
async def delete_all_po_batubara(user: dict = Depends(require_role(["admin"]))):
    result = await db.po_batubara.delete_many({})
    return {"message": f"Berhasil menghapus {result.deleted_count} data PO", "count": result.deleted_count}

@api_router.post("/upload/po-batubara")
async def upload_po_batubara_excel(file: UploadFile = File(...), user: dict = Depends(require_role(["admin", "operator"]))):
    try:
        contents = await file.read()
        df = pd.read_excel(io.BytesIO(contents))
        
        # Clean column names - remove newlines and extra spaces
        df.columns = df.columns.str.replace('\n', ' ').str.strip()
        
        def safe_float(val):
            if pd.isna(val) or val == '' or val is None:
                return None
            try:
                return float(val)
            except:
                return None
        
        def safe_str(val):
            if pd.isna(val) or val is None:
                return ""
            return str(val).strip()
        
        def parse_datetime(val):
            if pd.isna(val) or val is None:
                return None, None, None
            try:
                if isinstance(val, (int, float)):
                    # Excel serial date
                    dt = pd.to_datetime('1899-12-30') + pd.to_timedelta(val, 'D')
                elif isinstance(val, str):
                    dt = pd.to_datetime(val)
                else:
                    dt = val
                return str(dt), dt.year, dt.month
            except:
                return str(val), None, None
        
        def get_col(row, *possible_names):
            """Try multiple column names to get value"""
            for name in possible_names:
                val = row.get(name)
                if val is not None and not (isinstance(val, float) and pd.isna(val)):
                    return val
            return None
        
        records = []
        for _, row in df.iterrows():
            po_id = str(uuid.uuid4())
            completed_str, completed_year, completed_month = parse_datetime(get_col(row, "Completed"))
            
            po_doc = {
                "id": po_id,
                "district_code": safe_str(get_col(row, "District Code")),
                "district_name": safe_str(get_col(row, "District Name")),
                "periode": safe_str(get_col(row, "Periode")),
                "stock_code": safe_float(get_col(row, "Stock Code")),
                "warehouse": safe_float(get_col(row, "Warehouse")),
                "po_number": safe_str(get_col(row, "PO Number")),
                "supplier_code": safe_str(get_col(row, "Supplier Code")),
                "supplier_name": safe_str(get_col(row, "Supplier Name")),
                "spec": safe_str(get_col(row, "Spec")),
                "vessel_tugboat": safe_str(get_col(row, "Vessel / Tugboat", "Vessel/Tugboat")),
                "barge": safe_str(get_col(row, "Barge No", "Barge")),
                "no_jadwal": safe_str(get_col(row, "No Jadwal", "Jadwal Id BBO (No Pengiriman)")),
                "id_bbo_no_pengiriman": safe_str(get_col(row, "Id BBO (No Pengiriman)", "Jadwal Id BBO (No Pengiriman)")),
                "id_bbo_trans": safe_str(get_col(row, "Id BBO Trans")),
                "no_shipment": safe_str(get_col(row, "No Shipment")),
                "time_arrival": safe_str(get_col(row, "Time Arrival")),
                "completed": completed_str,
                "completed_year": completed_year,
                "completed_month": completed_month,
                "tonase_po": safe_float(get_col(row, "Tonase PO")),
                "tonase_po_1000": safe_float(get_col(row, "Tonase PO*1000")),
                "inventory_price": safe_float(get_col(row, "Inventory Price")),
                "freight_inventory_fob": safe_float(get_col(row, "Freight Inventory (FOB)", "Freight", "Inventory (FOB)")),
                "total": safe_float(get_col(row, "Total")),
                "created_at": datetime.now(timezone.utc).isoformat(),
                "created_by": user["id"]
            }
            records.append(po_doc)
        
        if records:
            await db.po_batubara.insert_many(records)
        
        return {"message": f"Berhasil mengimport {len(records)} data PO Batubara", "count": len(records)}
    except Exception as e:
        logger.error(f"Error uploading PO batubara excel: {e}")
        raise HTTPException(status_code=400, detail=f"Gagal memproses file: {str(e)}")

# ==================== MERIT ORDER ROUTES ====================

@api_router.get("/merit-order")
async def get_merit_orders(
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=500),
    year: Optional[int] = None,
    month: Optional[int] = None,
    search: Optional[str] = None,
    user: dict = Depends(get_current_user)
):
    query = {}
    if year:
        query["periode_year"] = year
    if month:
        query["periode_month"] = month
    if search:
        query["$or"] = [
            {"pemasok": {"$regex": search, "$options": "i"}},
            {"moda": {"$regex": search, "$options": "i"}},
            {"jenis_kontrak": {"$regex": search, "$options": "i"}}
        ]
    
    skip = (page - 1) * page_size
    total = await db.merit_order.count_documents(query)
    mo_list = await db.merit_order.find(query, {"_id": 0}).sort("periode", -1).skip(skip).limit(page_size).to_list(page_size)
    
    return {
        "items": mo_list,
        "total": total,
        "page": page,
        "page_size": page_size,
        "total_pages": (total + page_size - 1) // page_size
    }

@api_router.get("/merit-order/periods")
async def get_merit_order_periods(user: dict = Depends(get_current_user)):
    """Get list of available periods with summaries"""
    pipeline = [
        {"$match": {"periode_year": {"$ne": None}}},
        {"$group": {
            "_id": {"year": "$periode_year", "month": "$periode_month"},
            "count": {"$sum": 1},
            "avg_rp_kcal": {"$avg": "$rp_kcal"},
            "min_rp_kcal": {"$min": "$rp_kcal"},
            "max_rp_kcal": {"$max": "$rp_kcal"}
        }},
        {"$sort": {"_id.year": -1, "_id.month": -1}}
    ]
    results = await db.merit_order.aggregate(pipeline).to_list(1000)
    
    # Organize by year
    years_data = {}
    for r in results:
        year = r["_id"]["year"]
        month = r["_id"]["month"]
        if year not in years_data:
            years_data[year] = {"year": year, "months": {}, "total_count": 0}
        years_data[year]["months"][month] = {
            "month": month,
            "count": r["count"],
            "avg_rp_kcal": r["avg_rp_kcal"] or 0,
            "min_rp_kcal": r["min_rp_kcal"] or 0,
            "max_rp_kcal": r["max_rp_kcal"] or 0
        }
        years_data[year]["total_count"] += r["count"]
    
    return list(years_data.values())

@api_router.post("/merit-order", response_model=MeritOrderResponse)
async def create_merit_order(mo: MeritOrderCreate, user: dict = Depends(require_role(["admin", "operator"]))):
    mo_id = str(uuid.uuid4())
    mo_doc = {
        "id": mo_id,
        **mo.model_dump(),
        "created_at": datetime.now(timezone.utc).isoformat(),
        "created_by": user["id"]
    }
    await db.merit_order.insert_one(mo_doc)
    return MeritOrderResponse(**{k: v for k, v in mo_doc.items() if k != "_id"})

@api_router.get("/merit-order/{mo_id}", response_model=MeritOrderResponse)
async def get_merit_order_by_id(mo_id: str, user: dict = Depends(get_current_user)):
    mo = await db.merit_order.find_one({"id": mo_id}, {"_id": 0})
    if not mo:
        raise HTTPException(status_code=404, detail="Data Merit Order tidak ditemukan")
    return MeritOrderResponse(**mo)

@api_router.put("/merit-order/{mo_id}", response_model=MeritOrderResponse)
async def update_merit_order(mo_id: str, mo: MeritOrderCreate, user: dict = Depends(require_role(["admin", "operator"]))):
    existing = await db.merit_order.find_one({"id": mo_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Data Merit Order tidak ditemukan")
    await db.merit_order.update_one({"id": mo_id}, {"$set": mo.model_dump()})
    updated = await db.merit_order.find_one({"id": mo_id}, {"_id": 0})
    return MeritOrderResponse(**updated)

@api_router.delete("/merit-order/{mo_id}")
async def delete_merit_order(mo_id: str, user: dict = Depends(require_role(["admin"]))):
    result = await db.merit_order.delete_one({"id": mo_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Data Merit Order tidak ditemukan")
    return {"message": "Data Merit Order berhasil dihapus"}

@api_router.delete("/merit-order")
async def delete_all_merit_orders(user: dict = Depends(require_role(["admin"]))):
    result = await db.merit_order.delete_many({})
    return {"message": f"Berhasil menghapus {result.deleted_count} data Merit Order", "count": result.deleted_count}

@api_router.post("/upload/merit-order")
async def upload_merit_order_excel(file: UploadFile = File(...), user: dict = Depends(require_role(["admin", "operator"]))):
    try:
        contents = await file.read()
        df = pd.read_excel(io.BytesIO(contents))
        
        def safe_float(val):
            if pd.isna(val) or val == '' or val is None or val == '-':
                return None
            try:
                return float(val)
            except:
                return None
        
        def safe_str(val):
            if pd.isna(val) or val is None:
                return ""
            return str(val).strip()
        
        def parse_periode(val):
            if pd.isna(val) or val is None:
                return None, None, None
            try:
                if isinstance(val, str):
                    dt = pd.to_datetime(val)
                else:
                    dt = val
                return str(dt.date()), dt.year, dt.month
            except:
                return str(val), None, None
        
        records = []
        for _, row in df.iterrows():
            mo_id = str(uuid.uuid4())
            periode_str, periode_year, periode_month = parse_periode(row.get("Periode"))
            
            mo_doc = {
                "id": mo_id,
                "periode": periode_str,
                "periode_year": periode_year,
                "periode_month": periode_month,
                "pemasok": safe_str(row.get("Pemasok")),
                "moda": safe_str(row.get("Moda")),
                "tipikal_kcal_kg": safe_float(row.get("Tipikal (Kcal/Kg)")),
                "jenis_kontrak": safe_str(row.get("Jenis Kontrak")),
                "harga_batubara": safe_float(row.get("Harga Batubara (RP/Ton)")),
                "harga_freight": safe_float(row.get("Harga Freight (RP/Ton)")),
                "harga_cif": safe_float(row.get("Harga CIF(RP/Ton)")),
                "rp_kg": safe_float(row.get("RP/Kg")),
                "rp_kcal": safe_float(row.get("RP/Kcal")),
                "created_at": datetime.now(timezone.utc).isoformat(),
                "created_by": user["id"]
            }
            records.append(mo_doc)
        
        if records:
            await db.merit_order.insert_many(records)
        
        return {"message": f"Berhasil mengimport {len(records)} data Merit Order", "count": len(records)}
    except Exception as e:
        logger.error(f"Error uploading merit order excel: {e}")
        raise HTTPException(status_code=400, detail=f"Gagal memproses file: {str(e)}")

# ==================== EXCEL UPLOAD ROUTES ====================

@api_router.post("/upload/vessel")
async def upload_vessel_excel(file: UploadFile = File(...), user: dict = Depends(require_role(["admin", "operator"]))):
    try:
        contents = await file.read()
        df = pd.read_excel(io.BytesIO(contents))
        
        def safe_float(val):
            if pd.isna(val) or val == '' or val is None:
                return None
            try:
                return float(val)
            except:
                return None
        
        def safe_str(val):
            if pd.isna(val) or val is None:
                return ""
            return str(val).strip()
        
        records = []
        for _, row in df.iterrows():
            vessel_id = str(uuid.uuid4())
            vessel_doc = {
                "id": vessel_id,
                # Informasi Shipment
                "periode_ta": safe_str(row.get("Periode TA (Rakor)")),
                "periode_realisasi": safe_str(row.get("Periode Realisasi")),
                "shipment_code": safe_str(row.get("Shipment Code")),
                "voyage_code": safe_str(row.get("Voyage Code")),
                "suppliers": safe_str(row.get("Suppliers")),
                "voyage": safe_str(row.get("Voyage")),
                "name_of_vessel": safe_str(row.get("Name Of Vessel")),
                "coal_from": safe_str(row.get("Coal From")),
                # Waktu Operasional
                "time_arrival": safe_str(row.get("Time Arrival")),
                "berthed_time": safe_str(row.get("Berthed Time")),
                "commenced_unloading": safe_str(row.get("Commenced Unloading")),
                "completed_unloading": safe_str(row.get("Completed Unloading")),
                "durasi_pembongkaran_hari": safe_float(row.get("Durasi Pembongkaran (Hari)")),
                "durasi_pembongkaran_jam": safe_float(row.get("Durasi Pembongkaran (Jam)")),
                "waktu_tunggu_jam": safe_float(row.get("waktu tunggu (Jam)")),
                # Muatan
                "bl_mt": safe_float(row.get("B/L (MT)")),
                "ds_mt": safe_float(row.get("DS (MT)")),
                # COW
                "no_cow": safe_str(row.get("NO.COW")),
                "tgl_terbit_cow": safe_str(row.get("Tgl Terbit COW")),
                # Kualitas - GCV (handle multi-line headers)
                "gcv_arb": safe_float(row.get("GCV (Kcal/Kg)\nARB", row.get("GCV (Kcal/Kg) ARB"))),
                "gcv_adb": safe_float(row.get("GCV (Kcal/Kg)\nADB", row.get("GCV (Kcal/Kg) ADB"))),
                "gcv_db": safe_float(row.get("GCV (Kcal/Kg)\nDB", row.get("GCV (Kcal/Kg) DB"))),
                # Kualitas - Moisture
                "tm_arb": safe_float(row.get("TM (%wt)\nARB", row.get("TM (%wt) ARB"))),
                "im_adb": safe_float(row.get("IM (%wt) \nADB", row.get("IM (%wt) ADB"))),
                # Kualitas - Ash Content
                "ash_arb": safe_float(row.get("Ash \nContent (%wt) \nARB", row.get("Ash Content (%wt) ARB"))),
                "ash_adb": safe_float(row.get("Ash \nContent (%wt)\nADB", row.get("Ash Content (%wt) ADB"))),
                "ash_db": safe_float(row.get("Ash \nContent (%wt)\nDB", row.get("Ash Content (%wt) DB"))),
                # Kualitas - VM & FC
                "vm_arb": safe_float(row.get("VM (%wt)\nARB", row.get("VM (%wt) ARB"))),
                "vm_adb": safe_float(row.get("VM (%wt)\nADB", row.get("VM (%wt) ADB"))),
                "fc_arb": safe_float(row.get("FC (%wt)\nARB", row.get("FC (%wt) ARB"))),
                "fc_adb": safe_float(row.get("FC (%wt)\nADB", row.get("FC (%wt) ADB"))),
                # Kualitas - Sulfur
                "ts_arb": safe_float(row.get("Total Sulphur (%wt)\nARB", row.get("Total Sulphur (%wt) ARB"))),
                "ts_adb": safe_float(row.get("Total Sulphur (%wt)\nADB", row.get("Total Sulphur (%wt) ADB"))),
                "ts_db": safe_float(row.get("Total Sulphur (%wt)\nDB", row.get("Total Sulphur (%wt) DB"))),
                "ts_dafb": safe_float(row.get("Total Sulphur (%wt)\nDAFB", row.get("Total Sulphur (%wt) DAFB"))),
                # Ultimate Analysis
                "c_arb": safe_float(row.get("C (%wt)\nARB", row.get("C (%wt) ARB"))),
                "c_adb": safe_float(row.get("C (%wt)\nADB", row.get("C (%wt) ADB"))),
                "h_arb": safe_float(row.get("H (%wt)\nARB", row.get("H (%wt) ARB"))),
                "h_adb": safe_float(row.get("H (%wt)\nADB", row.get("H (%wt) ADB"))),
                "n_arb": safe_float(row.get("N (%wt)\nARB", row.get("N (%wt) ARB"))),
                "n_adb": safe_float(row.get("N (%wt)\nADB", row.get("N (%wt) ADB"))),
                "n_dafb": safe_float(row.get("N (%wt)\nDAFB", row.get("N (%wt) DAFB"))),
                "o_arb": safe_float(row.get("O (%wt)\nARB", row.get("O (%wt) ARB"))),
                "o_adb": safe_float(row.get("O (%wt)\nADB", row.get("O (%wt) ADB"))),
                # HGI & Index
                "hgi": safe_float(row.get("HGI (Point Index)")),
                "slagging_index": safe_str(row.get("Slagging (Index)")),
                "fouling_index": safe_str(row.get("Fouling (Index)")),
                "idt_reducing": safe_float(row.get("IDT Reducing (C°)")),
                # Ash Composition
                "sio2_db": safe_float(row.get("SiO2 (%DB)")),
                "al2o3_db": safe_float(row.get("Al2O3 (%DB)")),
                "tio2_db": safe_float(row.get("TiO2 (%DB)")),
                "fe2o3_db": safe_float(row.get("Fe2O3 (%DB)")),
                "cao_db": safe_float(row.get("CaO (%DB)")),
                "mgo_db": safe_float(row.get("MgO (%DB)")),
                "k2o_db": safe_float(row.get("K2O (%DB)")),
                "na2o_db": safe_float(row.get("Na2O (%DB)")),
                "so3_db": safe_float(row.get("SO3 (%DB)")),
                "p2o5_db": safe_float(row.get("P2O5 (%DB)")),
                "mno2_db": safe_float(row.get("MnO2 (%DB)")),
                "mn3o4_db": safe_float(row.get("Mn3O4 (%DB)")),
                # Size Analysis
                "size_70mm": safe_float(row.get("< 70 mm (%wt)")),
                "size_50mm": safe_float(row.get("< 50 mm (%wt)")),
                "size_32mm": safe_float(row.get("< 32 mm (%wt)")),
                "size_2_38mm": safe_float(row.get("< 2,38 mm (%wt)")),
                # COA
                "no_coa": safe_str(row.get("NO. COA")),
                "tgl_terbit_coa": safe_str(row.get("Tgl Terbit COA")),
                "durasi_terbit_coa": safe_str(row.get("DURASI TERBIT COA")),
                # Metadata
                "created_at": datetime.now(timezone.utc).isoformat(),
                "created_by": user["id"]
            }
            records.append(vessel_doc)
        
        if records:
            await db.vessels.insert_many(records)
        
        return {"message": f"Berhasil mengimport {len(records)} data vessel", "count": len(records)}
    except Exception as e:
        logger.error(f"Error uploading vessel excel: {e}")
        raise HTTPException(status_code=400, detail=f"Gagal memproses file: {str(e)}")

@api_router.post("/upload/barge")
async def upload_barge_excel(file: UploadFile = File(...), user: dict = Depends(require_role(["admin", "operator"]))):
    try:
        contents = await file.read()
        df = pd.read_excel(io.BytesIO(contents))
        
        def safe_float(val):
            if pd.isna(val) or val == '' or val is None:
                return None
            try:
                return float(val)
            except:
                return None
        
        def safe_str(val):
            if pd.isna(val) or val is None:
                return ""
            return str(val).strip()
        
        records = []
        for _, row in df.iterrows():
            barge_id = str(uuid.uuid4())
            barge_doc = {
                "id": barge_id,
                # Informasi Shipment
                "periode": safe_str(row.get("Periode")),
                "shipment_code": safe_str(row.get("Shipment Code")),
                "voyage_code": safe_str(row.get("Voyage Code")),
                "shipment": safe_str(row.get("Shipment")),
                "suppliers": safe_str(row.get("Suppliers")),
                "voyage": safe_str(row.get("Voyage")),
                "tb": safe_str(row.get("TB")),
                "bg": safe_str(row.get("BG")),
                "coal_from": safe_str(row.get("Coal From")),
                # Waktu Operasional
                "ta": safe_str(row.get("TA")),
                "berthed_time": safe_str(row.get("Berthed Time")),
                "commenced_unloading": safe_str(row.get("Commenced Unloading")),
                "completed_unloading": safe_str(row.get("Completed Unloading")),
                "durasi_pembongkaran_hari": safe_float(row.get("Durasi Pembongkaran (Hari)")),
                "durasi_pembongkaran_jam": safe_float(row.get("Durasi Pembongkaran (Jam)")),
                "waktu_tunggu_jam": safe_float(row.get("waktu tunggu (Jam)")),
                # Muatan
                "bl_mt": safe_float(row.get("B/L (MT)")),
                "ds_mt": safe_float(row.get("DS (MT)")),
                # COW
                "no_cow": safe_str(row.get("NO.COW")),
                "tgl_terbit_cow": safe_str(row.get("Tgl Terbit COW")),
                # Kualitas - GCV
                "gcv_arb": safe_float(row.get("GCV (Kcal/Kg)\nARB", row.get("GCV (Kcal/Kg) ARB"))),
                "gcv_adb": safe_float(row.get("GCV (Kcal/Kg)\nADB", row.get("GCV (Kcal/Kg) ADB"))),
                "gcv_db": safe_float(row.get("GCV (Kcal/Kg)\nDB", row.get("GCV (Kcal/Kg) DB"))),
                # Kualitas - Moisture
                "tm_arb": safe_float(row.get("TM (%wt)\nARB", row.get("TM (%wt) ARB"))),
                "im_adb": safe_float(row.get("IM (%wt) \nADB", row.get("IM (%wt) ADB"))),
                # Kualitas - Ash Content
                "ash_arb": safe_float(row.get("Ash \nContent (%wt) \nARB", row.get("Ash Content (%wt) ARB"))),
                "ash_adb": safe_float(row.get("Ash \nContent (%wt)\nADB", row.get("Ash Content (%wt) ADB"))),
                "ash_db": safe_float(row.get("Ash \nContent (%wt)\nDB", row.get("Ash Content (%wt) DB"))),
                # Kualitas - VM & FC
                "vm_arb": safe_float(row.get("VM (%wt)\nARB")), "vm_adb": safe_float(row.get("VM (%wt)\nADB")),
                "fc_arb": safe_float(row.get("FC (%wt)\nARB")), "fc_adb": safe_float(row.get("FC (%wt)\nADB")),
                # Kualitas - Sulfur
                "ts_arb": safe_float(row.get("Total Sulphur (%wt)\nARB")),
                "ts_adb": safe_float(row.get("Total Sulphur (%wt)\nADB")),
                "ts_db": safe_float(row.get("Total Sulphur (%wt)\nDB")),
                "ts_dafb": safe_float(row.get("Total Sulphur (%wt)\nDAFB")),
                # Ultimate Analysis
                "c_arb": safe_float(row.get("C (%wt)\nARB")), "c_adb": safe_float(row.get("C (%wt)\nADB")),
                "h_arb": safe_float(row.get("H (%wt)\nARB")), "h_adb": safe_float(row.get("H (%wt)\nADB")),
                "n_arb": safe_float(row.get("N (%wt)\nARB")), "n_adb": safe_float(row.get("N (%wt)\nADB")),
                "n_dafb": safe_float(row.get("N (%wt)\nDAFB")),
                "o_arb": safe_float(row.get("O (%wt)\nARB")), "o_adb": safe_float(row.get("O (%wt)\nADB")),
                # HGI & Index
                "hgi": safe_float(row.get("HGI (Point Index)")),
                "slagging_index": safe_str(row.get("Slagging (Index)")),
                "fouling_index": safe_str(row.get("Fouling (Index)")),
                "idt_reducing": safe_float(row.get("IDT Reducing (C°)")),
                # Ash Composition
                "sio2_db": safe_float(row.get("SiO2 (%DB)")), "al2o3_db": safe_float(row.get("Al2O3 (%DB)")),
                "tio2_db": safe_float(row.get("TiO2 (%DB)")), "fe2o3_db": safe_float(row.get("Fe2O3 (%DB)")),
                "cao_db": safe_float(row.get("CaO (%DB)")), "mgo_db": safe_float(row.get("MgO (%DB)")),
                "k2o_db": safe_float(row.get("K2O (%DB)")), "na2o_db": safe_float(row.get("Na2O (%DB)")),
                "so3_db": safe_float(row.get("SO3 (%DB)")), "p2o5_db": safe_float(row.get("P2O5 (%DB)")),
                "mno2_db": safe_float(row.get("MnO2 (%DB)")), "mn3o4_db": safe_float(row.get("Mn3O4 (%DB)")),
                # Size Analysis
                "size_70mm": safe_float(row.get("< 70 mm (%wt)")), "size_50mm": safe_float(row.get("< 50 mm (%wt)")),
                "size_32mm": safe_float(row.get("< 32 mm (%wt)")), "size_2_38mm": safe_float(row.get("< 2,38 mm (%wt)")),
                # COA
                "no_coa": safe_str(row.get("NO. COA")),
                "tgl_terbit_coa": safe_str(row.get("Tgl Terbit COA")),
                "durasi_terbit_coa": safe_str(row.get("DURASI TERBIT COA")),
                # Metadata
                "created_at": datetime.now(timezone.utc).isoformat(),
                "created_by": user["id"]
            }
            records.append(barge_doc)
        
        if records:
            await db.barges.insert_many(records)
        
        return {"message": f"Berhasil mengimport {len(records)} data barge", "count": len(records)}
    except Exception as e:
        logger.error(f"Error uploading barge excel: {e}")
        raise HTTPException(status_code=400, detail=f"Gagal memproses file: {str(e)}")

@api_router.post("/upload/trucking")
async def upload_trucking_excel(file: UploadFile = File(...), user: dict = Depends(require_role(["admin", "operator"]))):
    try:
        contents = await file.read()
        df = pd.read_excel(io.BytesIO(contents))
        
        def safe_float(val):
            if pd.isna(val) or val == '' or val is None:
                return None
            try:
                return float(val)
            except:
                return None
        
        def safe_str(val):
            if pd.isna(val) or val is None:
                return ""
            return str(val).strip()
        
        def safe_int(val):
            if pd.isna(val) or val == '' or val is None:
                return None
            try:
                return int(float(val))
            except:
                return None
        
        records = []
        for _, row in df.iterrows():
            trucking_id = str(uuid.uuid4())
            trucking_doc = {
                "id": trucking_id,
                # Informasi Shipment
                "periode_ta": safe_str(row.get("Periode TA (Rakor)")),
                "periode_realisasi": safe_str(row.get("Periode Realisasi")),
                "shipment_code": safe_str(row.get("Shipment Code")),
                "voyage_code": safe_str(row.get("Voyage Code")),
                "shipment": safe_str(row.get("Shipment")),
                "suppliers": safe_str(row.get("Suppliers")),
                "transportasi": safe_str(row.get("Transportasi")),
                "coal_from": safe_str(row.get("Coal From")),
                # Waktu Operasional
                "ta": safe_str(row.get("TA")),
                "berthed_time": safe_str(row.get("Berthed Time")),
                "commenced_unloading": safe_str(row.get("Commenced Unloading")),
                "completed_unloading": safe_str(row.get("Completed Unloading")),
                "durasi_pembongkaran_hari": safe_float(row.get("Durasi Pembongkaran (Hari)")),
                "durasi_pembongkaran_jam": safe_float(row.get("Durasi Pembongkaran (Jam)")),
                # Muatan
                "bl_mt": safe_float(row.get("B/L (MT)")),
                "ds_mt": safe_float(row.get("DS (MT)")),
                "rit": safe_int(row.get("RIT")),
                # COW
                "no_cow": safe_str(row.get("NO.COW")),
                "tgl_terbit_cow": safe_str(row.get("Tgl Terbit COW")),
                # Kualitas - GCV
                "gcv_arb": safe_float(row.get("GCV (Kcal/Kg)\nARB", row.get("GCV (Kcal/Kg) ARB"))),
                "gcv_adb": safe_float(row.get("GCV (Kcal/Kg)\nADB", row.get("GCV (Kcal/Kg) ADB"))),
                "gcv_db": safe_float(row.get("GCV (Kcal/Kg)\nDB", row.get("GCV (Kcal/Kg) DB"))),
                # Kualitas - Moisture
                "tm_arb": safe_float(row.get("TM (%wt)\nARB", row.get("TM (%wt) ARB"))),
                "im_adb": safe_float(row.get("IM (%wt) \nADB", row.get("IM (%wt) ADB"))),
                # Kualitas - Ash Content
                "ash_arb": safe_float(row.get("Ash \nContent (%wt) \nARB")),
                "ash_adb": safe_float(row.get("Ash \nContent (%wt)\nADB")),
                "ash_db": safe_float(row.get("Ash \nContent (%wt)\nDB")),
                # Kualitas - VM & FC
                "vm_arb": safe_float(row.get("VM (%wt)\nARB")), "vm_adb": safe_float(row.get("VM (%wt)\nADB")),
                "fc_arb": safe_float(row.get("FC (%wt)\nARB")), "fc_adb": safe_float(row.get("FC (%wt)\nADB")),
                # Kualitas - Sulfur
                "ts_arb": safe_float(row.get("Total Sulphur (%wt)\nARB")),
                "ts_adb": safe_float(row.get("Total Sulphur (%wt)\nADB")),
                "ts_db": safe_float(row.get("Total Sulphur (%wt)\nDB")),
                "ts_dafb": safe_float(row.get("Total Sulphur (%wt)\nDAFB")),
                # Ultimate Analysis
                "c_arb": safe_float(row.get("C (%wt)\nARB")), "c_adb": safe_float(row.get("C (%wt)\nADB")),
                "h_arb": safe_float(row.get("H (%wt)\nARB")), "h_adb": safe_float(row.get("H (%wt)\nADB")),
                "n_arb": safe_float(row.get("N (%wt)\nARB")), "n_adb": safe_float(row.get("N (%wt)\nADB")),
                "n_dafb": safe_float(row.get("N (%wt)\nDAFB")),
                "o_arb": safe_float(row.get("O (%wt)\nARB")), "o_adb": safe_float(row.get("O (%wt)\nADB")),
                # HGI & Index
                "hgi": safe_float(row.get("HGI (Point Index)")),
                "slagging_index": safe_str(row.get("Slagging (Index)")),
                "fouling_index": safe_str(row.get("Fouling (Index)")),
                "idt_reducing": safe_float(row.get("IDT Reducing (C°)")),
                # Ash Composition
                "sio2_db": safe_float(row.get("SiO2 (%DB)")), "al2o3_db": safe_float(row.get("Al2O3 (%DB)")),
                "tio2_db": safe_float(row.get("TiO2 (%DB)")), "fe2o3_db": safe_float(row.get("Fe2O3 (%DB)")),
                "cao_db": safe_float(row.get("CaO (%DB)")), "mgo_db": safe_float(row.get("MgO (%DB)")),
                "k2o_db": safe_float(row.get("K2O (%DB)")), "na2o_db": safe_float(row.get("Na2O (%DB)")),
                "so3_db": safe_float(row.get("SO3 (%DB)")), "p2o5_db": safe_float(row.get("P2O5 (%DB)")),
                "mno2_db": safe_float(row.get("MnO2 (%DB)")), "mn3o4_db": safe_float(row.get("Mn3O4 (%DB)")),
                # Size Analysis
                "size_70mm": safe_float(row.get("< 70 mm (%wt)")), "size_50mm": safe_float(row.get("< 50 mm (%wt)")),
                "size_32mm": safe_float(row.get("< 32 mm (%wt)")), "size_2_38mm": safe_float(row.get("< 2,38 mm (%wt)")),
                # COA
                "no_coa": safe_str(row.get("NO. COA")),
                "tgl_terbit_coa": safe_str(row.get("Tgl Terbit COA")),
                # Metadata
                "created_at": datetime.now(timezone.utc).isoformat(),
                "created_by": user["id"]
            }
            records.append(trucking_doc)
        
        if records:
            await db.trucking.insert_many(records)
        
        return {"message": f"Berhasil mengimport {len(records)} data trucking", "count": len(records)}
    except Exception as e:
        logger.error(f"Error uploading trucking excel: {e}")
        raise HTTPException(status_code=400, detail=f"Gagal memproses file: {str(e)}")

@api_router.post("/upload/biomassa")
async def upload_biomassa_excel(file: UploadFile = File(...), user: dict = Depends(require_role(["admin", "operator"]))):
    try:
        contents = await file.read()
        df = pd.read_excel(io.BytesIO(contents))
        
        # Clean column names - remove newlines and extra spaces
        df.columns = df.columns.str.replace('\n', ' ').str.strip()
        
        def safe_float(val):
            if pd.isna(val) or val == '' or val is None:
                return None
            try:
                return float(val)
            except:
                return None
        
        def safe_str(val):
            if pd.isna(val) or val is None:
                return ""
            return str(val).strip()
        
        records = []
        for _, row in df.iterrows():
            biomassa_id = str(uuid.uuid4())
            biomassa_doc = {
                "id": biomassa_id,
                # Informasi Shipment
                "periode": safe_str(row.get("Periode")),
                "shipment_code": safe_str(row.get("Shipment Code")),
                "voyage_code": safe_str(row.get("Voyage Code")),
                "lot": safe_str(row.get("Lot")),
                "suppliers": safe_str(row.get("Suppliers")),
                "shipper": safe_str(row.get("Shipper")),
                "lot_1": safe_str(row.get("Lot.1")),
                "tb": safe_str(row.get("TB")),
                "bg": safe_str(row.get("BG")),
                "biomass_type": safe_str(row.get("Biomass", row.get("Biomass "))),
                # Waktu Operasional
                "ta": safe_str(row.get("TA")),
                "berthed_time": safe_str(row.get("Berthed Time")),
                "commenced_unloading": safe_str(row.get("Commenced Unloading")),
                "completed_unloading": safe_str(row.get("Completed Unloading")),
                "durasi_pembongkaran_hari": safe_float(row.get("Durasi Pembongkaran (Hari)")),
                # Muatan
                "bl_mt": safe_float(row.get("B/L (MT)")),
                "jembatan_timbang_mt": safe_float(row.get("Jembatan Timbang (MT)")),
                "surveyor_unloading": safe_str(row.get("Surveyor Unloading", row.get(" Surveyor Unloading "))),
                # COW/ROW
                "no_cow_row": safe_str(row.get("NO.COW / ROW")),
                "tgl_terbit_cow": safe_str(row.get("Tgl Terbit COW")),
                "lama_terbit_row": safe_float(row.get("Lama terbit Row")),
                # Kualitas - handle column names with newlines converted to spaces
                "gcv_arb": safe_float(row.get("GCV (Kcal/Kg) ARB")),
                "gcv_adb": safe_float(row.get("GCV (Kcal/Kg) ADB")),
                "tm_arb": safe_float(row.get("TM (%wt) ARB")),
                "im_adb": safe_float(row.get("IM (%wt)  ADB", row.get("IM (%wt) ADB"))),
                # COA
                "no_coa": safe_str(row.get("NO. COA")),
                "tgl_terbit_coa": safe_str(row.get("Tgl Terbit COA")),
                "durasi_pembongkaran_hari_2": safe_float(row.get("Durasi Pembongkaran (Hari).1")),
                "waktu_tunggu_jam": safe_float(row.get("Waktu Tunggu (Jam)")),
                "durasi_terbit_coa": safe_float(row.get("DURASI TERBIT COA")),
                # Metadata
                "created_at": datetime.now(timezone.utc).isoformat(),
                "created_by": user["id"]
            }
            records.append(biomassa_doc)
        
        if records:
            await db.biomassa.insert_many(records)
        
        return {"message": f"Berhasil mengimport {len(records)} data biomassa", "count": len(records)}
    except Exception as e:
        logger.error(f"Error uploading biomassa excel: {e}")
        raise HTTPException(status_code=400, detail=f"Gagal memproses file: {str(e)}")

# ==================== DASHBOARD STATS ====================

@api_router.get("/dashboard/stats", response_model=DashboardStats)
async def get_dashboard_stats(user: dict = Depends(get_current_user)):
    total_vessel = await db.vessels.count_documents({})
    total_barge = await db.barges.count_documents({})
    total_trucking = await db.trucking.count_documents({})
    total_biomassa = await db.biomassa.count_documents({})
    
    # Calculate total tonase batubara from vessels and barges
    vessel_pipeline = [{"$group": {"_id": None, "total": {"$sum": {"$ifNull": ["$ds_mt", 0]}}}}]
    barge_pipeline = [{"$group": {"_id": None, "total": {"$sum": {"$ifNull": ["$ds_mt", 0]}}}}]
    biomassa_pipeline = [{"$group": {"_id": None, "total": {"$sum": {"$ifNull": ["$jembatan_timbang_mt", 0]}}}}]
    
    vessel_tonase = await db.vessels.aggregate(vessel_pipeline).to_list(1)
    barge_tonase = await db.barges.aggregate(barge_pipeline).to_list(1)
    biomassa_tonase = await db.biomassa.aggregate(biomassa_pipeline).to_list(1)
    
    total_tonase_batubara = (vessel_tonase[0]["total"] if vessel_tonase else 0) + (barge_tonase[0]["total"] if barge_tonase else 0)
    total_tonase_biomassa = biomassa_tonase[0]["total"] if biomassa_tonase else 0
    
    # Average GCV
    gcv_pipeline = [{"$match": {"gcv_arb": {"$ne": None}}}, {"$group": {"_id": None, "avg": {"$avg": "$gcv_arb"}}}]
    vessel_gcv = await db.vessels.aggregate(gcv_pipeline).to_list(1)
    avg_gcv = vessel_gcv[0]["avg"] if vessel_gcv else 0
    
    # Recent shipments
    recent_vessels = await db.vessels.find({}, {"_id": 0}).sort("created_at", -1).limit(5).to_list(5)
    recent_barges = await db.barges.find({}, {"_id": 0}).sort("created_at", -1).limit(5).to_list(5)
    recent_shipments = []
    for v in recent_vessels:
        recent_shipments.append({"type": "vessel", "name": v.get("name_of_vessel", ""), "code": v.get("shipment_code", ""), "date": v.get("created_at", "")})
    for b in recent_barges:
        recent_shipments.append({"type": "barge", "name": b.get("name_of_barge", ""), "code": b.get("shipment_code", ""), "date": b.get("created_at", "")})
    recent_shipments = sorted(recent_shipments, key=lambda x: x.get("date", ""), reverse=True)[:5]
    
    # Monthly trend (simplified)
    monthly_trend = [
        {"month": "Jan", "vessel": 5, "barge": 3, "trucking": 10, "biomassa": 8},
        {"month": "Feb", "vessel": 4, "barge": 5, "trucking": 12, "biomassa": 6},
        {"month": "Mar", "vessel": 6, "barge": 4, "trucking": 8, "biomassa": 10},
        {"month": "Apr", "vessel": 3, "barge": 6, "trucking": 15, "biomassa": 7},
        {"month": "May", "vessel": 7, "barge": 2, "trucking": 11, "biomassa": 9},
        {"month": "Jun", "vessel": 5, "barge": 4, "trucking": 9, "biomassa": 12}
    ]
    
    # Supplier stats
    supplier_pipeline = [
        {"$group": {"_id": "$suppliers", "count": {"$sum": 1}, "total_tonase": {"$sum": {"$ifNull": ["$ds_mt", 0]}}}},
        {"$sort": {"total_tonase": -1}},
        {"$limit": 10}
    ]
    supplier_stats_raw = await db.vessels.aggregate(supplier_pipeline).to_list(10)
    supplier_stats = [{"supplier": s["_id"], "count": s["count"], "tonase": s["total_tonase"]} for s in supplier_stats_raw if s["_id"]]
    
    return DashboardStats(
        total_vessel=total_vessel,
        total_barge=total_barge,
        total_trucking=total_trucking,
        total_biomassa=total_biomassa,
        total_tonase_batubara=total_tonase_batubara,
        total_tonase_biomassa=total_tonase_biomassa,
        avg_gcv=avg_gcv or 0,
        recent_shipments=recent_shipments,
        monthly_trend=monthly_trend,
        supplier_stats=supplier_stats
    )

@api_router.get("/dashboard/advanced")
async def get_dashboard_advanced(
    year: Optional[int] = None,
    month: Optional[int] = None,
    moda: Optional[str] = None,
    user: dict = Depends(get_current_user)
):
    """Advanced dashboard data with filters"""
    from datetime import date
    from dateutil.relativedelta import relativedelta
    
    # Calculate 6 months ago
    today = date.today()
    six_months_ago = today - relativedelta(months=6)
    
    # Build date filter based on completed_unloading field
    date_filter = {}
    if year and month:
        # Filter specific month
        start_date = f"{year}-{month:02d}-01"
        if month == 12:
            end_date = f"{year+1}-01-01"
        else:
            end_date = f"{year}-{month+1:02d}-01"
        date_filter = {"completed_unloading": {"$gte": start_date, "$lt": end_date}}
    
    # Moda filter for trucking
    moda_filter = {}
    
    # === 1. Contract Monitoring (Gauge) ===
    # Total DS MT from all sources
    vessel_ds = await db.vessels.aggregate([
        {"$group": {"_id": None, "total": {"$sum": {"$ifNull": ["$ds_mt", 0]}}}}
    ]).to_list(1)
    barge_ds = await db.barges.aggregate([
        {"$group": {"_id": None, "total": {"$sum": {"$ifNull": ["$ds_mt", 0]}}}}
    ]).to_list(1)
    trucking_ds = await db.trucking.aggregate([
        {"$group": {"_id": None, "total": {"$sum": {"$ifNull": ["$ds_mt", 0]}}}}
    ]).to_list(1)
    biomassa_ds = await db.biomassa.aggregate([
        {"$group": {"_id": None, "total": {"$sum": {"$ifNull": ["$jembatan_timbang_mt", 0]}}}}
    ]).to_list(1)
    
    total_ds_mt = (
        (vessel_ds[0]["total"] if vessel_ds else 0) +
        (barge_ds[0]["total"] if barge_ds else 0) +
        (trucking_ds[0]["total"] if trucking_ds else 0) +
        (biomassa_ds[0]["total"] if biomassa_ds else 0)
    )
    
    # Total Tonase PO from PO Batubara
    po_tonase = await db.po_batubara.aggregate([
        {"$group": {"_id": None, "total": {"$sum": {"$ifNull": ["$tonase_po", 0]}}}}
    ]).to_list(1)
    total_tonase_po = po_tonase[0]["total"] if po_tonase else 0
    
    contract_percentage = (total_ds_mt / total_tonase_po * 100) if total_tonase_po > 0 else 0
    
    # === 2. Fuel Composition (Donut) - Based on PO Batubara spec (LRC/MRC) and Biomassa types ===
    # Get coal spec (LRC, MRC) from PO Batubara with tonase
    po_spec_data = await db.po_batubara.aggregate([
        {"$match": {"spec": {"$ne": None}, "tonase_po": {"$ne": None}}},
        {"$group": {"_id": "$spec", "total_tonase": {"$sum": {"$ifNull": ["$tonase_po", 0]}}, "count": {"$sum": 1}}}
    ]).to_list(100)
    
    # Get biomass types from biomassa collection
    biomass_types = await db.biomassa.aggregate([
        {"$match": {"biomass_type": {"$ne": None}}},
        {"$group": {"_id": "$biomass_type", "total_tonase": {"$sum": {"$ifNull": ["$jembatan_timbang_mt", 0]}}, "count": {"$sum": 1}}}
    ]).to_list(100)
    
    fuel_composition = []
    
    # Add coal types (LRC, MRC) from PO
    for item in po_spec_data:
        spec_name = item["_id"]
        if spec_name:
            fuel_composition.append({
                "name": f"Batubara {spec_name}",
                "value": item["total_tonase"],
                "count": item["count"],
                "type": "coal"
            })
    
    # Add biomass types
    for item in biomass_types:
        btype = item["_id"]
        if btype:
            fuel_composition.append({
                "name": f"Biomassa {btype}",
                "value": item["total_tonase"],
                "count": item["count"],
                "type": "biomass"
            })
    
    # Sort by value descending
    fuel_composition.sort(key=lambda x: x["value"], reverse=True)
    
    # === 3. GCV Trend (Line Chart) ===
    # Get GCV data with dates from vessels and barges
    gcv_trend = []
    vessel_gcv_data = await db.vessels.find(
        {"gcv_arb": {"$ne": None}, "completed_unloading": {"$ne": None}},
        {"_id": 0, "completed_unloading": 1, "gcv_arb": 1, "name_of_vessel": 1}
    ).sort("completed_unloading", 1).to_list(500)
    
    barge_gcv_data = await db.barges.find(
        {"gcv_arb": {"$ne": None}, "completed_unloading": {"$ne": None}},
        {"_id": 0, "completed_unloading": 1, "gcv_arb": 1}
    ).sort("completed_unloading", 1).to_list(500)
    
    # Combine and sort
    all_gcv = []
    for v in vessel_gcv_data:
        try:
            date_str = str(v.get("completed_unloading", ""))[:10]
            all_gcv.append({"date": date_str, "gcv": v["gcv_arb"], "source": "vessel"})
        except:
            pass
    for b in barge_gcv_data:
        try:
            date_str = str(b.get("completed_unloading", ""))[:10]
            all_gcv.append({"date": date_str, "gcv": b["gcv_arb"], "source": "barge"})
        except:
            pass
    
    # Group by date and calculate average
    gcv_by_date = {}
    for item in all_gcv:
        d = item["date"]
        if d not in gcv_by_date:
            gcv_by_date[d] = []
        gcv_by_date[d].append(item["gcv"])
    
    for d, values in sorted(gcv_by_date.items()):
        gcv_trend.append({
            "date": d,
            "gcv_avg": sum(values) / len(values),
            "count": len(values),
            "target": 4000
        })
    
    # Limit to last 50 data points
    gcv_trend = gcv_trend[-50:]
    
    # === 4. Supplier Economy Analysis (Bar Chart) ===
    # Calculate Actual Rp/Kcal = (Harga CIF / GCV ARB) / 1000
    supplier_economy = []
    
    # Get merit order data with harga_cif
    merit_orders = await db.merit_order.find(
        {"harga_cif": {"$ne": None}, "pemasok": {"$ne": None}},
        {"_id": 0, "pemasok": 1, "harga_cif": 1, "rp_kcal": 1}
    ).to_list(1000)
    
    # Get GCV by supplier from vessels
    supplier_gcv = {}
    vessels_data = await db.vessels.find(
        {"gcv_arb": {"$ne": None}, "suppliers": {"$ne": None}},
        {"_id": 0, "suppliers": 1, "gcv_arb": 1}
    ).to_list(1000)
    
    for v in vessels_data:
        sup = v["suppliers"]
        if sup not in supplier_gcv:
            supplier_gcv[sup] = []
        supplier_gcv[sup].append(v["gcv_arb"])
    
    # Calculate Rp/Kcal for each supplier in merit order
    supplier_rp_kcal = {}
    for mo in merit_orders:
        sup = mo["pemasok"]
        if mo.get("rp_kcal"):
            if sup not in supplier_rp_kcal:
                supplier_rp_kcal[sup] = []
            supplier_rp_kcal[sup].append(mo["rp_kcal"])
    
    # Average and sort
    for sup, values in supplier_rp_kcal.items():
        avg_rp_kcal = sum(values) / len(values)
        supplier_economy.append({
            "supplier": sup[:30] + "..." if len(sup) > 30 else sup,
            "full_name": sup,
            "rp_kcal": avg_rp_kcal,
            "count": len(values)
        })
    
    # Sort by rp_kcal (lowest = most efficient) and take top 10
    supplier_economy.sort(key=lambda x: x["rp_kcal"])
    supplier_economy = supplier_economy[:10]
    
    # === 5. Slagging Risk Matrix (Heatmap) - Include Vessel, Barge, and Trucking ===
    slagging_matrix = []
    
    # Get from Vessels
    vessels_slagging = await db.vessels.find(
        {"$or": [{"slagging_index": {"$ne": None}}, {"fouling_index": {"$ne": None}}]},
        {"_id": 0, "name_of_vessel": 1, "suppliers": 1, "slagging_index": 1, "fouling_index": 1, "completed_unloading": 1}
    ).sort("completed_unloading", -1).limit(15).to_list(15)
    
    # Get from Barges
    barges_slagging = await db.barges.find(
        {"$or": [{"slagging_index": {"$ne": None}}, {"fouling_index": {"$ne": None}}]},
        {"_id": 0, "name_of_barge": 1, "suppliers": 1, "slagging_index": 1, "fouling_index": 1, "completed_unloading": 1}
    ).sort("completed_unloading", -1).limit(15).to_list(15)
    
    # Get from Trucking
    trucking_slagging = await db.trucking.find(
        {"$or": [{"slagging_index": {"$ne": None}}, {"fouling_index": {"$ne": None}}]},
        {"_id": 0, "suppliers": 1, "slagging_index": 1, "fouling_index": 1, "completed_unloading": 1}
    ).sort("completed_unloading", -1).limit(15).to_list(15)
    
    def get_risk_level(index_value):
        if not index_value:
            return "UNKNOWN"
        val = str(index_value).upper()
        if "SEVERE" in val:
            return "SEVERE"
        elif "HIGH" in val:
            return "HIGH"
        elif "MEDIUM" in val:
            return "MEDIUM"
        elif "LOW" in val:
            return "LOW"
        return "UNKNOWN"
    
    # Add Vessel data
    for v in vessels_slagging:
        slagging_matrix.append({
            "name": v.get("name_of_vessel", "Unknown Vessel"),
            "supplier": v.get("suppliers", ""),
            "slagging": v.get("slagging_index", ""),
            "slagging_risk": get_risk_level(v.get("slagging_index")),
            "fouling": v.get("fouling_index", ""),
            "fouling_risk": get_risk_level(v.get("fouling_index")),
            "date": str(v.get("completed_unloading", ""))[:10],
            "moda": "Vessel"
        })
    
    # Add Barge data
    for b in barges_slagging:
        slagging_matrix.append({
            "name": b.get("name_of_barge", b.get("suppliers", "Unknown Barge")),
            "supplier": b.get("suppliers", ""),
            "slagging": b.get("slagging_index", ""),
            "slagging_risk": get_risk_level(b.get("slagging_index")),
            "fouling": b.get("fouling_index", ""),
            "fouling_risk": get_risk_level(b.get("fouling_index")),
            "date": str(b.get("completed_unloading", ""))[:10],
            "moda": "Barge"
        })
    
    # Add Trucking data
    for t in trucking_slagging:
        slagging_matrix.append({
            "name": t.get("suppliers", "Unknown Trucking"),
            "supplier": t.get("suppliers", ""),
            "slagging": t.get("slagging_index", ""),
            "slagging_risk": get_risk_level(t.get("slagging_index")),
            "fouling": t.get("fouling_index", ""),
            "fouling_risk": get_risk_level(t.get("fouling_index")),
            "date": str(t.get("completed_unloading", ""))[:10],
            "moda": "Trucking"
        })
    
    # Sort by date descending
    slagging_matrix.sort(key=lambda x: x.get("date", ""), reverse=True)
    
    # === 6. Six Months Summary ===
    six_months_summary = []
    current_date = date.today()
    
    for i in range(6):
        target_date = current_date - relativedelta(months=i)
        year_val = target_date.year
        month_val = target_date.month
        month_name = target_date.strftime("%b %Y")
        
        # Count by month
        vessel_count = await db.vessels.count_documents({
            "completed_unloading": {"$regex": f"^{year_val}-{month_val:02d}"}
        })
        barge_count = await db.barges.count_documents({
            "completed_unloading": {"$regex": f"^{year_val}-{month_val:02d}"}
        })
        trucking_count = await db.trucking.count_documents({
            "completed_unloading": {"$regex": f"^{year_val}-{month_val:02d}"}
        })
        biomassa_count = await db.biomassa.count_documents({
            "completed_unloading": {"$regex": f"^{year_val}-{month_val:02d}"}
        })
        
        # Tonase by month
        vessel_ton = await db.vessels.aggregate([
            {"$match": {"completed_unloading": {"$regex": f"^{year_val}-{month_val:02d}"}}},
            {"$group": {"_id": None, "total": {"$sum": {"$ifNull": ["$ds_mt", 0]}}}}
        ]).to_list(1)
        barge_ton = await db.barges.aggregate([
            {"$match": {"completed_unloading": {"$regex": f"^{year_val}-{month_val:02d}"}}},
            {"$group": {"_id": None, "total": {"$sum": {"$ifNull": ["$ds_mt", 0]}}}}
        ]).to_list(1)
        
        six_months_summary.append({
            "month": month_name,
            "year": year_val,
            "month_num": month_val,
            "vessel": vessel_count,
            "barge": barge_count,
            "trucking": trucking_count,
            "biomassa": biomassa_count,
            "total_shipments": vessel_count + barge_count + trucking_count + biomassa_count,
            "vessel_tonase": vessel_ton[0]["total"] if vessel_ton else 0,
            "barge_tonase": barge_ton[0]["total"] if barge_ton else 0,
            "total_tonase": (vessel_ton[0]["total"] if vessel_ton else 0) + (barge_ton[0]["total"] if barge_ton else 0)
        })
    
    # Reverse to show oldest first
    six_months_summary.reverse()
    
    # === Available filters ===
    available_periods = []
    years = [2023, 2024, 2025, 2026]
    months = ["Januari", "Februari", "Maret", "April", "Mei", "Juni", 
              "Juli", "Agustus", "September", "Oktober", "November", "Desember"]
    for y in years:
        for m_idx, m_name in enumerate(months, 1):
            available_periods.append({"year": y, "month": m_idx, "label": f"{m_name} {y}"})
    
    available_moda = ["Vessel", "Barge", "Trucking", "Tongkang"]
    
    return {
        "total_ds_mt": total_ds_mt,
        "total_tonase_po": total_tonase_po,
        "contract_percentage": min(contract_percentage, 100),
        "fuel_composition": fuel_composition,
        "gcv_trend": gcv_trend,
        "supplier_economy": supplier_economy,
        "slagging_matrix": slagging_matrix,
        "six_months_summary": six_months_summary,
        "available_periods": available_periods[-24:],  # Last 2 years
        "available_moda": available_moda
    }

# ==================== AI INTELLIGENCE AGENT ====================
from emergentintegrations.llm.chat import LlmChat, UserMessage

# AI Chat History Collection
ai_chat_collection = db.ai_chat_history

class AIQueryRequest(BaseModel):
    query: str
    module: str = "general"  # general, blending, boiler_risk, contract, logistics
    session_id: Optional[str] = None
    parameters: Optional[dict] = None

class AISettingsUpdate(BaseModel):
    custom_api_key: Optional[str] = None
    llm_provider: Optional[str] = "gemini"
    llm_model: Optional[str] = "gemini-2.5-flash"

async def get_database_context(module: str, parameters: dict = None) -> str:
    """Gather relevant data from database based on module"""
    context_parts = []
    
    if module in ["general", "blending", "contract"]:
        # Get PO Batubara summary
        po_data = await db.po_batubara.aggregate([
            {"$group": {
                "_id": "$spec",
                "total_tonase": {"$sum": "$tonase_po"},
                "count": {"$sum": 1}
            }}
        ]).to_list(100)
        context_parts.append(f"PO Batubara Summary: {po_data}")
        
        # Get Merit Order data (top suppliers by efficiency)
        merit_data = await db.merit_order.find(
            {"rp_kcal": {"$ne": None}},
            {"_id": 0, "pemasok": 1, "moda": 1, "tipikal_kcal_kg": 1, "harga_cif": 1, "rp_kcal": 1}
        ).sort("rp_kcal", 1).limit(10).to_list(10)
        context_parts.append(f"Top 10 Efficient Suppliers (Merit Order): {merit_data}")
    
    if module in ["general", "blending", "boiler_risk"]:
        # Get recent vessel quality data
        vessel_quality = await db.vessels.find(
            {"gcv_arb": {"$ne": None}},
            {"_id": 0, "name_of_vessel": 1, "suppliers": 1, "gcv_arb": 1, "tm_arb": 1, 
             "ash_arb": 1, "slagging_index": 1, "fouling_index": 1, "completed_unloading": 1}
        ).sort("completed_unloading", -1).limit(20).to_list(20)
        context_parts.append(f"Recent Vessel Quality Data: {vessel_quality}")
        
        # Get barge quality data
        barge_quality = await db.barges.find(
            {"gcv_arb": {"$ne": None}},
            {"_id": 0, "suppliers": 1, "gcv_arb": 1, "tm_arb": 1, "ash_arb": 1,
             "slagging_index": 1, "fouling_index": 1, "completed_unloading": 1}
        ).sort("completed_unloading", -1).limit(20).to_list(20)
        context_parts.append(f"Recent Barge Quality Data: {barge_quality}")
    
    if module in ["general", "logistics"]:
        # Get logistics data (B/L vs DS differences)
        vessel_logistics = await db.vessels.find(
            {"bl_mt": {"$ne": None}, "ds_mt": {"$ne": None}},
            {"_id": 0, "name_of_vessel": 1, "suppliers": 1, "bl_mt": 1, "ds_mt": 1, 
             "commenced_unloading": 1, "completed_unloading": 1}
        ).sort("completed_unloading", -1).limit(30).to_list(30)
        context_parts.append(f"Vessel Logistics Data (B/L vs DS): {vessel_logistics}")
        
        barge_logistics = await db.barges.find(
            {"bl_mt": {"$ne": None}, "ds_mt": {"$ne": None}},
            {"_id": 0, "suppliers": 1, "bl_mt": 1, "ds_mt": 1,
             "commenced_unloading": 1, "completed_unloading": 1}
        ).sort("completed_unloading", -1).limit(30).to_list(30)
        context_parts.append(f"Barge Logistics Data: {barge_logistics}")
    
    if module in ["general", "contract"]:
        # Get contract compliance data
        po_summary = await db.po_batubara.aggregate([
            {"$group": {
                "_id": "$supplier_name",
                "total_po_tonase": {"$sum": "$tonase_po"},
                "po_count": {"$sum": 1}
            }},
            {"$sort": {"total_po_tonase": -1}},
            {"$limit": 15}
        ]).to_list(15)
        context_parts.append(f"PO by Supplier: {po_summary}")
        
        # Get actual receipts
        vessel_receipts = await db.vessels.aggregate([
            {"$group": {
                "_id": "$suppliers",
                "total_received": {"$sum": "$ds_mt"},
                "shipment_count": {"$sum": 1}
            }},
            {"$sort": {"total_received": -1}},
            {"$limit": 15}
        ]).to_list(15)
        context_parts.append(f"Vessel Receipts by Supplier: {vessel_receipts}")
    
    if module == "blending" and parameters:
        target_gcv = parameters.get("target_gcv", 4000)
        context_parts.append(f"User Target GCV: {target_gcv} Kcal/kg")
        
        # Get available stock for blending
        available_stock = await db.vessels.find(
            {"gcv_arb": {"$ne": None}, "ds_mt": {"$gt": 0}},
            {"_id": 0, "name_of_vessel": 1, "suppliers": 1, "gcv_arb": 1, "ds_mt": 1, "ash_arb": 1}
        ).sort("completed_unloading", -1).limit(15).to_list(15)
        context_parts.append(f"Available Vessel Stock for Blending: {available_stock}")
        
        biomass_stock = await db.biomassa.find(
            {"jembatan_timbang_mt": {"$gt": 0}},
            {"_id": 0, "biomass_type": 1, "jembatan_timbang_mt": 1, "gcv_arb": 1}
        ).sort("completed_unloading", -1).limit(10).to_list(10)
        context_parts.append(f"Available Biomass Stock: {biomass_stock}")
    
    # Smart Stock Module Data
    if module in ["general", "smart_stock"]:
        # Get smart stock penerimaan data
        penerimaan_data = await db.smart_stock.find(
            {},
            {"_id": 0, "source": 1, "supplier": 1, "cargo": 1, "tonase": 1, "gcv_arb": 1, "date": 1}
        ).sort("date", -1).limit(30).to_list(30)
        if penerimaan_data:
            context_parts.append(f"Smart Stock - Sumber Penerimaan (30 terbaru): {penerimaan_data}")
        
        # Get smart stock pemakaian data
        pemakaian_data = await db.sumber_pemakaian.find(
            {},
            {"_id": 0, "tanggal": 1, "energy_mwh": 1, "batubara_mt": 1, "biomassa_mt": 1, "sfc": 1}
        ).sort("tanggal", -1).limit(30).to_list(30)
        if pemakaian_data:
            context_parts.append(f"Smart Stock - Sumber Pemakaian (30 terbaru): {pemakaian_data}")
        
        # Calculate stock summary
        total_penerimaan = await db.smart_stock.aggregate([
            {"$group": {"_id": None, "total": {"$sum": "$tonase"}}}
        ]).to_list(1)
        total_pemakaian = await db.sumber_pemakaian.aggregate([
            {"$group": {"_id": None, "total_batubara": {"$sum": "$batubara_mt"}, "total_biomassa": {"$sum": "$biomassa_mt"}}}
        ]).to_list(1)
        
        if total_penerimaan and total_pemakaian:
            penerimaan = total_penerimaan[0].get("total", 0) if total_penerimaan else 0
            batubara_pakai = total_pemakaian[0].get("total_batubara", 0) if total_pemakaian else 0
            biomassa_pakai = total_pemakaian[0].get("total_biomassa", 0) if total_pemakaian else 0
            context_parts.append(f"Ringkasan Stok: Total Penerimaan={penerimaan} MT, Total Pemakaian Batubara={batubara_pakai} MT, Total Pemakaian Biomassa={biomassa_pakai} MT")
    
    # COA Reconciliation Module Data
    if module in ["general", "coa_reconciliation"]:
        # Get COA reconciliation data
        coa_data = await db.coa_reconciliation.find(
            {},
            {"_id": 0, "shipment": 1, "supplier": 1, "loading_gcv_arb": 1, "unloading_gcv_arb": 1, 
             "internal_gcv_arb": 1, "delta_loading_internal": 1, "status": 1, "umpire_status": 1,
             "completed_unloading": 1}
        ).sort("completed_unloading", -1).limit(30).to_list(30)
        if coa_data:
            context_parts.append(f"COA Reconciliation Data (30 terbaru): {coa_data}")
        
        # Get KPIs summary
        all_coa = await db.coa_reconciliation.find({}, {"_id": 0}).to_list(10000)
        if all_coa:
            kritis_count = sum(1 for c in all_coa if c.get("status") == "Kritis")
            umpire_count = sum(1 for c in all_coa if c.get("umpire_status") not in [None, "none", ""])
            
            # Calculate potential loss
            settings = await db.settings.find_one({"type": "coa"})
            price_per_kcal = settings.get("price_per_kcal_per_ton", 50) if settings else 50
            potential_loss = 0
            for c in all_coa:
                delta = c.get("delta_loading_internal", 0) or 0
                tonase = c.get("tonase", 0) or c.get("internal_tonase", 0) or 0
                if delta > 0:
                    potential_loss += delta * tonase * price_per_kcal
            
            context_parts.append(f"COA KPIs: Total Records={len(all_coa)}, Status Kritis={kritis_count}, Dalam Proses Umpire={umpire_count}, Potential Loss=Rp {potential_loss:,.0f}")
            
            # Supplier dengan deviasi tertinggi
            supplier_deviasi = {}
            for c in all_coa:
                supplier = c.get("supplier", "Unknown")
                delta = abs(c.get("delta_loading_internal", 0) or 0)
                if supplier not in supplier_deviasi:
                    supplier_deviasi[supplier] = []
                supplier_deviasi[supplier].append(delta)
            
            top_deviasi = sorted(
                [(s, sum(d)/len(d)) for s, d in supplier_deviasi.items() if d],
                key=lambda x: x[1], reverse=True
            )[:5]
            context_parts.append(f"Top 5 Supplier dengan Deviasi GCV Tertinggi: {top_deviasi}")
    
    return "\n\n".join(context_parts)

def get_system_prompt(module: str) -> str:
    """Get system prompt based on module"""
    base_prompt = """Anda adalah Tenayan Fuel Intelligence Agent - Asisten ahli Rendal Bahan Bakar PLTU Tenayan.
Tugas Anda adalah mengolah data dari database (Vessel, Barge, Trucking, Biomassa, PO, Merit Order, dan COA Reconciliation) 
untuk memberikan wawasan berbasis Data Science, Machine Learning, dan AI.

PENTING:
- Jawab dalam Bahasa Indonesia
- Berikan analisis yang akurat berdasarkan data
- Sertakan perhitungan numerik jika diperlukan
- Format output dalam tabel atau bullet points untuk kemudahan membaca
- Berikan rekomendasi actionable

FITUR APLIKASI YANG TERSEDIA:

1. **Dashboard**: Ringkasan KPI penerimaan bahan bakar, chart tren, dan statistik.

2. **Data Penerimaan**: Halaman untuk melihat, menambah, dan mengelola data penerimaan dari Vessel, Barge, Trucking, dan Biomassa.

3. **Merit Order**: Tabel ranking supplier berdasarkan harga per kCal (Rp/kCal) dari yang termurah.

4. **Smart Stock**: Monitoring stok batubara di stockpile dengan visualisasi.

5. **Laporan**: Generate laporan dalam format PDF dan Excel.

6. **COA Reconciliation (Triple Check)**: 
   - Fitur untuk membandingkan data kualitas batubara dari 3 sumber: Loading COA (dari supplier), Unloading COA (di pelabuhan), dan Lab Internal (hasil uji lab PLTU).
   - Menampilkan parameter kunci: GCV, TM (Total Moisture), Ash, dan Sulphur dari ketiga sumber secara berdampingan.
   - Menghitung Delta GCV = Loading GCV - Internal GCV untuk mengidentifikasi perbedaan klaim supplier vs hasil lab internal.
   - Status "Kritis" ditandai jika Delta GCV > 150 kCal/kg DAN Loading GCV > Internal GCV (supplier overclaim).
   - KPI Dashboard: High Deviation Alert, Potential Loss (Rp), dan Umpire Status.
   - Radar Chart untuk visualisasi perbandingan profil kualitas antar sumber.

7. **Dispute Monitor (Umpire Process)**:
   - Halaman untuk mengelola proses umpire/arbitrase ketika terjadi dispute kualitas batubara.
   - Workflow: Propose Umpire → Sedang Proses → Selesai (dengan input hasil lab umpire).
   - Mendukung "Quad Check" - membandingkan 4 sumber data termasuk hasil lab umpire.
   - Input hasil umpire: GCV, TM, Ash, Sulphur, Nama Lab, dan Tanggal Hasil.

8. **Settings**: Konfigurasi parameter aplikasi termasuk Harga per kCal per Ton untuk perhitungan Potential Loss.

9. **AI Intelligence**: Fitur AI Assistant (Anda) untuk analisis data dan rekomendasi.
"""
    
    module_prompts = {
        "blending": base_prompt + """
MODUL: Smart Blending Optimizer
Fokus: Optimasi campuran batubara & biomassa secara ekonomis.

Tugas spesifik:
1. Identifikasi kargo tersedia berdasarkan 'Completed Unloading' terbaru
2. Lakukan perhitungan optimasi untuk mencari porsi (%) campuran LRC, MRC, dan Biomassa
3. Batasan: Total campuran harus mencapai Target GCV, dengan Ash Content < 10%
4. Utamakan stok dengan Rp/Kcal terendah dari Merit Order
5. Tampilkan hasil: [Nama Supplier] | [Porsi %] | [Tonase Rekomendasi] | [Harga Estimasi]

Rumus GCV Campuran: GCV_mix = Σ(GCV_i × Porsi_i)
""",
        "boiler_risk": base_prompt + """
MODUL: Boiler Risk Warning
Fokus: Deteksi potensi kerusakan boiler dari data laboratorium.

Tugas spesifik:
1. Analisis data kualitas kimia: SiO2, Al2O3, Fe2O3, CaO, Na2O, K2O
2. Hitung indeks Slagging dan Fouling menggunakan rumus standar industri
3. Kriteria risiko:
   - Slagging Index > 0.6 atau Fouling Index > 0.4 = 'HIGH RISK'
   - Na2O > 2% = Peringatan khusus potensi kerak pipa
4. Tampilkan 'Alert Board' daftar kargo berisiko tinggi di stockpile
5. Berikan rekomendasi strategi sootblowing

Rumus Slagging Index: Rs = (Base/Acid) × (S content)
Rumus Fouling Index: Rf = (Base/Acid) × Na2O
""",
        "contract": base_prompt + """
MODUL: Contract Compliance & PO Tracker
Fokus: Digitalisasi monitoring kontrak tanpa rekap manual.

Tugas spesifik:
1. Lakukan vlookup otomatis antara PO BB dengan penerimaan (barge, vessel, trucking)
2. Hitung: Sisa Kuota = Tonase PO - Total DS MT yang diterima
3. Bandingkan GCV Kontrak (di PO) dengan GCV Realisasi (di Penerimaan)
4. Output Dashboard:
   - List PO yang hampir habis (< 10% sisa)
   - List Supplier dengan GCV di bawah spek kontrak (Defisit Kalori)
5. Berikan early warning untuk kontrak yang perlu diperpanjang
""",
        "logistics": base_prompt + """
MODUL: Logistic Efficiency & Loss Analysis
Fokus: Data Science pada efisiensi pengiriman.

Tugas spesifik:
1. Hitung selisih antara B/L (MT) dan DS (MT) - Draft Survey
2. Hitung rata-rata % Losses per supplier: ((B/L - DS) / B/L) × 100%
3. Hitung 'Durasi Pembongkaran' rata-rata per moda transportasi
4. Output:
   - 3 Supplier dengan tingkat penyusutan (losses) tertinggi
   - Tren durasi bongkar per bulan untuk deteksi inefisiensi di Jetty
   - Anomali pengiriman yang perlu investigasi
""",
        "general": base_prompt + """
MODUL: General Intelligence
Fokus: Menjawab pertanyaan umum tentang data bahan bakar PLTU Tenayan.

Anda dapat:
1. Memberikan ringkasan statistik dari semua data
2. Menjawab pertanyaan spesifik tentang supplier, kualitas, atau pengiriman
3. Memberikan rekomendasi berdasarkan analisis data
4. Menjelaskan tren dan pola dalam data historis
""",
        "smart_stock": base_prompt + """
MODUL: Smart Stock Management
Fokus: Analisis stok batubara dan biomassa di stockpile.

Tugas spesifik:
1. Monitoring stok aktual batubara di stockpile dari berbagai sumber (Vessel, Barge, Trucking, Biomassa)
2. Analisis pemakaian harian dan tren konsumsi bahan bakar
3. Perhitungan estimasi hari stok tersedia (Days of Supply)
4. Identifikasi anomali penerimaan vs pemakaian
5. Rekomendasi waktu optimal untuk pemesanan ulang

Data yang tersedia:
- Sumber Penerimaan: Data penerimaan batubara dari Vessel, Barge, Trucking, dan Biomassa
- Sumber Pemakaian: Data pemakaian harian batubara termasuk energi (MWh) dan tonase
- Merit Order: Ranking supplier berdasarkan harga per kCal

Output yang diharapkan:
- Total Stok Saat Ini (MT)
- Rata-rata Pemakaian Harian (MT/hari)
- Estimasi Hari Stok Tersisa
- Rekomendasi pengisian stok
""",
        "coa_reconciliation": base_prompt + """
MODUL: COA Reconciliation & Quality Dispute
Fokus: Rekonsiliasi kualitas batubara dari tiga sumber dan manajemen dispute.

Tugas spesifik:
1. Analisis Triple Check: Bandingkan Loading COA (supplier), Unloading COA (surveyor), dan Lab Internal (PLTU)
2. Identifikasi lot dengan deviasi GCV tinggi (Delta > 150 kCal/kg)
3. Deteksi potensi overclaim dari supplier (Loading GCV > Internal GCV)
4. Hitung Potential Loss (Rp) dari defisit kalori
5. Monitoring status umpire/arbitrase yang sedang berjalan
6. Analisis konsistensi supplier berdasarkan historis deviasi

Parameter Kunci yang dianalisis:
- GCV ARB (Gross Calorific Value As Received Basis)
- TM ARB (Total Moisture)
- Ash ARB (Kadar Abu)
- Sulphur ARB (Kadar Belerang)

Status Kritis: Ditandai jika Delta_GCV > 150 kCal/kg DAN Loading GCV > Internal GCV (overclaim)

Rumus Perhitungan:
- Delta GCV = Loading_GCV - Internal_GCV
- Potential Loss = Tonase × Delta_GCV × Harga_per_kCal_per_Ton

Output yang diharapkan:
- High Deviation Alert: Jumlah lot dengan status Kritis
- Potential Loss (Rp): Estimasi kerugian finansial
- Supplier dengan deviasi tertinggi
- Rekomendasi pengajuan umpire
"""
    }
    
    return module_prompts.get(module, module_prompts["general"])

@api_router.post("/ai/query")
async def ai_query(request: AIQueryRequest, user: dict = Depends(get_current_user)):
    """Process AI query with database context"""
    try:
        # Get user's custom API key or use default
        user_settings = await db.user_settings.find_one({"user_id": user["id"]})
        
        api_key = os.environ.get("EMERGENT_LLM_KEY")
        llm_provider = "gemini"
        llm_model = "gemini-2.5-flash"
        
        if user_settings:
            if user_settings.get("custom_api_key"):
                api_key = user_settings["custom_api_key"]
            if user_settings.get("llm_provider"):
                llm_provider = user_settings["llm_provider"]
            if user_settings.get("llm_model"):
                llm_model = user_settings["llm_model"]
        
        # Generate session ID if not provided
        session_id = request.session_id or f"tenayan-ai-{user['id']}-{uuid.uuid4()}"
        
        # Get database context
        db_context = await get_database_context(request.module, request.parameters)
        
        # Get system prompt
        system_prompt = get_system_prompt(request.module)
        
        # Initialize LLM Chat
        chat = LlmChat(
            api_key=api_key,
            session_id=session_id,
            system_message=system_prompt
        ).with_model(llm_provider, llm_model)
        
        # Prepare message with context
        full_query = f"""DATA CONTEXT:
{db_context}

USER QUERY:
{request.query}

Berikan analisis dan jawaban berdasarkan data di atas."""
        
        user_message = UserMessage(text=full_query)
        
        # Get AI response
        response = await chat.send_message(user_message)
        
        # Save to chat history
        chat_entry = {
            "id": str(uuid.uuid4()),
            "user_id": user["id"],
            "session_id": session_id,
            "module": request.module,
            "query": request.query,
            "response": response,
            "parameters": request.parameters,
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await ai_chat_collection.insert_one(chat_entry)
        
        return {
            "response": response,
            "session_id": session_id,
            "module": request.module
        }
        
    except Exception as e:
        logger.error(f"AI Query Error: {e}")
        raise HTTPException(status_code=500, detail=f"AI Query failed: {str(e)}")

@api_router.get("/ai/history")
async def get_ai_chat_history(
    session_id: Optional[str] = None,
    limit: int = Query(20, ge=1, le=100),
    user: dict = Depends(get_current_user)
):
    """Get AI chat history for user"""
    query = {"user_id": user["id"]}
    if session_id:
        query["session_id"] = session_id
    
    history = await ai_chat_collection.find(
        query, {"_id": 0}
    ).sort("created_at", -1).limit(limit).to_list(limit)
    
    return history

@api_router.delete("/ai/history")
async def clear_ai_chat_history(user: dict = Depends(get_current_user)):
    """Clear AI chat history for user"""
    result = await ai_chat_collection.delete_many({"user_id": user["id"]})
    return {"message": f"Berhasil menghapus {result.deleted_count} riwayat chat"}

@api_router.get("/ai/settings")
async def get_ai_settings(user: dict = Depends(get_current_user)):
    """Get user's AI settings"""
    settings = await db.user_settings.find_one({"user_id": user["id"]}, {"_id": 0})
    if not settings:
        return {
            "custom_api_key": None,
            "llm_provider": "gemini",
            "llm_model": "gemini-2.5-flash",
            "using_default": True
        }
    return {
        **settings,
        "using_default": not bool(settings.get("custom_api_key"))
    }

@api_router.put("/ai/settings")
async def update_ai_settings(settings: AISettingsUpdate, user: dict = Depends(get_current_user)):
    """Update user's AI settings"""
    update_data = {
        "user_id": user["id"],
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    if settings.custom_api_key is not None:
        update_data["custom_api_key"] = settings.custom_api_key if settings.custom_api_key else None
    if settings.llm_provider:
        update_data["llm_provider"] = settings.llm_provider
    if settings.llm_model:
        update_data["llm_model"] = settings.llm_model
    
    await db.user_settings.update_one(
        {"user_id": user["id"]},
        {"$set": update_data},
        upsert=True
    )
    
    return {"message": "Pengaturan AI berhasil disimpan"}

# Quick Analysis Endpoints for Dashboard Modules
@api_router.get("/ai/quick/blending-suggestion")
async def get_blending_suggestion(target_gcv: int = 4000, user: dict = Depends(get_current_user)):
    """Get quick blending suggestion without full AI query"""
    # Get available coal stock
    vessels = await db.vessels.find(
        {"gcv_arb": {"$ne": None}, "ds_mt": {"$gt": 0}},
        {"_id": 0, "name_of_vessel": 1, "suppliers": 1, "gcv_arb": 1, "ds_mt": 1, "ash_arb": 1}
    ).sort("completed_unloading", -1).limit(10).to_list(10)
    
    biomass = await db.biomassa.find(
        {"jembatan_timbang_mt": {"$gt": 0}},
        {"_id": 0, "biomass_type": 1, "jembatan_timbang_mt": 1}
    ).limit(5).to_list(5)
    
    return {
        "target_gcv": target_gcv,
        "available_coal": vessels,
        "available_biomass": biomass,
        "recommendation": "Gunakan modul AI untuk analisis blending yang lebih detail"
    }

@api_router.get("/ai/quick/boiler-alerts")
async def get_boiler_alerts(user: dict = Depends(get_current_user)):
    """Get quick boiler risk alerts"""
    # Get high risk items
    high_risk_vessels = await db.vessels.find(
        {"$or": [
            {"slagging_index": {"$regex": "HIGH|SEVERE", "$options": "i"}},
            {"fouling_index": {"$regex": "HIGH|SEVERE", "$options": "i"}}
        ]},
        {"_id": 0, "name_of_vessel": 1, "suppliers": 1, "slagging_index": 1, "fouling_index": 1}
    ).sort("completed_unloading", -1).limit(10).to_list(10)
    
    high_risk_barges = await db.barges.find(
        {"$or": [
            {"slagging_index": {"$regex": "HIGH|SEVERE", "$options": "i"}},
            {"fouling_index": {"$regex": "HIGH|SEVERE", "$options": "i"}}
        ]},
        {"_id": 0, "suppliers": 1, "slagging_index": 1, "fouling_index": 1}
    ).sort("completed_unloading", -1).limit(10).to_list(10)
    
    return {
        "high_risk_vessels": high_risk_vessels,
        "high_risk_barges": high_risk_barges,
        "total_alerts": len(high_risk_vessels) + len(high_risk_barges)
    }

@api_router.get("/ai/quick/contract-status")
async def get_contract_status(user: dict = Depends(get_current_user)):
    """Get quick contract compliance status"""
    # Get PO summary
    po_summary = await db.po_batubara.aggregate([
        {"$group": {
            "_id": "$supplier_name",
            "total_po": {"$sum": "$tonase_po"},
            "po_count": {"$sum": 1}
        }},
        {"$sort": {"total_po": -1}},
        {"$limit": 10}
    ]).to_list(10)
    
    # Get receipts
    receipts = await db.vessels.aggregate([
        {"$group": {
            "_id": "$suppliers",
            "total_received": {"$sum": "$ds_mt"}
        }}
    ]).to_list(100)
    
    receipts_map = {r["_id"]: r["total_received"] for r in receipts if r["_id"]}
    
    contract_status = []
    for po in po_summary:
        supplier = po["_id"]
        received = receipts_map.get(supplier, 0)
        remaining = po["total_po"] - received
        percentage = (received / po["total_po"] * 100) if po["total_po"] > 0 else 0
        
        contract_status.append({
            "supplier": supplier,
            "total_po": po["total_po"],
            "received": received,
            "remaining": remaining,
            "percentage": percentage,
            "status": "CRITICAL" if percentage > 90 else "WARNING" if percentage > 70 else "OK"
        })
    
    return {"contracts": contract_status}

@api_router.get("/ai/quick/logistics-losses")
async def get_logistics_losses(user: dict = Depends(get_current_user)):
    """Get quick logistics losses analysis"""
    # Calculate losses from vessels
    vessels = await db.vessels.find(
        {"bl_mt": {"$ne": None}, "ds_mt": {"$ne": None}},
        {"_id": 0, "suppliers": 1, "bl_mt": 1, "ds_mt": 1}
    ).to_list(1000)
    
    supplier_losses = {}
    for v in vessels:
        supplier = v.get("suppliers", "Unknown")
        bl = v.get("bl_mt", 0) or 0
        ds = v.get("ds_mt", 0) or 0
        if bl > 0:
            loss_pct = ((bl - ds) / bl) * 100
            if supplier not in supplier_losses:
                supplier_losses[supplier] = []
            supplier_losses[supplier].append(loss_pct)
    
    # Calculate averages
    losses_summary = []
    for supplier, losses in supplier_losses.items():
        avg_loss = sum(losses) / len(losses)
        losses_summary.append({
            "supplier": supplier,
            "avg_loss_pct": avg_loss,
            "shipment_count": len(losses)
        })
    
    # Sort by highest loss
    losses_summary.sort(key=lambda x: x["avg_loss_pct"], reverse=True)
    
    return {
        "top_losses": losses_summary[:5],
        "lowest_losses": losses_summary[-5:] if len(losses_summary) > 5 else []
    }

@api_router.get("/ai/quick/smart-stock")
async def get_smart_stock_summary(user: dict = Depends(get_current_user)):
    """Get quick smart stock summary"""
    # Get total penerimaan
    total_penerimaan = await db.smart_stock.aggregate([
        {"$group": {"_id": None, "total": {"$sum": "$tonase"}}}
    ]).to_list(1)
    
    # Get total pemakaian
    total_pemakaian = await db.sumber_pemakaian.aggregate([
        {"$group": {
            "_id": None, 
            "total_batubara": {"$sum": "$batubara_mt"},
            "total_biomassa": {"$sum": "$biomassa_mt"}
        }}
    ]).to_list(1)
    
    # Get average daily usage (last 30 days)
    avg_usage = await db.sumber_pemakaian.aggregate([
        {"$sort": {"tanggal": -1}},
        {"$limit": 30},
        {"$group": {
            "_id": None,
            "avg_batubara": {"$avg": "$batubara_mt"},
            "avg_energy": {"$avg": "$energy_mwh"}
        }}
    ]).to_list(1)
    
    penerimaan = total_penerimaan[0].get("total", 0) if total_penerimaan else 0
    batubara_pakai = total_pemakaian[0].get("total_batubara", 0) if total_pemakaian else 0
    biomassa_pakai = total_pemakaian[0].get("total_biomassa", 0) if total_pemakaian else 0
    avg_daily = avg_usage[0].get("avg_batubara", 0) if avg_usage else 0
    
    current_stock = penerimaan - batubara_pakai - biomassa_pakai
    days_of_supply = int(current_stock / avg_daily) if avg_daily > 0 else 0
    
    return {
        "total_penerimaan": penerimaan,
        "total_pemakaian": batubara_pakai + biomassa_pakai,
        "current_stock": current_stock,
        "avg_daily_usage": avg_daily,
        "days_of_supply": days_of_supply,
        "status": "CRITICAL" if days_of_supply < 7 else "WARNING" if days_of_supply < 14 else "OK"
    }

@api_router.get("/ai/quick/coa-alerts")
async def get_coa_alerts(user: dict = Depends(get_current_user)):
    """Get quick COA reconciliation alerts"""
    # Get all COA data
    all_coa = await db.coa_reconciliation.find({}, {"_id": 0}).to_list(10000)
    
    if not all_coa:
        return {
            "total_records": 0,
            "kritis_count": 0,
            "umpire_count": 0,
            "potential_loss": 0,
            "top_supplier_deviasi": None
        }
    
    # Count kritis status
    kritis_count = sum(1 for c in all_coa if c.get("status") == "Kritis")
    
    # Count umpire in progress
    umpire_count = sum(1 for c in all_coa if c.get("umpire_status") not in [None, "none", ""])
    
    # Calculate potential loss
    settings = await db.settings.find_one({"type": "coa"})
    price_per_kcal = settings.get("price_per_kcal_per_ton", 50) if settings else 50
    
    potential_loss = 0
    supplier_deviasi = {}
    
    for c in all_coa:
        delta = c.get("delta_loading_internal", 0) or 0
        tonase = c.get("tonase", 0) or c.get("internal_tonase", 0) or 0
        supplier = c.get("supplier", "Unknown")
        
        if delta > 0:
            potential_loss += delta * tonase * price_per_kcal
        
        if supplier not in supplier_deviasi:
            supplier_deviasi[supplier] = []
        supplier_deviasi[supplier].append(abs(delta))
    
    # Find supplier with highest average deviation
    top_supplier = None
    max_avg_deviasi = 0
    for supplier, deviasi_list in supplier_deviasi.items():
        if deviasi_list:
            avg = sum(deviasi_list) / len(deviasi_list)
            if avg > max_avg_deviasi:
                max_avg_deviasi = avg
                top_supplier = {"name": supplier, "avg_deviasi": avg}
    
    return {
        "total_records": len(all_coa),
        "kritis_count": kritis_count,
        "umpire_count": umpire_count,
        "potential_loss": potential_loss,
        "top_supplier_deviasi": top_supplier
    }

# ==================== AI CONVERSATION SESSIONS ====================

@api_router.get("/ai/sessions")
async def get_ai_sessions(
    page: int = Query(1, ge=1),
    page_size: int = Query(10, ge=1, le=50),
    user: dict = Depends(get_current_user)
):
    """Get list of user's AI conversation sessions"""
    skip = (page - 1) * page_size
    
    # Aggregate to get unique sessions with their last message
    pipeline = [
        {"$match": {"user_id": user["id"]}},
        {"$sort": {"created_at": -1}},
        {"$group": {
            "_id": "$session_id",
            "module": {"$first": "$module"},
            "last_query": {"$first": "$query"},
            "last_response": {"$first": "$response"},
            "message_count": {"$sum": 1},
            "created_at": {"$min": "$created_at"},
            "updated_at": {"$max": "$created_at"}
        }},
        {"$sort": {"updated_at": -1}},
        {"$skip": skip},
        {"$limit": page_size}
    ]
    
    sessions = await ai_chat_collection.aggregate(pipeline).to_list(page_size)
    
    # Count total unique sessions
    total_pipeline = [
        {"$match": {"user_id": user["id"]}},
        {"$group": {"_id": "$session_id"}},
        {"$count": "total"}
    ]
    total_result = await ai_chat_collection.aggregate(total_pipeline).to_list(1)
    total = total_result[0]["total"] if total_result else 0
    
    return {
        "items": sessions,
        "total": total,
        "page": page,
        "page_size": page_size,
        "total_pages": (total + page_size - 1) // page_size
    }

@api_router.get("/ai/sessions/{session_id}")
async def get_session_messages(session_id: str, user: dict = Depends(get_current_user)):
    """Get all messages in a specific session"""
    messages = await ai_chat_collection.find(
        {"session_id": session_id, "user_id": user["id"]},
        {"_id": 0}
    ).sort("created_at", 1).to_list(100)
    
    if not messages:
        raise HTTPException(status_code=404, detail="Session tidak ditemukan")
    
    return {
        "session_id": session_id,
        "module": messages[0].get("module") if messages else "general",
        "messages": messages,
        "message_count": len(messages)
    }

@api_router.delete("/ai/sessions/{session_id}")
async def delete_session(session_id: str, user: dict = Depends(get_current_user)):
    """Delete a specific conversation session"""
    result = await ai_chat_collection.delete_many({"session_id": session_id, "user_id": user["id"]})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Session tidak ditemukan")
    return {"message": f"Berhasil menghapus {result.deleted_count} pesan dalam sesi"}

@api_router.post("/ai/sessions/new")
async def create_new_session(module: str = "general", user: dict = Depends(get_current_user)):
    """Create a new conversation session"""
    session_id = f"tenayan-ai-{user['id']}-{uuid.uuid4()}"
    return {
        "session_id": session_id,
        "module": module,
        "message": "Sesi percakapan baru telah dibuat"
    }

# ==================== SMART STOCK ENDPOINTS ====================

class SmartStockEntry(BaseModel):
    date: str
    stock_awal: float
    suppliers: dict  # {"RIAU_MITRA": {"A": 100, "B": 200, "C": 0}, ...}
    total_penerimaan: Optional[float] = 0.0

class SumberPemakaianEntry(BaseModel):
    date: str
    stock_awal: float
    suppliers: dict  # {"RIAU_MITRA": {"UNIT1": {"A": 100, "B": 200, "C": 0}, "UNIT2": {...}}, ...}
    total_pemakaian: Optional[float] = 0.0

@api_router.get("/smart-stock")
async def get_smart_stock(
    limit: int = Query(100, ge=1, le=500),
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    user: dict = Depends(get_current_user)
):
    """Get smart stock data with optional date range filter"""
    
    query = {}
    if start_date or end_date:
        date_query = {}
        if start_date:
            date_query["$gte"] = start_date
        if end_date:
            date_query["$lte"] = end_date
        query["date"] = date_query
    
    stocks = await db.smartstock.find(query, {"_id": 0}).sort("date", -1).limit(limit).to_list(limit)
    
    # Get last 30 days for chart
    from datetime import datetime, timezone, timedelta
    thirty_days_ago = (datetime.now(timezone.utc) - timedelta(days=30)).strftime("%Y-%m-%d")
    recent_stocks = await db.smartstock.find(
        {"date": {"$gte": thirty_days_ago}},
        {"_id": 0}
    ).sort("date", 1).to_list(100)
    
    # Calculate supplier totals for stacked bar chart
    supplier_totals = {}
    for stock in recent_stocks:
        suppliers = stock.get("suppliers", {})
        for supplier_name, zones in suppliers.items():
            if supplier_name not in supplier_totals:
                supplier_totals[supplier_name] = 0
            for zone, value in zones.items():
                supplier_totals[supplier_name] += value if value else 0
    
    return {
        "data": stocks,
        "recent_30_days": recent_stocks,
        "supplier_totals": supplier_totals,
        "total_count": len(stocks)
    }

@api_router.post("/smart-stock/entry")
async def create_smart_stock_entry(
    entry: SmartStockEntry,
    user: dict = Depends(get_current_user)
):
    """Create new manual smart stock entry"""
    
    # Calculate total penerimaan from suppliers
    total = 0
    for supplier, zones in entry.suppliers.items():
        for zone, value in zones.items():
            total += value if value else 0
    
    entry.total_penerimaan = total
    
    new_entry = {
        "id": str(uuid.uuid4()),
        "date": entry.date,
        "stock_awal": entry.stock_awal,
        "suppliers": entry.suppliers,
        "total_penerimaan": entry.total_penerimaan,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.smartstock.insert_one(new_entry)
    return {"message": "Entry created successfully", "id": new_entry["id"]}

@api_router.post("/smart-stock/upload")
async def upload_smart_stock_excel(
    file: UploadFile = File(...),
    user: dict = Depends(get_current_user)
):
    """Upload Excel file and parse smart stock data"""
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="File must be Excel format (.xlsx or .xls)")
    
    try:
        contents = await file.read()
        df = pd.read_excel(io.BytesIO(contents), header=None)
        
        # Get header rows for smart parsing
        header_row_0 = df.iloc[0].tolist()
        header_row_1 = df.iloc[1].tolist() if len(df) > 1 else []
        
        logger.info(f"Header row 0 (first 15): {header_row_0[:15]}")
        
        # Keywords to SKIP - these are not supplier names
        skip_keywords = [
            'TANGGAL', 'STOCK', 'AWAL', 'SUMBER', 'PENERIMAAN', 'TOTAL', 
            'TOTALPENERIMAAN', 'TOTAL_PENERIMAAN', 'AKHIR', 'MT'
        ]
        
        # Find supplier columns - start scanning from column 4
        supplier_columns = {}
        current_supplier = None
        col_start = None
        
        for col_idx in range(4, len(header_row_0)):
            cell_value = header_row_0[col_idx]
            
            if pd.notna(cell_value) and str(cell_value).strip() != '':
                cell_str = str(cell_value).strip().upper().replace(" ", "")
                
                # Check if this is a skip keyword
                is_skip = any(kw in cell_str for kw in skip_keywords)
                
                if is_skip:
                    # End current supplier if we hit a skip keyword
                    if current_supplier and col_start is not None:
                        supplier_columns[current_supplier] = (col_start, col_idx)
                        current_supplier = None
                        col_start = None
                    continue
                
                # This is likely a supplier name
                if current_supplier and col_start is not None:
                    supplier_columns[current_supplier] = (col_start, col_idx)
                
                current_supplier = str(cell_value).strip()
                col_start = col_idx
        
        # Add last supplier
        if current_supplier and col_start is not None:
            supplier_columns[current_supplier] = (col_start, len(df.columns))
        
        logger.info(f"Found {len(supplier_columns)} suppliers: {list(supplier_columns.keys())}")
        
        # Parse data rows (starting from row 2)
        inserted_count = 0
        for idx in range(2, len(df)):
            row = df.iloc[idx]
            
            # Skip if date is empty
            if pd.isna(row.iloc[0]):
                continue
            
            # Parse date (Excel serial date or datetime)
            try:
                date_val = row.iloc[0]
                if isinstance(date_val, (int, float)):
                    date_value = pd.to_datetime(date_val, origin='1899-12-30', unit='D')
                else:
                    date_value = pd.to_datetime(date_val)
                date_str = date_value.strftime("%Y-%m-%d")
            except Exception as e:
                logger.warning(f"Date parsing error at row {idx}: {e}")
                continue
            
            # Get stock awal from column 1
            stock_awal = _safe_float(row.iloc[1])
            
            # Get total penerimaan - look for the TOTAL PENERIMAAN column
            total_penerimaan = 0.0
            for col_idx, header in enumerate(header_row_0):
                if pd.notna(header):
                    header_str = str(header).upper().replace(" ", "")
                    if 'TOTALPENERIMAAN' in header_str or header_str == 'TOTAL':
                        total_penerimaan = _safe_float(row.iloc[col_idx])
                        break
            
            # Fallback to column 3 if not found
            if total_penerimaan == 0.0:
                total_penerimaan = _safe_float(row.iloc[3]) if len(row) > 3 else 0.0
            
            # Parse suppliers data
            suppliers_data = {}
            for supplier_name, (start_col, end_col) in supplier_columns.items():
                # Normalize supplier name
                supplier_key = (supplier_name.strip()
                    .replace(" ", "_")
                    .replace("(", "")
                    .replace(")", "")
                    .replace("&", "")
                    .replace("-", "_")
                    .upper())
                
                # Get A, B, C values (exactly 3 columns per supplier)
                zones = {"A": 0.0, "B": 0.0, "C": 0.0}
                zone_keys = ["A", "B", "C"]
                
                for i, col in enumerate(range(start_col, min(start_col + 3, end_col))):
                    if col < len(row) and i < 3:
                        zones[zone_keys[i]] = _safe_float(row.iloc[col])
                
                suppliers_data[supplier_key] = zones
            
            # Calculate stock akhir
            stock_akhir = stock_awal + total_penerimaan
            
            # Check if entry already exists for this date
            existing = await db.smartstock.find_one({"date": date_str})
            
            if existing:
                # Update existing entry
                await db.smartstock.update_one(
                    {"date": date_str},
                    {"$set": {
                        "stock_awal": stock_awal,
                        "suppliers": suppliers_data,
                        "total_penerimaan": total_penerimaan,
                        "stock_akhir": stock_akhir,
                        "updated_at": datetime.now(timezone.utc).isoformat()
                    }}
                )
            else:
                # Insert new entry
                new_entry = {
                    "id": str(uuid.uuid4()),
                    "date": date_str,
                    "stock_awal": stock_awal,
                    "suppliers": suppliers_data,
                    "total_penerimaan": total_penerimaan,
                    "stock_akhir": stock_akhir,
                    "created_at": datetime.now(timezone.utc).isoformat(),
                    "updated_at": datetime.now(timezone.utc).isoformat()
                }
                await db.smartstock.insert_one(new_entry)
            
            inserted_count += 1
        
        return {
            "message": f"Successfully processed {inserted_count} entries",
            "count": inserted_count,
            "suppliers_found": list(supplier_columns.keys())
        }
    
    except Exception as e:
        logger.error(f"Error processing Excel file: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error processing file: {str(e)}")

def _safe_float(value) -> float:
    """Safely convert value to float"""
    if pd.isna(value) or value == '' or value is None:
        return 0.0
    try:
        return float(value)
    except (ValueError, TypeError):
        return 0.0

@api_router.delete("/smart-stock/{entry_id}")
async def delete_smart_stock_entry(
    entry_id: str,
    user: dict = Depends(get_current_user)
):
    """Delete a smart stock entry"""
    
    result = await db.smartstock.delete_one({"id": entry_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Entry not found")
    
    return {"message": "Entry deleted successfully"}

@api_router.delete("/smart-stock")
async def delete_all_smart_stock(
    user: dict = Depends(get_current_user)
):
    """Delete all smart stock entries"""
    
    result = await db.smartstock.delete_many({})
    
    return {
        "message": f"Successfully deleted {result.deleted_count} entries",
        "deleted_count": result.deleted_count
    }

# ==================== SUMBER PEMAKAIAN ENDPOINTS ====================

@api_router.get("/sumber-pemakaian")
async def get_sumber_pemakaian(
    limit: int = Query(100, ge=1, le=500),
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    user: dict = Depends(get_current_user)
):
    """Get sumber pemakaian data with optional date range filter"""
    
    query = {}
    if start_date or end_date:
        date_query = {}
        if start_date:
            date_query["$gte"] = start_date
        if end_date:
            date_query["$lte"] = end_date
        query["date"] = date_query
    
    pemakaian = await db.sumberpemakaian.find(query, {"_id": 0}).sort("date", -1).limit(limit).to_list(limit)
    
    # Get last 30 days for chart
    from datetime import datetime, timezone, timedelta
    thirty_days_ago = (datetime.now(timezone.utc) - timedelta(days=30)).strftime("%Y-%m-%d")
    recent_pemakaian = await db.sumberpemakaian.find(
        {"date": {"$gte": thirty_days_ago}},
        {"_id": 0}
    ).sort("date", 1).to_list(100)
    
    # Calculate stats from LATEST ENTRY (not today's date)
    latest_data = await db.sumberpemakaian.find_one({}, {"_id": 0}, sort=[("date", -1)])
    
    stats = {
        "total_burn_today": 0,
        "unit1_load": 0,
        "unit2_load": 0,
        "dominant_source": "N/A",
        "latest_date": None
    }
    
    if latest_data and latest_data.get("suppliers"):
        stats["latest_date"] = latest_data["date"]
        supplier_totals = {}
        
        for supplier, units in latest_data["suppliers"].items():
            supplier_total = 0
            for unit, zones in units.items():
                unit_total = sum(zones.values())
                if unit == "UNIT1":
                    stats["unit1_load"] += unit_total
                elif unit == "UNIT2":
                    stats["unit2_load"] += unit_total
                supplier_total += unit_total
            supplier_totals[supplier] = supplier_total
        
        stats["total_burn_today"] = stats["unit1_load"] + stats["unit2_load"]
        
        # Find dominant source (highest total)
        if supplier_totals:
            dominant = max(supplier_totals.items(), key=lambda x: x[1])
            if dominant[1] > 0:  # Only show if there's actual consumption
                stats["dominant_source"] = dominant[0].replace("_", " ")
    
    return {
        "data": pemakaian,
        "recent_30_days": recent_pemakaian,
        "stats": stats,
        "total_count": len(pemakaian)
    }

@api_router.post("/sumber-pemakaian/entry")
async def create_sumber_pemakaian_entry(
    entry: SumberPemakaianEntry,
    user: dict = Depends(get_current_user)
):
    """Create new manual sumber pemakaian entry"""
    
    # Calculate total pemakaian from suppliers
    total = 0
    for supplier, units in entry.suppliers.items():
        for unit, zones in units.items():
            for zone, value in zones.items():
                total += value if value else 0
    
    entry.total_pemakaian = total
    
    new_entry = {
        "id": str(uuid.uuid4()),
        "date": entry.date,
        "stock_awal": entry.stock_awal,
        "suppliers": entry.suppliers,
        "total_pemakaian": entry.total_pemakaian,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.sumberpemakaian.insert_one(new_entry)
    return {"message": "Entry created successfully", "id": new_entry["id"]}

@api_router.post("/sumber-pemakaian/upload")
async def upload_sumber_pemakaian_excel(
    file: UploadFile = File(...),
    user: dict = Depends(get_current_user)
):
    """Upload Excel file and parse sumber pemakaian data with correct column mapping"""
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="File must be Excel format (.xlsx or .xls)")
    
    try:
        contents = await file.read()
        df = pd.read_excel(io.BytesIO(contents), header=None)
        
        # CORRECT Column mapping for Sumber Pemakaian.xlsx:
        # Row 0: TANGGAL | STOCK AWAL | (merged cells, not used)
        # Row 1: Supplier names starting from Col 2
        # Row 2: UNIT1/UNIT2 indicators
        # Row 3: A, B, C zone indicators
        # Row 4+: Data rows
        
        # Get supplier names from ROW 1 (starting from column 2)
        suppliers_row = df.iloc[1, 2:].tolist()
        
        # Identify unique suppliers and count columns per supplier
        supplier_columns = []
        current_supplier = None
        col_count = 0
        
        for i, cell_value in enumerate(suppliers_row):
            actual_col = i + 2
            if pd.notna(cell_value) and str(cell_value).strip() != '':
                if current_supplier and col_count > 0:
                    supplier_columns.append((current_supplier, col_count))
                current_supplier = str(cell_value).strip()
                col_count = 1
            else:
                col_count += 1
        
        # Add last supplier
        if current_supplier:
            supplier_columns.append((current_supplier, col_count))
        
        logger.info(f"Found suppliers for pemakaian: {[s[0] for s in supplier_columns]}")
        
        # Parse data rows (starting from row 4)
        inserted_count = 0
        for idx in range(4, len(df)):
            row = df.iloc[idx]
            
            if pd.isna(row[0]):
                continue
            
            # Parse date
            try:
                if isinstance(row[0], (int, float)):
                    date_value = pd.to_datetime(row[0], origin='1899-12-30', unit='D')
                else:
                    date_value = pd.to_datetime(row[0])
                date_str = date_value.strftime("%Y-%m-%d")
            except Exception as e:
                logger.warning(f"Date parsing error at row {idx}: {e}")
                continue
            
            stock_awal = float(row[1]) if pd.notna(row[1]) and row[1] != '' else 0.0
            
            # Parse suppliers data
            suppliers_data = {}
            col_offset = 2
            
            for supplier_name, num_cols in supplier_columns:
                supplier_key = supplier_name.strip().replace(" ", "_").replace("(", "").replace(")", "").replace("&", "").replace("-", "_").upper()
                
                # Each supplier has 6 columns: UNIT1 (A,B,C) and UNIT2 (A,B,C)
                units_data = {
                    "UNIT1": {"A": 0.0, "B": 0.0, "C": 0.0},
                    "UNIT2": {"A": 0.0, "B": 0.0, "C": 0.0}
                }
                
                # Read the columns for this supplier
                for i in range(min(num_cols, 6)):
                    col = col_offset + i
                    if col < len(row):
                        unit = "UNIT1" if i < 3 else "UNIT2"
                        zone = ["A", "B", "C"][i % 3]
                        try:
                            units_data[unit][zone] = float(row[col]) if pd.notna(row[col]) and row[col] != '' else 0.0
                        except:
                            units_data[unit][zone] = 0.0
                
                suppliers_data[supplier_key] = units_data
                col_offset += num_cols
            
            # Calculate total pemakaian: SUM from column 2 onwards
            total_pemakaian = 0
            for col in range(2, len(row)):
                try:
                    val = float(row[col]) if pd.notna(row[col]) and row[col] != '' else 0.0
                    total_pemakaian += val
                except:
                    pass
            
            # Check if entry already exists
            existing = await db.sumberpemakaian.find_one({"date": date_str})
            
            if existing:
                await db.sumberpemakaian.update_one(
                    {"date": date_str},
                    {"$set": {
                        "stock_awal": stock_awal,
                        "suppliers": suppliers_data,
                        "total_pemakaian": total_pemakaian,
                        "updated_at": datetime.now(timezone.utc).isoformat()
                    }}
                )
            else:
                new_entry = {
                    "id": str(uuid.uuid4()),
                    "date": date_str,
                    "stock_awal": stock_awal,
                    "suppliers": suppliers_data,
                    "total_pemakaian": total_pemakaian,
                    "created_at": datetime.now(timezone.utc).isoformat(),
                    "updated_at": datetime.now(timezone.utc).isoformat()
                }
                await db.sumberpemakaian.insert_one(new_entry)
            
            inserted_count += 1
        
        return {
            "message": f"Successfully processed {inserted_count} entries",
            "count": inserted_count
        }
    
    except Exception as e:
        logger.error(f"Error processing Excel file: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error processing file: {str(e)}")

@api_router.delete("/sumber-pemakaian")
async def delete_all_sumber_pemakaian(
    user: dict = Depends(get_current_user)
):
    """Delete all sumber pemakaian entries"""
    
    result = await db.sumberpemakaian.delete_many({})
    
    return {
        "message": f"Successfully deleted {result.deleted_count} entries",
        "deleted_count": result.deleted_count
    }

# ==================== SMART BLENDING AI ENDPOINTS ====================

class SmartBlendingRequest(BaseModel):
    target_gcv: float  # Target GCV in kcal/kg (3700-4700)
    max_ash: float  # Max Ash content in % (3.3-6.0)
    max_sulphur: float  # Max Sulphur in % (0.13-2.2)
    max_total_moisture: float = 35.0  # Max Total Moisture in % (25-40)
    max_inherent_moisture: float = 18.0  # Max Inherent Moisture in % (13.8-25)
    min_volatile_matter: float = 35.0  # Min Volatile Matter in % (27.9-40)
    min_fixed_carbon: float = 25.0  # Min Fixed Carbon in % (23-41)
    target_quantity: float  # Target quantity in MT

@api_router.post("/smart-blending/recommend")
async def get_smart_blending_recommendation(
    request: SmartBlendingRequest,
    user: dict = Depends(get_current_user)
):
    """Get AI-powered smart blending recommendation using Gemini"""
    
    try:
        # Calculate 6 months ago date for filtering suppliers
        from dateutil.relativedelta import relativedelta
        six_months_ago = datetime.now(timezone.utc) - relativedelta(months=6)
        
        # 1. Fetch quality data from suppliers in the last 6 months
        # Filter by various date fields that exist in the collections
        date_filter_vessel = {
            "$or": [
                {"time_arrival": {"$gte": six_months_ago.isoformat()}},
                {"periode_realisasi": {"$gte": six_months_ago.isoformat()}},
                {"created_at": {"$gte": six_months_ago.isoformat()}}
            ]
        }
        date_filter_barge = {
            "$or": [
                {"ta": {"$gte": six_months_ago.isoformat()}},
                {"periode": {"$gte": six_months_ago.isoformat()}},
                {"created_at": {"$gte": six_months_ago.isoformat()}}
            ]
        }
        date_filter_trucking = {
            "$or": [
                {"ta": {"$gte": six_months_ago.isoformat()}},
                {"periode_ta": {"$gte": six_months_ago.isoformat()}},
                {"created_at": {"$gte": six_months_ago.isoformat()}}
            ]
        }
        
        vessels = await db.vessels.find(date_filter_vessel, {"_id": 0}).sort("time_arrival", -1).limit(50).to_list(50)
        barges = await db.barges.find(date_filter_barge, {"_id": 0}).sort("ta", -1).limit(50).to_list(50)
        trucking = await db.trucking.find(date_filter_trucking, {"_id": 0}).sort("ta", -1).limit(50).to_list(50)
        
        # 2. Fetch stock availability from Smart Stock (Sumber Penerimaan)
        latest_stock = await db.smartstock.find_one({}, {"_id": 0}, sort=[("date", -1)])
        
        # 3. Fetch pricing from Merit Order
        merit_order = await db.merit_order.find({}, {"_id": 0}).to_list(100)
        
        # 4. Prepare data for AI analysis
        coal_inventory = []
        
        # Process vessels
        for v in vessels:
            supplier_name = v.get("suppliers", "Unknown")
            if supplier_name and supplier_name != "Unknown":
                coal_inventory.append({
                    "source": "Vessel",
                    "supplier": supplier_name,
                    "type": "MRC" if "MRC" in supplier_name.upper() else "LRC",
                    "gcv_ar": v.get("gcv_arb"),
                    "ash_ar": v.get("ash_arb"),
                    "ts_ar": v.get("ts_arb"),
                    "tm_ar": v.get("tm_arb"),  # Total Moisture
                    "im_adb": v.get("im_adb"),  # Inherent Moisture
                    "vm_ar": v.get("vm_arb"),  # Volatile Matter
                    "fc_ar": v.get("fc_arb"),  # Fixed Carbon
                    "available_mt": v.get("bl_mt", 0),
                    "date": v.get("time_arrival")
                })
        
        # Process barges
        for b in barges:
            supplier_name = b.get("suppliers", "Unknown")
            if supplier_name and supplier_name != "Unknown":
                coal_inventory.append({
                    "source": "Barge",
                    "supplier": supplier_name,
                    "type": "MRC" if "MRC" in supplier_name.upper() else "LRC",
                    "gcv_ar": b.get("gcv_arb"),
                    "ash_ar": b.get("ash_arb"),
                    "ts_ar": b.get("ts_arb"),
                    "tm_ar": b.get("tm_arb"),  # Total Moisture
                    "im_adb": b.get("im_adb"),  # Inherent Moisture
                    "vm_ar": b.get("vm_arb"),  # Volatile Matter
                    "fc_ar": b.get("fc_arb"),  # Fixed Carbon
                    "available_mt": b.get("bl_mt", 0),
                    "date": b.get("time_arrival")
                })
        
        # Process trucking
        for t in trucking:
            supplier_name = t.get("suppliers", "Unknown")
            if supplier_name and supplier_name != "Unknown":
                coal_inventory.append({
                    "source": "Trucking",
                    "supplier": supplier_name,
                    "type": "MRC" if "MRC" in supplier_name.upper() else "LRC",
                    "gcv_ar": t.get("gcv_arb"),
                    "ash_ar": t.get("ash_arb"),
                    "ts_ar": t.get("ts_arb"),
                    "tm_ar": t.get("tm_arb"),  # Total Moisture
                    "im_adb": t.get("im_adb"),  # Inherent Moisture
                    "vm_ar": t.get("vm_arb"),  # Volatile Matter
                    "fc_ar": t.get("fc_arb"),  # Fixed Carbon
                    "available_mt": t.get("quantity", 0),
                    "date": t.get("date_received")
                })
        
        # 5. Create AI prompt for Gemini
        ai_prompt = f"""Anda adalah Ahli Kimia Batubara Digital untuk PLTU Tenayan. Tugas Anda adalah memberikan rekomendasi blending batubara yang optimal.

**PERSYARATAN BLENDING:**
- Target GCV: {request.target_gcv} kcal/kg (As Received)
- Maksimal Kandungan Abu: {request.max_ash}% (As Received)
- Maksimal Total Sulphur: {request.max_sulphur}% (As Received)
- Maksimal Total Moisture: {request.max_total_moisture}% (As Received)
- Maksimal Inherent Moisture: {request.max_inherent_moisture}% (Air Dried Basis)
- Minimal Volatile Matter: {request.min_volatile_matter}% (As Received)
- Minimal Fixed Carbon: {request.min_fixed_carbon}% (As Received)
- Total Kuantitas yang Dibutuhkan: {request.target_quantity:,.0f} MT

**INVENTORI BATUBARA TERSEDIA:**
{json.dumps(coal_inventory, indent=2)}

**CATATAN PENTING:**
- LRC (Low Rank Coal): GCV lebih rendah (3000-4500 kcal/kg), moisture tinggi, lebih murah
- MRC (Medium Rank Coal): GCV lebih tinggi (4500-6000 kcal/kg), moisture rendah, lebih mahal
- Formula blending: Hasil_GCV = (Batubara1_GCV × Batubara1_%) + (Batubara2_GCV × Batubara2_%)
- Prioritas: CAPAI TARGET GCV DULU, kemudian pertimbangkan parameter lainnya
- WAJIB: Gunakan NAMA SUPPLIER ASLI dari data di atas, JANGAN buat nama dummy seperti "Supplier Alpha"
- tm_ar = Total Moisture (As Received), im_adb = Inherent Moisture (Air Dried Basis)
- vm_ar = Volatile Matter (As Received), fc_ar = Fixed Carbon (As Received)

**TUGAS ANDA:**
1. Analisis batubara yang tersedia (pertimbangkan semua parameter: GCV, Abu, Sulphur, Moisture, VM, FC)
2. Hitung persentase blend yang optimal
3. Pastikan target GCV tercapai
4. Jaga semua parameter dalam batas yang ditentukan
5. Rekomendasikan 2-4 batubara untuk blending
6. GUNAKAN nama supplier SEBENARNYA dari data (contoh: "PT PLN BATUBARA", "PT BUKIT ASAM")

**FORMAT OUTPUT (JSON):**
{{
  "recommendation": [
    {{
      "supplier": "PT PLN BATUBARA",
      "source": "Vessel",
      "type": "LRC",
      "percentage": 60.0,
      "tonnage": 6000.0,
      "gcv": 4277,
      "ash": 4.4,
      "sulphur": 0.21,
      "total_moisture": 34.0,
      "inherent_moisture": 15.3,
      "volatile_matter": 31.9,
      "fixed_carbon": 29.7
    }}
  ],
  "predicted_quality": {{
    "gcv": 4050,
    "ash": 4.8,
    "sulphur": 0.25,
    "total_moisture": 35.0,
    "inherent_moisture": 16.0,
    "volatile_matter": 33.0,
    "fixed_carbon": 27.0
  }},
  "meets_target": true,
  "reasoning": "Jelaskan mengapa blend ini optimal dalam Bahasa Indonesia",
  "cost_warning": "Catatan jika menggunakan batubara mahal (dalam Bahasa Indonesia)"
}}

Respons HANYA dengan JSON yang valid, tanpa teks tambahan. WAJIB gunakan nama supplier asli dari data."""

        # 6. Get API key from user settings or use default
        user_settings = await db.user_settings.find_one({"user_id": user["id"]}, {"_id": 0})
        llm_key = None
        if user_settings and user_settings.get("custom_api_key"):
            llm_key = user_settings.get("custom_api_key")
        if not llm_key:
            llm_key = os.environ.get("EMERGENT_LLM_KEY")
        
        if not llm_key:
            raise HTTPException(status_code=400, detail="API Key tidak ditemukan. Silakan konfigurasi di Pengaturan AI Intelligence.")
        
        chat = LlmChat(
            api_key=llm_key,
            session_id=f"smart-blending-{uuid.uuid4()}",
            system_message="You are an expert coal blending optimization AI for power plants."
        ).with_model("gemini", "gemini-2.5-flash")
        
        user_message = UserMessage(text=ai_prompt)
        response = await chat.send_message(user_message)
        
        # 7. Parse AI response
        try:
            # Clean response (remove markdown code blocks if present)
            clean_response = response.strip()
            if clean_response.startswith("```json"):
                clean_response = clean_response[7:]
            if clean_response.startswith("```"):
                clean_response = clean_response[3:]
            if clean_response.endswith("```"):
                clean_response = clean_response[:-3]
            clean_response = clean_response.strip()
            
            ai_result = json.loads(clean_response)
        except json.JSONDecodeError:
            # If JSON parsing fails, return raw response
            ai_result = {
                "error": "Failed to parse AI response",
                "raw_response": response
            }
        
        return {
            "request": request.dict(),
            "ai_recommendation": ai_result,
            "data_sources": {
                "vessels_count": len(vessels),
                "barges_count": len(barges),
                "trucking_count": len(trucking),
                "latest_stock_date": latest_stock.get("date") if latest_stock else None
            }
        }
    
    except Exception as e:
        logger.error(f"Smart Blending AI error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"AI recommendation failed: {str(e)}")

# ==================== COA SETTINGS ====================

class COASettingsUpdate(BaseModel):
    price_per_kcal_per_ton: float

@api_router.get("/settings/coa")
async def get_coa_settings(user: dict = Depends(get_current_user)):
    """Get COA settings including price per kCal"""
    settings = await db.app_settings.find_one({"type": "coa"}, {"_id": 0})
    if not settings:
        return {"price_per_kcal_per_ton": None}
    return settings

@api_router.put("/settings/coa")
async def update_coa_settings(data: COASettingsUpdate, user: dict = Depends(require_role(["admin"]))):
    """Update COA settings"""
    await db.app_settings.update_one(
        {"type": "coa"},
        {"$set": {
            "type": "coa",
            "price_per_kcal_per_ton": data.price_per_kcal_per_ton,
            "updated_at": datetime.now(timezone.utc).isoformat(),
            "updated_by": user["id"]
        }},
        upsert=True
    )
    return {"message": "Pengaturan COA berhasil disimpan", "price_per_kcal_per_ton": data.price_per_kcal_per_ton}

# ==================== COA RECONCILIATION & DISPUTE MONITOR ====================

from services.coa_reconciliation import (
    parse_coa_excel, 
    merge_coa_data, 
    calculate_kpis, 
    get_gcv_trend_data,
    get_radar_chart_data
)

class UmpireProposal(BaseModel):
    reconciliation_id: str
    sample_number: str
    notes: Optional[str] = None

@api_router.get("/coa-reconciliation")
async def get_coa_reconciliation(
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=500),
    status: Optional[str] = None,
    search: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    user: dict = Depends(get_current_user)
):
    """Get COA reconciliation data with pagination and date filter"""
    query = {}
    if status and status != "all":
        query["status"] = status
    if search:
        query["$or"] = [
            {"shipment": {"$regex": search, "$options": "i"}},
            {"suppliers": {"$regex": search, "$options": "i"}}
        ]
    
    # Date range filter on completed_unloading
    if date_from or date_to:
        date_filter = {}
        if date_from:
            date_filter["$gte"] = date_from
        if date_to:
            date_filter["$lte"] = date_to + "T23:59:59"
        if date_filter:
            query["completed_unloading"] = date_filter
    
    skip = (page - 1) * page_size
    total = await db.coa_reconciliation.count_documents(query)
    # Sort by completed_unloading descending (newest first)
    items = await db.coa_reconciliation.find(query, {"_id": 0}).sort("completed_unloading", -1).skip(skip).limit(page_size).to_list(page_size)
    
    return {
        "items": items,
        "total": total,
        "page": page,
        "page_size": page_size,
        "total_pages": (total + page_size - 1) // page_size
    }

@api_router.get("/coa-reconciliation/kpis")
async def get_coa_kpis(user: dict = Depends(get_current_user)):
    """Get KPIs for COA Reconciliation dashboard"""
    all_data = await db.coa_reconciliation.find({}, {"_id": 0}).to_list(10000)
    
    # Get price setting from database
    coa_settings = await db.app_settings.find_one({"type": "coa"}, {"_id": 0})
    price_per_kcal = coa_settings.get("price_per_kcal_per_ton") if coa_settings else None
    
    if not all_data:
        return {
            "total_records": 0,
            "high_deviation_count": 0,
            "potential_loss_rp": 0,
            "total_tonnage_problem": 0,
            "umpire_status": {"proposed": 0, "in_progress": 0, "completed": 0, "total": 0},
            "supplier_deviations": [],
            "worst_supplier": None,
            "avg_accuracy": 100,
            "critical_count": 0,
            "warning_count": 0,
            "normal_count": 0,
            "price_per_kcal_per_ton": price_per_kcal,
            "price_not_set": price_per_kcal is None
        }
    
    kpis = calculate_kpis(all_data, price_per_kcal)
    kpis["price_per_kcal_per_ton"] = price_per_kcal
    kpis["price_not_set"] = price_per_kcal is None
    return kpis

@api_router.get("/coa-reconciliation/trend")
async def get_coa_trend(months: int = Query(3, ge=1, le=12), user: dict = Depends(get_current_user)):
    """Get GCV trend data for line chart"""
    all_data = await db.coa_reconciliation.find({}, {"_id": 0}).to_list(10000)
    return get_gcv_trend_data(all_data, months)

@api_router.get("/coa-reconciliation/supplier-consistency")
async def get_supplier_consistency(user: dict = Depends(get_current_user)):
    """Get supplier consistency data for bar chart"""
    all_data = await db.coa_reconciliation.find({}, {"_id": 0}).to_list(10000)
    kpis = calculate_kpis(all_data)
    return kpis.get("supplier_deviations", [])

# Dispute Monitor endpoints - MUST be before {record_id}
@api_router.get("/coa-reconciliation/dispute-monitor")
async def get_dispute_monitor(
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=500),
    umpire_status: Optional[str] = None,
    user: dict = Depends(get_current_user)
):
    """Get all records that have umpire activity (proposed, in_progress, or completed)"""
    query = {"umpire_status": {"$ne": "none"}}
    if umpire_status and umpire_status != "all":
        query["umpire_status"] = umpire_status
    
    skip = (page - 1) * page_size
    total = await db.coa_reconciliation.count_documents(query)
    items = await db.coa_reconciliation.find(query, {"_id": 0}).sort("umpire_proposed_at", -1).skip(skip).limit(page_size).to_list(page_size)
    
    # Calculate summary stats
    all_disputes = await db.coa_reconciliation.find({"umpire_status": {"$ne": "none"}}, {"_id": 0, "umpire_status": 1}).to_list(10000)
    summary = {
        "proposed": sum(1 for d in all_disputes if d.get("umpire_status") == "proposed"),
        "in_progress": sum(1 for d in all_disputes if d.get("umpire_status") == "in_progress"),
        "completed": sum(1 for d in all_disputes if d.get("umpire_status") == "completed"),
        "total": len(all_disputes)
    }
    
    return {
        "items": items,
        "total": total,
        "page": page,
        "page_size": page_size,
        "total_pages": (total + page_size - 1) // page_size,
        "summary": summary
    }

@api_router.get("/coa-reconciliation/{record_id}")
async def get_coa_reconciliation_detail(record_id: str, user: dict = Depends(get_current_user)):
    """Get single COA reconciliation record with radar chart data"""
    record = await db.coa_reconciliation.find_one({"id": record_id}, {"_id": 0})
    if not record:
        raise HTTPException(status_code=404, detail="Data tidak ditemukan")
    
    radar_data = get_radar_chart_data(record)
    return {
        "record": record,
        "radar_chart": radar_data
    }

@api_router.get("/coa-reconciliation/shipment/{shipment}")
async def get_coa_by_shipment(shipment: str, user: dict = Depends(get_current_user)):
    """Get COA reconciliation record by shipment ID (string)"""
    record = await db.coa_reconciliation.find_one({"shipment": shipment}, {"_id": 0})
    if not record:
        raise HTTPException(status_code=404, detail="Data tidak ditemukan")
    
    radar_data = get_radar_chart_data(record)
    return {
        "record": record,
        "radar_chart": radar_data
    }

@api_router.post("/coa-reconciliation/propose-umpire")
async def propose_umpire(data: UmpireProposal, user: dict = Depends(require_role(["admin", "operator"]))):
    """Propose umpire testing for a reconciliation record"""
    result = await db.coa_reconciliation.update_one(
        {"id": data.reconciliation_id},
        {"$set": {
            "umpire_status": "proposed",
            "umpire_sample_number": data.sample_number,
            "umpire_notes": data.notes,
            "umpire_proposed_at": datetime.now(timezone.utc).isoformat(),
            "umpire_proposed_by": user["id"]
        }}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Data tidak ditemukan")
    return {"message": "Umpire testing berhasil diajukan", "sample_number": data.sample_number}

@api_router.post("/coa-reconciliation/update-umpire-status/{record_id}")
async def update_umpire_status(
    record_id: str, 
    status: str = Query(..., regex="^(none|proposed|in_progress|completed)$"),
    user: dict = Depends(require_role(["admin", "operator"]))
):
    """Update umpire testing status"""
    update_data = {"umpire_status": status}
    if status == "completed":
        update_data["umpire_completed_at"] = datetime.now(timezone.utc).isoformat()
    if status == "in_progress":
        update_data["umpire_started_at"] = datetime.now(timezone.utc).isoformat()
    
    result = await db.coa_reconciliation.update_one({"id": record_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Data tidak ditemukan")
    return {"message": f"Status umpire berhasil diubah ke {status}"}

class UmpireResultInput(BaseModel):
    reconciliation_id: str
    umpire_gcv_arb: float
    umpire_tm_arb: Optional[float] = None
    umpire_ash_arb: Optional[float] = None
    umpire_ts_arb: Optional[float] = None
    umpire_lab_name: str
    umpire_result_date: str
    notes: Optional[str] = None

@api_router.post("/coa-reconciliation/submit-umpire-result")
async def submit_umpire_result(data: UmpireResultInput, user: dict = Depends(require_role(["admin", "operator"]))):
    """Submit umpire test results for a reconciliation record"""
    result = await db.coa_reconciliation.update_one(
        {"id": data.reconciliation_id},
        {"$set": {
            "umpire_status": "completed",
            "umpire_gcv_arb": data.umpire_gcv_arb,
            "umpire_tm_arb": data.umpire_tm_arb,
            "umpire_ash_arb": data.umpire_ash_arb,
            "umpire_ts_arb": data.umpire_ts_arb,
            "umpire_lab_name": data.umpire_lab_name,
            "umpire_result_date": data.umpire_result_date,
            "umpire_result_notes": data.notes,
            "umpire_completed_at": datetime.now(timezone.utc).isoformat(),
            "umpire_completed_by": user["id"]
        }}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Data tidak ditemukan")
    return {"message": "Hasil umpire berhasil disimpan", "reconciliation_id": data.reconciliation_id}

@api_router.post("/coa-reconciliation/upload")
async def upload_coa_files(
    loading_file: UploadFile = File(..., description="Loading.xlsx"),
    unloading_file: UploadFile = File(..., description="Unloading.xlsx"),
    internal_file: UploadFile = File(..., description="Lab Internal.xlsx"),
    user: dict = Depends(require_role(["admin", "operator"]))
):
    """Upload and process COA files (Loading, Unloading, Lab Internal)"""
    try:
        # Parse all three files
        loading_contents = await loading_file.read()
        unloading_contents = await unloading_file.read()
        internal_contents = await internal_file.read()
        
        loading_data = parse_coa_excel(loading_contents, "loading")
        unloading_data = parse_coa_excel(unloading_contents, "unloading")
        internal_data = parse_coa_excel(internal_contents, "internal")
        
        # Merge data
        merged_data = merge_coa_data(loading_data, unloading_data, internal_data)
        
        # Save to database (replace existing data)
        if merged_data:
            # Clear existing data
            await db.coa_reconciliation.delete_many({})
            # Insert new data
            for record in merged_data:
                record["uploaded_by"] = user["id"]
                record["uploaded_at"] = datetime.now(timezone.utc).isoformat()
            await db.coa_reconciliation.insert_many(merged_data)
        
        return {
            "message": f"Berhasil memproses dan menyimpan {len(merged_data)} data rekonsiliasi COA",
            "count": len(merged_data),
            "sources": {
                "loading": len(loading_data),
                "unloading": len(unloading_data),
                "internal": len(internal_data)
            }
        }
    except Exception as e:
        logger.error(f"Error processing COA files: {e}")
        raise HTTPException(status_code=400, detail=f"Gagal memproses file: {str(e)}")

@api_router.delete("/coa-reconciliation")
async def delete_all_coa_reconciliation(user: dict = Depends(require_role(["admin"]))):
    """Delete all COA reconciliation data"""
    result = await db.coa_reconciliation.delete_many({})
    return {"message": f"Berhasil menghapus {result.deleted_count} data rekonsiliasi COA", "count": result.deleted_count}

class COAManualInput(BaseModel):
    shipment: str  # Changed to string to support "Lot XX" format
    suppliers: str
    periode: Optional[str] = None
    tb: Optional[str] = None
    bg: Optional[str] = None
    ds_mt: Optional[float] = None
    completed_unloading: Optional[str] = None
    loading_gcv_arb: Optional[float] = None
    loading_tm_arb: Optional[float] = None
    loading_ash_arb: Optional[float] = None
    loading_ts_arb: Optional[float] = None
    unloading_gcv_arb: Optional[float] = None
    unloading_tm_arb: Optional[float] = None
    unloading_ash_arb: Optional[float] = None
    unloading_ts_arb: Optional[float] = None
    internal_gcv_arb: Optional[float] = None
    internal_tm_arb: Optional[float] = None
    internal_ash_arb: Optional[float] = None
    internal_ts_arb: Optional[float] = None

@api_router.post("/coa-reconciliation/manual")
async def add_coa_manual(data: COAManualInput, user: dict = Depends(require_role(["admin", "operator"]))):
    """Add COA reconciliation data manually"""
    import uuid
    
    # Check if shipment already exists
    existing = await db.coa_reconciliation.find_one({"shipment": data.shipment})
    if existing:
        raise HTTPException(status_code=400, detail=f"Shipment {data.shipment} sudah ada. Gunakan fitur edit atau hapus terlebih dahulu.")
    
    # Calculate deltas
    delta_loading_internal = None
    delta_unloading_internal = None
    delta_loading_unloading = None
    
    if data.loading_gcv_arb and data.internal_gcv_arb:
        delta_loading_internal = data.loading_gcv_arb - data.internal_gcv_arb
    if data.unloading_gcv_arb and data.internal_gcv_arb:
        delta_unloading_internal = data.unloading_gcv_arb - data.internal_gcv_arb
    if data.loading_gcv_arb and data.unloading_gcv_arb:
        delta_loading_unloading = data.loading_gcv_arb - data.unloading_gcv_arb
    
    # Determine status (Loading vs Internal)
    # KRITIS: Loading > Internal (supplier overclaim, Anda RUGI)
    status = "normal"
    if delta_loading_internal is not None:
        if delta_loading_internal > 150:  # Hanya positif = RUGI
            status = "critical"
        elif delta_loading_internal > 100:
            status = "warning"
    
    record = {
        "id": str(uuid.uuid4()),
        "shipment": data.shipment,  # Now string
        "suppliers": data.suppliers,
        "periode": data.periode or datetime.now(timezone.utc).strftime("%Y-%m-%d"),
        "tb": data.tb or "",
        "bg": data.bg or "",
        "ds_mt": data.ds_mt,
        "completed_unloading": data.completed_unloading,
        "loading_gcv_arb": data.loading_gcv_arb,
        "loading_tm_arb": data.loading_tm_arb,
        "loading_ash_arb": data.loading_ash_arb,
        "loading_ts_arb": data.loading_ts_arb,
        "loading_surveyor": None,
        "loading_no_coa": None,
        "unloading_gcv_arb": data.unloading_gcv_arb,
        "unloading_tm_arb": data.unloading_tm_arb,
        "unloading_ash_arb": data.unloading_ash_arb,
        "unloading_ts_arb": data.unloading_ts_arb,
        "unloading_surveyor": None,
        "unloading_slagging": None,
        "unloading_fouling": None,
        "internal_gcv_arb": data.internal_gcv_arb,
        "internal_tm_arb": data.internal_tm_arb,
        "internal_ash_arb": data.internal_ash_arb,
        "internal_ts_arb": data.internal_ts_arb,
        "delta_loading_internal": delta_loading_internal,
        "delta_unloading_internal": delta_unloading_internal,
        "delta_loading_unloading": delta_loading_unloading,
        "status": status,
        "umpire_status": "none",
        "umpire_sample_number": None,
        "umpire_proposed_at": None,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "created_by": user["id"],
        "input_method": "manual"
    }
    
    await db.coa_reconciliation.insert_one(record)
    return {"message": f"Berhasil menambahkan data COA Shipment {data.shipment}", "id": record["id"], "status": status}

@api_router.get("/coa-reconciliation/export/excel")
async def export_coa_to_excel(
    status_filter: str = Query("all"),
    user: dict = Depends(get_current_user)
):
    """Export COA reconciliation data to Excel"""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    # Build query
    query = {}
    if status_filter == "kritis":
        query["status"] = "Kritis"
    elif status_filter == "umpire":
        query["umpire_status"] = {"$nin": [None, "none", ""]}
    
    # Get data
    data = await db.coa_reconciliation.find(query, {"_id": 0}).sort("completed_unloading", -1).to_list(10000)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "COA Reconciliation"
    
    # Styles
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    kritis_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = [
        "No", "Shipment", "Supplier", "Tanggal Unloading", "DS (MT)",
        "Loading GCV", "Loading TM", "Loading Ash", "Loading S",
        "Unloading GCV", "Unloading TM", "Unloading Ash", "Unloading S",
        "Internal GCV", "Internal TM", "Internal Ash", "Internal S",
        "Delta GCV", "Status", "Umpire Status", "Umpire GCV"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Data rows
    for row_idx, record in enumerate(data, 2):
        row_data = [
            row_idx - 1,
            record.get("shipment", ""),
            record.get("suppliers", ""),
            record.get("completed_unloading", "")[:10] if record.get("completed_unloading") else "",
            record.get("ds_mt", 0),
            record.get("loading_gcv_arb"),
            record.get("loading_tm_arb"),
            record.get("loading_ash_arb"),
            record.get("loading_ts_arb"),
            record.get("unloading_gcv_arb"),
            record.get("unloading_tm_arb"),
            record.get("unloading_ash_arb"),
            record.get("unloading_ts_arb"),
            record.get("internal_gcv_arb"),
            record.get("internal_tm_arb"),
            record.get("internal_ash_arb"),
            record.get("internal_ts_arb"),
            record.get("delta_loading_internal"),
            record.get("status", "normal").upper(),
            record.get("umpire_status", "none"),
            record.get("umpire_result", {}).get("gcv") if record.get("umpire_result") else None
        ]
        
        is_kritis = record.get("status") == "Kritis"
        
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = border
            cell.alignment = Alignment(horizontal='center' if col <= 4 else 'right')
            if is_kritis:
                cell.fill = kritis_fill
    
    # Adjust column widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15
    ws.column_dimensions['B'].width = 20  # Shipment
    ws.column_dimensions['C'].width = 30  # Supplier
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"COA_Reconciliation_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@api_router.get("/coa-reconciliation/export/pdf")
async def export_coa_to_pdf(
    status_filter: str = Query("all"),
    user: dict = Depends(get_current_user)
):
    """Export COA reconciliation data to PDF"""
    import io
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch, cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    
    # Build query
    query = {}
    if status_filter == "kritis":
        query["status"] = "Kritis"
    elif status_filter == "umpire":
        query["umpire_status"] = {"$nin": [None, "none", ""]}
    
    # Get data
    data = await db.coa_reconciliation.find(query, {"_id": 0}).sort("completed_unloading", -1).to_list(10000)
    
    # Get KPIs
    all_coa = await db.coa_reconciliation.find({}, {"_id": 0}).to_list(10000)
    kritis_count = sum(1 for c in all_coa if c.get("status") == "Kritis")
    umpire_count = sum(1 for c in all_coa if c.get("umpire_status") not in [None, "none", ""])
    
    settings = await db.settings.find_one({"type": "coa"})
    price_per_kcal = settings.get("price_per_kcal_per_ton", 50) if settings else 50
    potential_loss = 0
    for c in all_coa:
        delta = c.get("delta_loading_internal", 0) or 0
        tonase = c.get("ds_mt", 0) or 0
        if delta > 0:
            potential_loss += delta * tonase * price_per_kcal
    
    # Create PDF
    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=landscape(A4), leftMargin=0.5*inch, rightMargin=0.5*inch)
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=16, alignment=TA_CENTER, spaceAfter=20)
    elements.append(Paragraph("LAPORAN COA RECONCILIATION - PLTU TENAYAN", title_style))
    elements.append(Paragraph(f"Tanggal: {datetime.now().strftime('%d %B %Y')}", ParagraphStyle('SubTitle', parent=styles['Normal'], alignment=TA_CENTER, spaceAfter=20)))
    
    # KPI Summary
    kpi_data = [
        ["Total Records", "Status Kritis", "Proses Umpire", "Potential Loss"],
        [str(len(all_coa)), str(kritis_count), str(umpire_count), f"Rp {potential_loss:,.0f}"]
    ]
    kpi_table = Table(kpi_data, colWidths=[2.5*inch, 2.5*inch, 2.5*inch, 2.5*inch])
    kpi_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E3A5F')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F0F4F8')),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 11),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#CCCCCC')),
    ]))
    elements.append(kpi_table)
    elements.append(Spacer(1, 20))
    
    # Main Data Table
    headers = ["No", "Shipment", "Supplier", "Tanggal", "Loading GCV", "Unloading GCV", "Internal GCV", "Delta", "Status"]
    table_data = [headers]
    
    for idx, record in enumerate(data[:100], 1):  # Limit to 100 rows for PDF
        status = record.get("status", "normal")
        row = [
            str(idx),
            str(record.get("shipment", ""))[:15],
            str(record.get("suppliers", ""))[:20],
            str(record.get("completed_unloading", ""))[:10] if record.get("completed_unloading") else "-",
            str(record.get("loading_gcv_arb") or "-"),
            str(record.get("unloading_gcv_arb") or "-"),
            str(record.get("internal_gcv_arb") or "-"),
            str(int(record.get("delta_loading_internal") or 0)),
            status.upper()
        ]
        table_data.append(row)
    
    main_table = Table(table_data, colWidths=[0.4*inch, 1.2*inch, 1.8*inch, 0.9*inch, 1*inch, 1*inch, 1*inch, 0.8*inch, 0.8*inch])
    
    table_style = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E3A5F')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F5F5')]),
    ]
    
    # Highlight Kritis rows
    for idx, record in enumerate(data[:100], 1):
        if record.get("status") == "Kritis":
            table_style.append(('BACKGROUND', (0, idx), (-1, idx), colors.HexColor('#FFCCCC')))
    
    main_table.setStyle(TableStyle(table_style))
    elements.append(main_table)
    
    # Footer
    elements.append(Spacer(1, 20))
    footer_style = ParagraphStyle('Footer', parent=styles['Normal'], fontSize=8, alignment=TA_LEFT)
    elements.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Filter: {status_filter.upper()}", footer_style))
    
    doc.build(elements)
    output.seek(0)
    
    filename = f"COA_Reconciliation_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    
    return Response(
        content=output.getvalue(),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# ==================== HEALTH CHECK ====================

@api_router.get("/")
async def root():
    return {"message": "PLTU Tenayan Fuel Management System API", "status": "running"}

@api_router.get("/health")
async def health_check():
    return {"status": "healthy", "timestamp": datetime.now(timezone.utc).isoformat()}

# Include the router in the main app
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
