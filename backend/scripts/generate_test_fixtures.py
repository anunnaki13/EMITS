"""ONE-TIME SCRIPT: Generate 4 sanitized synthetic xlsx fixtures for Phase 4 tests.

To regenerate fixtures:
    cd pltu-tenayan-full-backup/backend
    .venv/bin/python scripts/generate_test_fixtures.py
    git add tests/fixtures/excel/*.xlsx
    git commit -m "chore: regenerate synthetic test fixtures"

Headers are the plain (non-newline) variants that the server.py parsers accept via
their fallback row.get("GCV (Kcal/Kg) ARB") calls. See RESEARCH §Focus 6 for details.

All rows use:
  shipment_code = SHP-TEST-001..005  (round-trip assertion target in Plan 04-03)
  gcv_arb       = 4200.0 + i*50      (deterministic value for round-trip assertion)
  supplier      = PT TEST SUPPLIER    (no real PT names)
  dates within Jan-2026              (within 6-month window for /api/dashboard/advanced)
"""
from pathlib import Path
from openpyxl import Workbook

FIXTURE_DIR = Path(__file__).parent.parent / "tests" / "fixtures" / "excel"
FIXTURE_DIR.mkdir(parents=True, exist_ok=True)


def write_vessel():
    wb = Workbook()
    ws = wb.active
    headers = [
        "Periode TA (Rakor)", "Periode Realisasi", "Shipment Code", "Voyage Code",
        "Suppliers", "Voyage", "Name Of Vessel", "Coal From",
        "Time Arrival", "Berthed Time", "Commenced Unloading", "Completed Unloading",
        "Durasi Pembongkaran (Hari)", "Durasi Pembongkaran (Jam)", "waktu tunggu (Jam)",
        "B/L (MT)", "DS (MT)", "NO.COW", "Tgl Terbit COW",
        "GCV (Kcal/Kg) ARB", "TM (%wt) ARB", "Ash Content (%wt) ARB",
        "Total Sulphur (%wt) ARB", "NO. COA", "Tgl Terbit COA", "DURASI TERBIT COA",
    ]
    ws.append(headers)
    for i in range(1, 6):
        ws.append([
            "Jan-2026", "Jan-2026", f"SHP-TEST-{i:03d}", f"VYG-{i:03d}",
            "PT TEST SUPPLIER", f"V{i:03d}", f"MV TEST VESSEL {i}", "Kalimantan",
            "2026-01-15 08:00", "2026-01-15 10:00", "2026-01-15 11:00", "2026-01-16 10:00",
            1.0, 23.0, 2.5,
            float(5000 + i * 100), float(4950 + i * 100),
            f"COW-{i:03d}", "2026-01-17",
            float(4200 + i * 50), 30.5, 5.0, 0.3,
            f"COA-{i:03d}", "2026-01-18", "3 days",
        ])
    wb.save(FIXTURE_DIR / "vessel_minimal.xlsx")
    print(f"Written: {FIXTURE_DIR / 'vessel_minimal.xlsx'}")


def write_barge():
    wb = Workbook()
    ws = wb.active
    headers = [
        "Periode", "Shipment Code", "Voyage Code", "Shipment", "Suppliers",
        "Voyage", "TB", "BG", "Coal From", "TA", "Berthed Time",
        "Commenced Unloading", "Completed Unloading",
        "Durasi Pembongkaran (Hari)", "Durasi Pembongkaran (Jam)",
        "waktu tunggu (Jam)", "B/L (MT)", "DS (MT)", "NO.COW",
        "Tgl Terbit COW", "GCV (Kcal/Kg) ARB", "TM (%wt) ARB",
        "Ash Content (%wt) ARB", "Total Sulphur (%wt) ARB",
        "NO. COA", "Tgl Terbit COA", "DURASI TERBIT COA",
    ]
    ws.append(headers)
    for i in range(1, 6):
        ws.append([
            "Jan-2026", f"SHP-TEST-{i:03d}", f"VYG-{i:03d}", f"SHP-TEST-{i:03d}",
            "PT TEST SUPPLIER", f"V{i:03d}", f"TB-TEST-{i:03d}", f"BG-{i:03d}",
            "Kalimantan", "2026-01-15 08:00", "2026-01-15 10:00",
            "2026-01-15 11:00", "2026-01-16 10:00",
            1.0, 23.0, 2.5,
            float(3000 + i * 100), float(2950 + i * 100),
            f"COW-{i:03d}", "2026-01-17",
            float(4200 + i * 50), 30.5, 5.0, 0.3,
            f"COA-{i:03d}", "2026-01-18", "3 days",
        ])
    wb.save(FIXTURE_DIR / "barge_minimal.xlsx")
    print(f"Written: {FIXTURE_DIR / 'barge_minimal.xlsx'}")


def write_trucking():
    wb = Workbook()
    ws = wb.active
    headers = [
        "Periode TA (Rakor)", "Periode Realisasi", "Shipment Code", "Voyage Code",
        "Shipment", "Suppliers", "Transportasi", "Coal From", "TA",
        "Berthed Time", "Commenced Unloading", "Completed Unloading",
        "Durasi Pembongkaran (Hari)", "Durasi Pembongkaran (Jam)",
        "B/L (MT)", "DS (MT)", "RIT", "NO.COW", "Tgl Terbit COW",
        "GCV (Kcal/Kg) ARB", "TM (%wt) ARB", "NO. COA", "Tgl Terbit COA",
    ]
    ws.append(headers)
    for i in range(1, 6):
        ws.append([
            "Jan-2026", "Jan-2026", f"SHP-TEST-{i:03d}", f"VYG-{i:03d}",
            f"SHP-TEST-{i:03d}", "PT TEST SUPPLIER", f"TRUCK-TEST-{i:03d}",
            "Kalimantan", "2026-01-15 08:00",
            "2026-01-15 10:00", "2026-01-15 11:00", "2026-01-16 10:00",
            0.0, 2.0,
            float(500 + i * 50), float(490 + i * 50),
            float(10 + i),
            f"COW-{i:03d}", "2026-01-17",
            float(4200 + i * 50), 30.5,
            f"COA-{i:03d}", "2026-01-18",
        ])
    wb.save(FIXTURE_DIR / "trucking_minimal.xlsx")
    print(f"Written: {FIXTURE_DIR / 'trucking_minimal.xlsx'}")


def write_biomassa():
    wb = Workbook()
    ws = wb.active
    # Biomassa parser normalizes headers via df.columns.str.replace('\n', ' ').str.strip()
    # Use plain single-line headers.
    headers = [
        "Periode", "Shipment Code", "Voyage Code", "Lot", "Suppliers",
        "Shipper", "Lot.1", "TB", "BG", "Biomass",
        "TA", "Berthed Time", "Commenced Unloading", "Completed Unloading",
        "Durasi Pembongkaran (Hari)", "B/L (MT)", "Jembatan Timbang (MT)",
        "Surveyor Unloading", "NO.COW / ROW", "Tgl Terbit COW",
        "GCV (Kcal/Kg) ARB", "GCV (Kcal/Kg) ADB", "TM (%wt) ARB",
        "IM (%wt) ADB", "NO. COA", "Tgl Terbit COA",
    ]
    ws.append(headers)
    for i in range(1, 6):
        ws.append([
            "Jan-2026", f"SHP-TEST-{i:03d}", f"VYG-{i:03d}", f"LOT-{i:03d}",
            "PT TEST BIOMASSA SUPPLIER", "PT SHIPPER", f"LOT-{i:03d}",
            f"TB-TEST-{i:03d}", f"BG-{i:03d}", "Biomassa",
            "2026-01-15 08:00", "2026-01-15 10:00",
            "2026-01-15 11:00", "2026-01-16 10:00",
            1.0,
            float(1000 + i * 100), float(980 + i * 100),
            "PT SURVEYOR", f"COW-{i:03d}", "2026-01-17",
            float(3000 + i * 50), float(3100 + i * 50), 45.0,
            12.0,
            f"COA-{i:03d}", "2026-01-18",
        ])
    wb.save(FIXTURE_DIR / "biomassa_minimal.xlsx")
    print(f"Written: {FIXTURE_DIR / 'biomassa_minimal.xlsx'}")


if __name__ == "__main__":
    write_vessel()
    write_barge()
    write_trucking()
    write_biomassa()
    print("Done. 4 fixtures written.")
