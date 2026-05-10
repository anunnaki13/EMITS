"""Phase 4 fixture validation tests.

Verifies that the 4 synthetic xlsx files exist, can be opened by openpyxl,
contain the expected headers, and have exactly 5 data rows with deterministic
shipment codes SHP-TEST-001 through SHP-TEST-005.

These tests run standalone (no backend needed).
"""
from pathlib import Path
import pytest

FIXTURE_DIR = Path(__file__).parent


def _load_ws(filename: str):
    import openpyxl
    path = FIXTURE_DIR / filename
    assert path.exists(), f"Fixture missing: {path}"
    return openpyxl.load_workbook(path).active


def test_vessel_xlsx_parseable():
    """Vessel fixture opens cleanly and has expected critical headers."""
    ws = _load_ws("vessel_minimal.xlsx")
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    assert "Periode TA (Rakor)" in headers, f"Missing 'Periode TA (Rakor)' in headers: {headers}"
    assert "GCV (Kcal/Kg) ARB" in headers, f"Missing 'GCV (Kcal/Kg) ARB' in headers: {headers}"


def test_barge_xlsx_parseable():
    """Barge fixture opens cleanly and has expected critical headers."""
    ws = _load_ws("barge_minimal.xlsx")
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    assert "TB" in headers, f"Missing 'TB' in headers: {headers}"
    assert "BG" in headers, f"Missing 'BG' in headers: {headers}"
    assert "GCV (Kcal/Kg) ARB" in headers, f"Missing 'GCV (Kcal/Kg) ARB' in headers: {headers}"


def test_trucking_xlsx_parseable():
    """Trucking fixture opens cleanly and has expected critical headers."""
    ws = _load_ws("trucking_minimal.xlsx")
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    assert "Transportasi" in headers, f"Missing 'Transportasi' in headers: {headers}"
    assert "RIT" in headers, f"Missing 'RIT' in headers: {headers}"


def test_biomassa_xlsx_parseable():
    """Biomassa fixture opens cleanly and has expected critical headers."""
    ws = _load_ws("biomassa_minimal.xlsx")
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    assert "Biomass" in headers, f"Missing 'Biomass' in headers: {headers}"
    assert "GCV (Kcal/Kg) ADB" in headers, f"Missing 'GCV (Kcal/Kg) ADB' in headers: {headers}"


def test_all_have_5_rows():
    """Each fixture has exactly 5 data rows with shipment codes SHP-TEST-001..005."""
    fixtures = [
        "vessel_minimal.xlsx",
        "barge_minimal.xlsx",
        "trucking_minimal.xlsx",
        "biomassa_minimal.xlsx",
    ]
    for filename in fixtures:
        ws = _load_ws(filename)
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

        # Find shipment code column index
        try:
            sc_idx = headers.index("Shipment Code")
        except ValueError:
            pytest.fail(f"{filename}: 'Shipment Code' column not found. Headers: {headers}")

        # Collect data rows (rows 2-6)
        data_rows = list(ws.iter_rows(min_row=2, values_only=True))
        assert len(data_rows) == 5, (
            f"{filename}: expected 5 data rows, got {len(data_rows)}"
        )

        # Check shipment codes
        expected_codes = {f"SHP-TEST-{i:03d}" for i in range(1, 6)}
        actual_codes = {row[sc_idx] for row in data_rows if row[sc_idx] is not None}
        assert actual_codes == expected_codes, (
            f"{filename}: shipment codes {actual_codes} != expected {expected_codes}"
        )
